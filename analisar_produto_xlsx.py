#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Análise do arquivo produto.xlsx para verificar imagens
"""

import openpyxl
import os
import zipfile

def analisar_produto_xlsx():
    """Analisa o arquivo produto.xlsx para verificar imagens"""
    
    print("🔍 Analisando arquivo produto.xlsx")
    print("=" * 40)
    
    arquivo = "produto.xlsx"
    
    if not os.path.exists(arquivo):
        print(f"❌ Arquivo {arquivo} não encontrado")
        return
    
    try:
        # Análise 1: Estrutura do arquivo
        print(f"📁 Informações do arquivo:")
        file_size = os.path.getsize(arquivo)
        print(f"   Tamanho: {file_size} bytes ({file_size/1024/1024:.2f} MB)")
        
        # Análise 2: Estrutura interna do Excel
        print(f"\n📊 Estrutura interna do Excel:")
        with zipfile.ZipFile(arquivo, 'r') as zip_file:
            file_list = zip_file.namelist()
            print(f"   Total de arquivos internos: {len(file_list)}")
            
            # Procura por arquivos de imagem
            image_files = [f for f in file_list if any(f.endswith(ext) for ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp'])]
            print(f"   Arquivos de imagem encontrados: {len(image_files)}")
            
            if image_files:
                print(f"   📷 Arquivos de imagem:")
                for img_file in image_files:
                    # Verifica tamanho da imagem
                    img_data = zip_file.read(img_file)
                    print(f"      {img_file} - {len(img_data)} bytes")
                    
                    # Verifica cabeçalho
                    if len(img_data) >= 8:
                        header = img_data[:8]
                        print(f"         Cabeçalho: {header.hex()}")
                        
                        if header.startswith(b'\xff\xd8'):
                            print(f"         ✅ Formato JPEG")
                        elif header.startswith(b'\x89PNG'):
                            print(f"         ✅ Formato PNG")
                        else:
                            print(f"         ❓ Formato desconhecido")
            
            # Procura por arquivos de desenho
            drawing_files = [f for f in file_list if 'drawing' in f.lower()]
            print(f"   Arquivos de desenho encontrados: {len(drawing_files)}")
            
            if drawing_files:
                print(f"   🎨 Arquivos de desenho:")
                for draw_file in drawing_files:
                    print(f"      {draw_file}")
        
        # Análise 3: Conteúdo das planilhas
        print(f"\n📋 Conteúdo das planilhas:")
        workbook = openpyxl.load_workbook(arquivo, data_only=True, read_only=False)
        
        print(f"   Planilhas disponíveis: {workbook.sheetnames}")
        
        total_imagens_todas_planilhas = 0
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            print(f"\n   📊 Planilha: {sheet_name}")
            print(f"      Dimensões: {worksheet.max_row} linhas x {worksheet.max_column} colunas")
            
            # Verifica imagens nesta planilha
            total_imagens = len(worksheet._images)
            total_imagens_todas_planilhas += total_imagens
            print(f"      Imagens embebidas: {total_imagens}")
            
            if total_imagens > 0:
                print(f"      🖼️ Imagens embebidas encontradas:")
                for i, image in enumerate(worksheet._images):
                    print(f"         Imagem {i+1}: {image.width}x{image.height}")
                    
                    # Verifica posição
                    try:
                        anchor = image.anchor
                        if hasattr(anchor, '_from'):
                            col_from = anchor._from.col
                            row_from = anchor._from.row
                        else:
                            col_from = anchor.col
                            row_from = anchor.row
                        
                        col_letter = openpyxl.utils.get_column_letter(col_from + 1)
                        row_number = row_from + 1
                        
                        print(f"            Posição: {col_letter}{row_number}")
                        
                        if col_from == 7:  # Coluna H
                            print(f"            ✅ Está na coluna H")
                        else:
                            print(f"            ⚠️ Está na coluna {col_letter}")
                    
                    except Exception as e:
                        print(f"            ❌ Erro ao verificar posição: {e}")
            
            # Verifica conteúdo das primeiras linhas
            print(f"      📝 Primeiras linhas:")
            for row in range(1, min(6, worksheet.max_row + 1)):
                row_data = []
                for col in range(1, min(6, worksheet.max_column + 1)):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        row_data.append(str(cell_value)[:20])  # Limita tamanho
                    else:
                        row_data.append("")
                print(f"         Linha {row}: {row_data}")
        
        workbook.close()
        
        # Resumo geral
        print(f"\n📊 RESUMO GERAL:")
        print(f"   • Total de imagens embebidas: {total_imagens_todas_planilhas}")
        print(f"   • Arquivos de imagem internos: {len(image_files)}")
        
        if total_imagens_todas_planilhas > 0:
            print(f"\n✅ Arquivo tem imagens embebidas nas células!")
            print(f"   Pode ser usado diretamente para exportação")
        elif len(image_files) > 0:
            print(f"\n⚠️ Arquivo tem imagens internas mas não embebidas")
            print(f"   Seria necessário inserir as imagens nas células")
        else:
            print(f"\n❌ Arquivo não tem imagens")
            print(f"   Não pode ser usado para exportação")
        
    except Exception as e:
        print(f"❌ Erro na análise: {e}")
        import traceback
        traceback.print_exc()

def testar_extração_imagens():
    """Testa extração de imagens do arquivo produto.xlsx"""
    
    print(f"\n🧪 Testando extração de imagens:")
    
    try:
        workbook = openpyxl.load_workbook("produto.xlsx", data_only=True, read_only=False)
        
        total_imagens_extraidas = 0
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            total_imagens = len(worksheet._images)
            
            if total_imagens > 0:
                print(f"📊 Planilha {sheet_name}: {total_imagens} imagens embebidas")
                
                for i, image in enumerate(worksheet._images):
                    try:
                        # Método _data()
                        image_bytes = image._data()
                        print(f"   ✅ Imagem {i+1}: {len(image_bytes)} bytes")
                        
                        # Verifica cabeçalho
                        if len(image_bytes) >= 8:
                            header = image_bytes[:8]
                            print(f"      Cabeçalho: {header.hex()}")
                            
                            if header.startswith(b'\xff\xd8'):
                                print(f"      ✅ Formato JPEG")
                            elif header.startswith(b'\x89PNG'):
                                print(f"      ✅ Formato PNG")
                            else:
                                print(f"      ❓ Formato desconhecido")
                        
                        total_imagens_extraidas += 1
                        
                    except Exception as e:
                        print(f"   ❌ Erro na imagem {i+1}: {e}")
            else:
                print(f"📊 Planilha {sheet_name}: 0 imagens embebidas")
        
        workbook.close()
        
        print(f"\n📊 Total de imagens extraídas com sucesso: {total_imagens_extraidas}")
        
    except Exception as e:
        print(f"❌ Erro: {e}")

def verificar_refs_e_imagens():
    """Verifica REFs e correspondência com imagens"""
    
    print(f"\n🔍 Verificando REFs e correspondência com imagens:")
    
    try:
        workbook = openpyxl.load_workbook("produto.xlsx", data_only=True, read_only=False)
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            print(f"\n📊 Planilha: {sheet_name}")
            
            # Verifica REFs na coluna A
            refs_encontradas = []
            for row in range(1, min(worksheet.max_row + 1, 100)):  # Limita para performance
                cell_value = worksheet[f'A{row}'].value
                if cell_value and str(cell_value).strip():
                    ref_value = str(cell_value).strip()
                    if ref_value.upper() not in ['TOTAL', 'SUBTOTAL', 'SUM', 'COUNT', 'REF']:
                        refs_encontradas.append((row, ref_value))
            
            print(f"   📋 REFs encontradas: {len(refs_encontradas)}")
            
            # Mostra algumas REFs
            for i, (row, ref) in enumerate(refs_encontradas[:10]):
                print(f"      Linha {row}: {ref}")
            
            if len(refs_encontradas) > 10:
                print(f"      ... e mais {len(refs_encontradas) - 10} REFs")
            
            # Verifica correspondência REF ↔ Imagem
            if len(worksheet._images) > 0:
                imagens_com_refs = 0
                for row, ref in refs_encontradas[:20]:  # Limita para performance
                    tem_imagem = False
                    for image in worksheet._images:
                        try:
                            anchor = image.anchor
                            if hasattr(anchor, '_from'):
                                col_from = anchor._from.col
                                row_from = anchor._from.row
                            else:
                                col_from = anchor.col
                                row_from = anchor.row
                            
                            # Verifica se está na coluna H (7) e linha correspondente
                            if col_from == 7 and (row_from + 1) == row:
                                tem_imagem = True
                                break
                        
                        except:
                            continue
                    
                    if tem_imagem:
                        imagens_com_refs += 1
                        print(f"   ✅ REF {ref} (linha {row}) tem imagem na coluna H")
                
                print(f"   📊 REFs com imagens: {imagens_com_refs}")
            else:
                print(f"   ⚠️ Nenhuma imagem embebida para verificar correspondência")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Erro: {e}")

def verificar_estrutura_completa():
    """Verifica a estrutura completa da planilha"""
    
    print(f"\n📋 Verificando estrutura completa da planilha:")
    
    try:
        workbook = openpyxl.load_workbook("produto.xlsx", data_only=True, read_only=False)
        worksheet = workbook.active
        
        # Verifica todas as colunas
        print(f"📊 Todas as colunas:")
        for col in range(1, min(worksheet.max_column + 1, 15)):  # Primeiras 15 colunas
            col_letter = openpyxl.utils.get_column_letter(col)
            cell_value = worksheet[f'{col_letter}1'].value
            if cell_value:
                print(f"   Coluna {col_letter}: {cell_value}")
        
        # Verifica se há coluna PHOTO
        print(f"\n🔍 Procurando coluna PHOTO:")
        for col in range(1, worksheet.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            cell_value = worksheet[f'{col_letter}1'].value
            if cell_value and 'photo' in str(cell_value).lower():
                print(f"   ✅ Coluna PHOTO encontrada: {col_letter} - {cell_value}")
                
                # Verifica se há conteúdo nesta coluna
                print(f"   📊 Conteúdo da coluna {col_letter}:")
                for row in range(2, min(worksheet.max_row + 1, 10)):
                    cell_value = worksheet[f'{col_letter}{row}'].value
                    if cell_value:
                        print(f"      Linha {row}: {cell_value}")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Erro: {e}")

if __name__ == "__main__":
    analisar_produto_xlsx()
    testar_extração_imagens()
    verificar_refs_e_imagens()
    verificar_estrutura_completa()
    
    print(f"\n💡 CONCLUSÕES:")
    print(f"1. Se há imagens embebidas, pode usar diretamente para exportação")
    print(f"2. Se há imagens internas mas não embebidas, seria necessário inserir")
    print(f"3. Se não há imagens, arquivo não pode ser usado para exportação")

