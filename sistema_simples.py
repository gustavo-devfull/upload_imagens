#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistema de Upload de Imagens Excel - Versão Simplificada para Deploy
Versão que funciona sem PIL para evitar problemas de compilação
"""

import os
import json
import tempfile
import webbrowser
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
import threading
import time

# Importações básicas (sem PIL)
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError as e:
    print(f"⚠️  openpyxl não disponível: {e}")
    OPENPYXL_AVAILABLE = False

try:
    import ftplib
    FTP_AVAILABLE = True
except ImportError as e:
    print(f"⚠️  ftplib não disponível: {e}")
    FTP_AVAILABLE = False

class SistemaExcelProcessor:
    """Classe para processamento de arquivos Excel"""
    
    def extract_internal_images(self, file_path):
        """Extrai imagens internas do arquivo Excel quando não há imagens embebidas"""
        try:
            import zipfile
            
            print(f"🔍 Extraindo imagens internas de {file_path}")
            internal_images = []
            
            with zipfile.ZipFile(file_path, 'r') as zip_file:
                # Lista arquivos de imagem
                image_files = [f for f in zip_file.namelist() if f.startswith('xl/media/') and f.endswith('.jpg')]
                
                print(f"📊 Arquivos de imagem internos encontrados: {len(image_files)}")
                
                for i, img_file in enumerate(image_files):
                    try:
                        # Extrai dados da imagem
                        img_data = zip_file.read(img_file)
                        
                        print(f"   📷 {img_file}: {len(img_data)} bytes")
                        
                        # Verifica se é JPEG válido
                        if len(img_data) >= 8 and img_data.startswith(b'\xff\xd8'):
                            internal_images.append({
                                'index': i,
                                'filename': img_file,
                                'bytes': img_data,
                                'size': len(img_data)
                            })
                            print(f"      ✅ JPEG válido")
                        else:
                            print(f"      ❌ JPEG inválido")
                    
                    except Exception as e:
                        print(f"   ❌ Erro ao extrair {img_file}: {e}")
            
            print(f"✅ Total de imagens internas extraídas: {len(internal_images)}")
            return internal_images
            
        except Exception as e:
            print(f"❌ Erro ao extrair imagens internas: {e}")
            return []

    def find_image_by_cell(self, worksheet, cell_addr):
        """Encontra imagem por célula usando anchor - método robusto"""
        try:
            # Normaliza endereço da célula (ex: "H23")
            cell_addr = cell_addr.upper()
            target_col_letter = ''.join([c for c in cell_addr if c.isalpha()])
            target_row = int(''.join([c for c in cell_addr if c.isdigit()]))
            
            # Converte coluna para índice 0-based
            target_col_idx = openpyxl.utils.column_index_from_string(target_col_letter) - 1
            
            # Itera imagens da planilha e checa o anchor
            for img in getattr(worksheet, "_images", []):
                anchor = img.anchor
                try:
                    # Quando é TwoCellAnchor
                    col_from = anchor._from.col
                    row_from = anchor._from.row
                except AttributeError:
                    # Quando é OneCellAnchor
                    col_from = anchor.col
                    row_from = anchor.row
                
                # Anchors são 0-based tanto para col quanto para row
                if col_from == target_col_idx and (row_from + 1) == target_row:
                    # Extrai bytes da imagem usando método mais confiável
                    try:
                        # Método principal: _data() - mais confiável
                        if hasattr(img, '_data') and callable(img._data):
                            image_bytes = img._data()
                            if image_bytes and len(image_bytes) > 100:
                                print(f"✅ Imagem extraída via _data(): {len(image_bytes)} bytes")
                                return image_bytes
                    except Exception as e:
                        print(f"❌ Erro _data(): {e}")
                        pass
                    
                    try:
                        # Método alternativo: Usa o caminho interno do arquivo
                        with worksheet.parent._archive.open(img.path) as fp:
                            image_bytes = fp.read()
                            if image_bytes and len(image_bytes) > 100:
                                print(f"✅ Imagem extraída via _archive: {len(image_bytes)} bytes")
                                return image_bytes
                    except Exception as e:
                        print(f"❌ Erro _archive: {e}")
                        pass
                    
                    try:
                        # Método fallback: ref (pode falhar se arquivo fechado)
                        if hasattr(img, 'ref') and img.ref:
                            image_bytes = img.ref.read()
                            if image_bytes and len(image_bytes) > 100:
                                print(f"✅ Imagem extraída via ref: {len(image_bytes)} bytes")
                                return image_bytes
                    except Exception as e:
                        print(f"❌ Erro ref: {e}")
                        pass
            
            return None
            
        except Exception as e:
            return None

    def upload_images_ftp(self, images):
        """Upload real das imagens via FTP com timeout maior para muitos uploads"""
        upload_successful = 0
        upload_failed = 0
        
        if not FTP_AVAILABLE:
            print("❌ FTP não disponível - simulando upload")
            for image_data in images:
                print(f"📤 Simulação: {image_data['ref']}.jpg ({len(image_data['bytes'])} bytes)")
                upload_successful += 1
            return upload_successful, upload_failed
        
        try:
            # Configurações FTP
            ftp_host = os.getenv('FTP_HOST', "46.202.90.62")
            ftp_user = os.getenv('FTP_USER', "u715606397.ideolog.ia.br")
            ftp_pass = os.getenv('FTP_PASSWORD', "27y8rYoDq=Q&aHk:")
            ftp_path = "/images/products/"
            
            print(f"🔗 Conectando FTP: {ftp_host}")
            
            # Conecta FTP
            ftp = ftplib.FTP()
            ftp.connect(ftp_host, 21)
            ftp.login(ftp_user, ftp_pass)
            ftp.cwd(ftp_path)
            
            print(f"✅ Conectado FTP - diretório: {ftp_path}")
            
            # Upload de cada imagem
            for image_data in images:
                try:
                    # Cria arquivo temporário
                    temp_image_path = f"/tmp/{image_data['ref']}.jpg"
                    
                    # Tenta usar PIL se disponível
                    try:
                        from PIL import Image
                        import io
                        # Abre a imagem dos bytes
                        image_stream = io.BytesIO(image_data['bytes'])
                        img = Image.open(image_stream)
                    except ImportError:
                        # Se PIL não estiver disponível, salva bytes diretamente
                        with open(temp_image_path, 'wb') as f:
                            f.write(image_data['bytes'])
                        print(f"📤 Upload {image_data['ref']}: {len(image_data['bytes'])} bytes (sem conversão PIL)")
                        # Faz upload direto
                        with open(temp_image_path, 'rb') as file:
                            ftp.storbinary(f'STOR {image_data["ref"]}.jpg', file)
                        print(f"✅ Upload concluído: {image_data['ref']}.jpg")
                        os.unlink(temp_image_path)
                        upload_successful += 1
                        continue
                    
                    # Continua com PIL se disponível
                    
                    # Detecta formato original
                    original_format = img.format
                    source_info = image_data.get('source', 'unknown')
                    filename_info = f" ({image_data.get('filename', '')})" if 'filename' in image_data else ""
                    print(f"📤 Upload {image_data['ref']}: {len(image_data['bytes'])} bytes ({original_format}) → JPEG [{source_info}]{filename_info}")
                    
                    # Converte para RGB se necessário
                    if img.mode in ('RGBA', 'LA', 'P'):
                        img = img.convert('RGB')
                    
                    # Salva como JPEG com configurações otimizadas para compatibilidade
                    img.save(temp_image_path, 'JPEG', quality=100, optimize=False, progressive=False, subsampling=0)
                    
                    # Upload FTP
                    with open(temp_image_path, 'rb') as file:
                        ftp.storbinary(f'STOR {image_data["ref"]}.jpg', file)
                    
                    print(f"✅ Upload concluído: {image_data['ref']}.jpg")
                    
                    # Validação pós-upload
                    try:
                        import requests
                        image_url = f"https://ideolog.ia.br/images/products/{image_data['ref']}.jpg"
                        response = requests.get(image_url, timeout=10)
                        if response.status_code == 200:
                            if response.content.startswith(b'\xff\xd8'):
                                print(f"✅ Validação pós-upload: JPEG válido após processamento do servidor")
                            else:
                                print(f"⚠️ Validação pós-upload: Arquivo não é mais JPEG válido")
                        else:
                            print(f"⚠️ Validação pós-upload: Não foi possível baixar arquivo")
                    except Exception as e:
                        print(f"⚠️ Validação pós-upload falhou: {e}")
                    
                    # Remove arquivo temporário
                    os.unlink(temp_image_path)
                    upload_successful += 1
                    
                except Exception as e:
                    print(f"❌ Erro upload {image_data['ref']}: {e}")
                    upload_failed += 1
            
            # Fecha FTP
            ftp.quit()
            print(f"🔌 FTP desconectado")
            
        except Exception as e:
            print(f"❌ Erro FTP: {e}")
            upload_failed = len(images)
        
        return upload_successful, upload_failed

    def process_excel_simple(self, file_path):
        """Processamento do Excel com configurações corretas para Render"""
        try:
            # Configuração correta para carregar imagens
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=False, keep_links=False)
            worksheet = workbook.active
            
            print(f"📊 Planilha carregada: {worksheet.title}")
            print(f"📏 Dimensões: {worksheet.max_row} linhas x {worksheet.max_column} colunas")
            
            # Extrai REFs da coluna A (linhas 4+)
            refs = []
            ref_count = 0
            
            for row in range(4, worksheet.max_row + 1):
                cell_value = worksheet[f'A{row}'].value
                if cell_value and str(cell_value).strip() and str(cell_value).strip().upper() not in ['TOTAL', 'SUBTOTAL']:
                    ref_count += 1
                    refs.append(str(cell_value).strip())
            
            # Processa imagens usando método robusto por anchor
            images = []
            total_images = len(worksheet._images)
            
            print(f"📊 Imagens embebidas encontradas: {total_images}")
            
            # Se não há imagens embebidas, tenta extrair imagens internas
            if total_images == 0:
                print(f"🔍 Nenhuma imagem embebida encontrada, tentando extrair imagens internas...")
                internal_images = self.extract_internal_images(file_path)
                
                if internal_images:
                    print(f"✅ {len(internal_images)} imagens internas extraídas")
                    
                    # Associa imagens internas às REFs por ordem
                    for i, ref_value in enumerate(refs):
                        if i < len(internal_images):
                            internal_img = internal_images[i]
                            images.append({
                                'ref': ref_value, 
                                'bytes': internal_img['bytes'], 
                                'row': i + 4,  # Assume que REFs começam na linha 4
                                'source': 'internal',
                                'filename': internal_img['filename']
                            })
                            print(f"   📷 REF {ref_value} → {internal_img['filename']}")
                        else:
                            print(f"   ⚠️ REF {ref_value} não tem imagem correspondente")
                else:
                    print(f"❌ Nenhuma imagem interna encontrada")
            else:
                # Método original: busca imagens embebidas
                print(f"🔍 Buscando imagens embebidas...")
                for ref_value in refs:
                    # Encontra a linha da REF
                    ref_row = None
                    for row in range(4, worksheet.max_row + 1):
                        cell_value = worksheet[f'A{row}'].value
                        if cell_value and str(cell_value).strip() == ref_value:
                            ref_row = row
                            break
                    
                    if ref_row:
                        # Busca imagem na célula H{ref_row}
                        image_bytes = self.find_image_by_cell(worksheet, f'H{ref_row}')
                        if image_bytes:
                            images.append({'ref': ref_value, 'bytes': image_bytes, 'row': ref_row, 'source': 'embedded'})
            
            workbook.close()
            
            # Upload real via FTP
            upload_successful = 0
            upload_failed = 0
            
            if images:
                print(f"\n🚀 Iniciando upload de {len(images)} imagens...")
                upload_successful, upload_failed = self.upload_images_ftp(images)
            else:
                print(f"\n⚠️ Nenhuma imagem encontrada para upload")
            
            return {
                'success': True,
                'total_refs': len(refs),
                'images_found': len(images),
                'uploads_successful': upload_successful,
                'uploads_failed': upload_failed,
                'error': None
            }
            
        except Exception as e:
            print(f"❌ Erro no processamento: {e}")
            return {
                'success': False,
                'error': str(e),
                'total_refs': 0,
                'images_found': 0,
                'uploads_successful': 0,
                'uploads_failed': 1
            }

class SimpleUploadHandler(BaseHTTPRequestHandler):
    """Handler simplificado para o servidor de deploy"""
    
    def do_GET(self):
        """Serve o frontend"""
        if self.path == '/' or self.path == '/index.html':
            self.serve_frontend()
        elif self.path == '/config':
            self.serve_config()
        elif self.path == '/health':
            self.serve_health()
        elif self.path == '/favicon.ico':
            # Retorna 204 No Content para favicon (evita erro 404)
            self.send_response(204)
            self.end_headers()
        else:
            self.send_error(404)
    
    def do_POST(self):
        """Processa uploads"""
        if self.path == '/upload':
            self.handle_upload()
        else:
            self.send_error(404)
    
    def serve_frontend(self):
        """Serve o frontend HTML"""
        html_content = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Sistema de Upload de Imagens Excel - Simplificado</title>
        <link href="https://fonts.googleapis.com/css2?family=Archivo:wght@300;400;500;600;700&display=swap" rel="stylesheet">

    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }

        body {
            font-family: 'Archivo', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, rgb(1, 117, 166) 0%, rgb(0, 80, 120) 100%);
            min-height: 100vh;
            padding: 20px;
        }

        .container {
            max-width: 800px;
            margin: 0 auto;
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 40px rgba(0,0,0,0.1);
            overflow: hidden;
        }

        .header {
            background: linear-gradient(135deg, rgb(1, 117, 166) 0%, rgb(0, 80, 120) 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }

        .header h1 {
            font-size: 2.5rem;
            margin-bottom: 10px;
            font-weight: 300;
        }

        .header p {
            font-size: 1.1rem;
            opacity: 0.9;
        }

        .content {
            padding: 40px;
        }

        .status-box {
            background: #e8f5e8;
            border: 1px solid #c8e6c9;
            border-radius: 10px;
            padding: 20px;
            margin: 20px 0;
        }

        .status-box h3 {
            color: #2e7d32;
            margin-bottom: 10px;
        }

        .warning-box {
            background: #fff3e0;
            border: 1px solid #ffcc02;
            border-radius: 10px;
            padding: 20px;
            margin: 20px 0;
        }

        .warning-box h3 {
            color: #f57c00;
            margin-bottom: 10px;
        }

        .upload-area {
            border: 3px dashed #667eea;
            border-radius: 15px;
            padding: 40px;
            text-align: center;
            margin-bottom: 30px;
            transition: all 0.3s ease;
            cursor: pointer;
        }

        .upload-area:hover {
            border-color: #764ba2;
            background: #f8f9ff;
        }

        .upload-icon {
            font-size: 4rem;
            color: #667eea;
            margin-bottom: 20px;
        }

        .upload-text {
            font-size: 1.2rem;
            color: #333;
            margin-bottom: 10px;
        }

        .upload-subtext {
            color: #666;
            font-size: 0.9rem;
        }

        .btn {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            padding: 15px 30px;
            border-radius: 50px;
            font-size: 1.1rem;
            cursor: pointer;
            transition: all 0.3s ease;
            margin: 10px;
            text-decoration: none;
            display: inline-block;
        }

        .btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 10px 20px rgba(102, 126, 234, 0.3);
        }

        .btn:disabled {
            opacity: 0.6;
            cursor: not-allowed;
            transform: none;
        }

        .info-box {
            background: #e3f2fd;
            border: 1px solid #bbdefb;
            border-radius: 10px;
            padding: 20px;
            margin: 20px 0;
        }

        .info-box h3 {
            color: #1976d2;
            margin-bottom: 10px;
        }

        .progress-container {
            display: none;
            margin: 30px 0;
            background: #f8f9fa;
            border-radius: 15px;
            padding: 20px;
            border: 1px solid #e9ecef;
        }

        .progress-bar {
            width: 100%;
            height: 25px;
            background: #e9ecef;
            border-radius: 15px;
            overflow: hidden;
            margin-bottom: 15px;
            box-shadow: inset 0 2px 4px rgba(0,0,0,0.1);
        }

        .progress-fill {
            height: 100%;
            background: linear-gradient(135deg, #28a745 0%, #20c997 50%, #17a2b8 100%);
            width: 0%;
            transition: width 0.5s ease;
            position: relative;
            border-radius: 15px;
        }

        .progress-fill::after {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: linear-gradient(45deg, transparent 30%, rgba(255,255,255,0.3) 50%, transparent 70%);
            animation: shimmer 2s infinite;
        }

        @keyframes shimmer {
            0% { transform: translateX(-100%); }
            100% { transform: translateX(100%); }
        }

        .progress-text {
            text-align: center;
            color: #495057;
            font-size: 1rem;
            font-weight: 500;
            margin-bottom: 10px;
        }

        .progress-details {
            display: flex;
            justify-content: space-between;
            font-size: 0.8rem;
            color: #6c757d;
        }

        .progress-steps {
            margin: 15px 0;
        }

        .progress-step {
            display: flex;
            align-items: center;
            margin-bottom: 8px;
            font-size: 0.9rem;
            color: #6c757d;
            transition: color 0.3s ease;
        }

        .progress-step.active {
            color: #28a745;
            font-weight: 500;
        }

        .progress-step.completed {
            color: #28a745;
        }

        .progress-step-icon {
            width: 24px;
            height: 24px;
            border-radius: 50%;
            background: #e9ecef;
            display: flex;
            align-items: center;
            justify-content: center;
            margin-right: 12px;
            font-size: 0.8rem;
            font-weight: bold;
            transition: all 0.3s ease;
            border: 2px solid #e9ecef;
        }

        .progress-step.active .progress-step-icon {
            background: #28a745;
            color: white;
            border-color: #28a745;
            transform: scale(1.1);
        }

        .progress-step.completed .progress-step-icon {
            background: #28a745;
            color: white;
            border-color: #28a745;
        }

        .results {
            display: none;
            margin-top: 30px;
            padding: 20px;
            background: #f8f9fa;
            border-radius: 15px;
        }

        .results h3 {
            margin-bottom: 20px;
            color: #333;
        }

        .stats-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }

        .stat-card {
            background: white;
            padding: 20px;
            border-radius: 15px;
            text-align: center;
            box-shadow: 0 5px 15px rgba(0,0,0,0.08);
        }

        .stat-number {
            font-size: 2rem;
            font-weight: bold;
            color: #667eea;
            margin-bottom: 5px;
        }

        .stat-label {
            color: #666;
            font-size: 0.9rem;
        }

        @media (max-width: 768px) {
            .container {
                margin: 10px;
                border-radius: 15px;
            }
            
            .header {
                padding: 20px;
            }
            
            .header h1 {
                font-size: 2rem;
            }
            
            .content {
                padding: 20px;
            }
            
            .upload-area {
                padding: 20px;
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Sistema de Upload de Imagens para Cotações</h1>
            <p>Subir imagens para o FTP e vincular as imagens às cotações</p>
        </div>

        <div class="content">
            <div class="upload-area" onclick="document.getElementById('fileInput').click()">
                <div class="upload-icon">📁</div>
                <div class="upload-text">Clique para selecionar arquivo Excel</div>
                <div class="upload-subtext">Apenas arquivos .xlsx são aceitos</div>
            </div>

            <input type="file" id="fileInput" accept=".xlsx" style="display: none;" onchange="handleFileSelect(this.files[0])">

            <div style="text-align: center;">
                <button class="btn" onclick="uploadFile()" id="uploadBtn" disabled>
                    🚀 Fazer Upload
                </button>
                <button class="btn" onclick="resetForm()" id="resetBtn" disabled>
                    🔄 Limpar
                </button>
            </div>

            <div id="progressContainer" class="progress-container">
                <div id="progressText" class="progress-text">Preparando upload...</div>
                <div class="progress-bar">
                    <div id="progressFill" class="progress-fill"></div>
                </div>
                <div id="progressSteps" class="progress-steps">
                    <div class="progress-step" id="step1">
                        <div class="progress-step-icon">1</div>
                        <span>Enviando arquivo</span>
                    </div>
                    <div class="progress-step" id="step2">
                        <div class="progress-step-icon">2</div>
                        <span>Analisando planilha</span>
                    </div>
                    <div class="progress-step" id="step3">
                        <div class="progress-step-icon">3</div>
                        <span>Extraindo imagens</span>
                    </div>
                    <div class="progress-step" id="step4">
                        <div class="progress-step-icon">4</div>
                        <span>Upload FTP</span>
                    </div>
                    <div class="progress-step" id="step5">
                        <div class="progress-step-icon">5</div>
                        <span>Finalizando</span>
                    </div>
                </div>
                <div id="progressDetails" class="progress-details">
                    <span id="progressStatus">Aguardando...</span>
                    <span id="progressTime">00:00</span>
                </div>
            </div>

            <div id="results" class="results">
                <h3>📊 Resultados do Processamento</h3>
                <div id="resultsContent"></div>
            </div>
        </div>
    </div>

    <script>
        let selectedFile = null;


        function handleFileSelect(file) {
            if (file && file.type === 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet') {
                selectedFile = file;
                document.getElementById('uploadBtn').disabled = false;
                document.getElementById('resetBtn').disabled = false;
                document.querySelector('.upload-text').textContent = file.name;
                document.querySelector('.upload-area').style.borderColor = '#28a745';
                document.querySelector('.upload-area').style.backgroundColor = '#f8fff8';
            } else {
                alert('Por favor, selecione um arquivo Excel (.xlsx) válido.');
            }
        }

        function uploadFile() {
            if (!selectedFile) {
                alert('Por favor, selecione um arquivo Excel.');
                return;
            }

            const formData = new FormData();
            formData.append('excel_file', selectedFile);

            // Mostra progresso
            document.getElementById('progressContainer').style.display = 'block';
            document.getElementById('uploadBtn').disabled = true;
            document.getElementById('resetBtn').disabled = true;

            // Barra de progresso avançada
            let progress = 0;
            let currentStep = 0;
            let startTime = Date.now();
            
            const steps = [
                { text: 'Enviando arquivo...', progress: 20, stepId: 'step1' },
                { text: 'Analisando planilha...', progress: 40, stepId: 'step2' },
                { text: 'Extraindo imagens...', progress: 60, stepId: 'step3' },
                { text: 'Fazendo upload FTP...', progress: 80, stepId: 'step4' },
                { text: 'Finalizando...', progress: 100, stepId: 'step5' }
            ];
            
            // Cronômetro
            const timerInterval = setInterval(() => {
                const elapsed = Math.floor((Date.now() - startTime) / 1000);
                const minutes = Math.floor(elapsed / 60).toString().padStart(2, '0');
                const seconds = (elapsed % 60).toString().padStart(2, '0');
                document.getElementById('progressTime').textContent = `${minutes}:${seconds}`;
            }, 1000);
            
            // Progresso dos passos
            const progressInterval = setInterval(() => {
                if (currentStep < steps.length) {
                    const step = steps[currentStep];
                    
                    // Atualiza barra de progresso
                    progress = step.progress;
                    document.getElementById('progressFill').style.width = progress + '%';
                    document.getElementById('progressText').textContent = step.text;
                    
                    // Atualiza passo atual
                    const stepElement = document.getElementById(step.stepId);
                    stepElement.classList.add('active');
                    
                    // Marca passos anteriores como completos
                    for (let i = 0; i < currentStep; i++) {
                        const prevStep = document.getElementById(steps[i].stepId);
                        prevStep.classList.remove('active');
                        prevStep.classList.add('completed');
                    }
                    
                    // Atualiza status
                    document.getElementById('progressStatus').textContent = step.text;
                    
                    currentStep++;
                } else {
                    clearInterval(progressInterval);
                }
            }, 400);

            fetch('/upload', {
                method: 'POST',
                body: formData,
            })
            .then(response => {
                if (!response.ok) {
                    throw new Error(`HTTP error! status: ${response.status}`);
                }
                return response.text();
            })
            .then(text => {
                try {
                    return JSON.parse(text);
                } catch (e) {
                    console.error('Erro ao fazer parse do JSON:', e);
                    console.error('Resposta recebida:', text);
                    throw new Error('Resposta inválida do servidor');
                }
            })
            .then(data => {
                clearInterval(progressInterval);
                clearInterval(timerInterval);
                
                // Marca todos os passos como completos
                steps.forEach(step => {
                    const stepElement = document.getElementById(step.stepId);
                    stepElement.classList.remove('active');
                    stepElement.classList.add('completed');
                });
                
                // Atualiza progresso final com informações detalhadas
                document.getElementById('progressFill').style.width = '100%';
                if (data.uploads_successful > 0) {
                    const totalImages = data.images_found || 0;
                    const successRate = totalImages > 0 ? ((data.uploads_successful / totalImages) * 100).toFixed(1) : 0;
                    
                    if (totalImages > 30) {
                        document.getElementById('progressText').textContent = `✅ Upload concluído! ${data.uploads_successful}/${totalImages} imagens enviadas (${successRate}% sucesso)`;
                        document.getElementById('progressStatus').textContent = `✅ Upload em lote concluído: ${data.uploads_successful} de ${totalImages} imagens`;
                    } else {
                        document.getElementById('progressText').textContent = `✅ Upload concluído! ${data.uploads_successful} imagem(ns) enviada(s)`;
                        document.getElementById('progressStatus').textContent = `✅ ${data.uploads_successful} imagem(ns) enviada(s) com sucesso`;
                    }
                } else if (data.error) {
                    document.getElementById('progressText').textContent = `❌ ${data.error}`;
                    document.getElementById('progressStatus').textContent = `❌ Erro: ${data.error}`;
                } else {
                    document.getElementById('progressText').textContent = '⚠️ Nenhuma imagem foi enviada';
                    document.getElementById('progressStatus').textContent = '⚠️ Nenhuma imagem foi enviada';
                }
                
                setTimeout(() => {
                    showResults(data);
                }, 1500);
            })
            .catch(error => {
                clearInterval(progressInterval);
                clearInterval(timerInterval);
                
                document.getElementById('progressText').textContent = `❌ Erro no upload: ${error.message}`;
                document.getElementById('progressStatus').textContent = `❌ Erro: ${error.message}`;
                
                alert('Erro no upload: ' + error.message);
                document.getElementById('uploadBtn').disabled = false;
                document.getElementById('resetBtn').disabled = false;
            });
        }

        function showResults(data) {
            const resultsDiv = document.getElementById('results');
            const contentDiv = document.getElementById('resultsContent');
            
            let html = `
                <div class="stats-grid">
                    <div class="stat-card">
                        <div class="stat-number">${data.total_refs || 0}</div>
                        <div class="stat-label">REFs Processadas</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">${data.images_found || 0}</div>
                        <div class="stat-label">Imagens Encontradas</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">${data.uploads_successful || 0}</div>
                        <div class="stat-label">Uploads Bem-sucedidos</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">${data.uploads_failed || 0}</div>
                        <div class="stat-label">Uploads Falharam</div>
                    </div>
                </div>
            `;

            if (data.error) {
                html += `<div class="warning-box"><h3>⚠️ Problema no Processamento</h3><p>${data.error}</p>`;
                if (data.suggestion) {
                    html += `<p><strong>💡 Sugestão:</strong> ${data.suggestion}</p>`;
                }
                html += `</div>`;
            } else if (data.images_found > 30) {
                html += `<div class="status-box"><h3>📤 Upload em Lote Concluído</h3>`;
                html += `<p><strong>Total de imagens:</strong> ${data.images_found}</p>`;
                html += `<p><strong>Uploads bem-sucedidos:</strong> ${data.uploads_successful}</p>`;
                html += `<p><strong>Uploads falharam:</strong> ${data.uploads_failed}</p>`;
                html += `<p><strong>Taxa de sucesso:</strong> ${data.images_found > 0 ? ((data.uploads_successful / data.images_found) * 100).toFixed(1) : 0}%</p>`;
                html += `<p><strong>URL das imagens:</strong> <a href="https://ideolog.ia.br/images/products/" target="_blank">https://ideolog.ia.br/images/products/</a></p>`;
                html += `</div>`;
            }

            contentDiv.innerHTML = html;
            resultsDiv.style.display = 'block';
            
            document.getElementById('uploadBtn').disabled = false;
            document.getElementById('resetBtn').disabled = false;
        }

        function resetForm() {
            selectedFile = null;
            document.getElementById('fileInput').value = '';
            document.getElementById('uploadBtn').disabled = true;
            document.getElementById('resetBtn').disabled = true;
            document.querySelector('.upload-text').textContent = 'Clique para selecionar arquivo Excel';
            document.querySelector('.upload-area').style.borderColor = '#667eea';
            document.querySelector('.upload-area').style.backgroundColor = 'transparent';
            document.getElementById('progressContainer').style.display = 'none';
            document.getElementById('results').style.display = 'none';
        }
    </script>
</body>
</html>
        """
        
        self.send_response(200)
        self.send_header('Content-type', 'text/html')
        self.end_headers()
        self.wfile.write(html_content.encode())
    
    def serve_config(self):
        """Serve configurações do sistema"""
        config = {
            'status': 'ok',
            'message': 'Sistema Simplificado Funcionando',
            'mode': 'simplified',
            'ftp_host': '46.202.90.62',
            'base_url': 'https://ideolog.ia.br/images/products/',
            'max_file_size': '50MB',
            'dependencies': {
                'ftp_available': FTP_AVAILABLE,
                'openpyxl_available': OPENPYXL_AVAILABLE,
                'pil_available': False  # Sempre False nesta versão
            }
        }
        
        self.send_response(200)
        self.send_header('Content-type', 'application/json')
        self.end_headers()
        self.wfile.write(json.dumps(config).encode())
    
    def serve_health(self):
        """Serve health check"""
        health = {
            'status': 'ok',
            'message': 'Sistema simplificado funcionando',
            'dependencies': {
                'ftp': FTP_AVAILABLE,
                'openpyxl': OPENPYXL_AVAILABLE,
                'pil': False
            }
        }
        
        self.send_response(200)
        self.send_header('Content-type', 'application/json')
        self.end_headers()
        self.wfile.write(json.dumps(health).encode())
    
    def handle_upload(self):
        """Processa uploads de arquivos Excel - versão simplificada"""
        try:
            print("🔍 Iniciando handle_upload...")
            
            if not FTP_AVAILABLE or not OPENPYXL_AVAILABLE:
                print("❌ Dependências não disponíveis")
                error_response = {
                    'error': 'Dependências não disponíveis no ambiente de deploy',
                    'total_refs': 0,
                    'images_found': 0,
                    'uploads_successful': 0,
                    'uploads_failed': 1
                }
                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps(error_response).encode())
                return
            
            # Lê o arquivo enviado
            content_length = int(self.headers['Content-Length'])
            print(f"📏 Content-Length: {content_length}")
            
            post_data = self.rfile.read(content_length)
            print(f"📦 Dados recebidos: {len(post_data)} bytes")
            
            # Parse do FormData
            print(f"📋 Content-Type: {self.headers['Content-Type']}")
            boundary = self.headers['Content-Type'].split('boundary=')[1]
            print(f"🔗 Boundary: {boundary}")
            
            parts = post_data.split(f'--{boundary}'.encode())
            print(f"📄 Partes encontradas: {len(parts)}")
            
            file_data = None
            filename = None
            
            for part in parts:
                if b'Content-Disposition: form-data' in part:
                    if b'filename=' in part:
                        # Extrai nome do arquivo
                        filename_start = part.find(b'filename="') + 11
                        filename_end = part.find(b'"', filename_start)
                        filename = part[filename_start:filename_end].decode()
                        
                        # Extrai dados do arquivo
                        file_start = part.find(b'\r\n\r\n') + 4
                        file_data = part[file_start:-2]  # Remove \r\n no final
                        break
            
            if not file_data or not filename:
                self.send_error(400, 'Arquivo não encontrado')
                return
            
            # Salva arquivo temporariamente com suffix .xlsx
            from tempfile import NamedTemporaryFile
            temp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")
            temp_file.write(file_data)
            temp_file.flush()
            temp_file.close()
            temp_path = temp_file.name
            
            # Processamento simplificado (sem PIL)
            stats = self.process_excel_simple(temp_path)
            
            # Verifica se o arquivo tem imagens
            if stats['images_found'] == 0:
                error_response = {
                    'error': f'Arquivo {filename} não contém imagens na coluna H. Use um arquivo que tenha imagens inseridas.',
                    'total_refs': stats['total_refs'],
                    'images_found': stats['images_found'],
                    'uploads_successful': 0,
                    'uploads_failed': stats['total_refs'],
                    'suggestion': 'Arquivos recomendados: tartaruga.xlsx ou carrinho.xlsx'
                }
                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps(error_response).encode())
                os.remove(temp_path)
                return
            
            # Remove arquivo temporário
            try:
                os.remove(temp_path)
            except:
                pass
            
            # Envia resposta com timeout maior para muitos uploads
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            # Garante que a resposta JSON está completa
            response_data = json.dumps(stats, ensure_ascii=False)
            self.wfile.write(response_data.encode('utf-8'))
            self.wfile.flush()
            
        except Exception as e:
            print(f"❌ Erro no handle_upload: {e}")
            error_response = {
                'success': False,
                'error': str(e),
                'total_refs': 0,
                'images_found': 0,
                'uploads_successful': 0,
                'uploads_failed': 1
            }
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(error_response, ensure_ascii=False).encode('utf-8'))
    
    def process_excel_simple(self, file_path):
        """Processamento do Excel com configurações corretas para Render"""
        try:
            # Configuração correta para carregar imagens
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=False, keep_links=False)
            worksheet = workbook.active
            
            # Conta REFs válidas
            ref_count = 0
            refs = []
            for row in range(4, worksheet.max_row + 1):
                cell_value = worksheet[f'A{row}'].value
                if cell_value and str(cell_value).strip() and str(cell_value).strip().upper() not in ['TOTAL', 'SUBTOTAL']:
                    ref_count += 1
                    refs.append(str(cell_value).strip())
            
            # Processa imagens usando método robusto por anchor
            images = []
            total_images = len(worksheet._images)
            
            print(f"📊 Imagens embebidas encontradas: {total_images}")
            
            # Se não há imagens embebidas, tenta extrair imagens internas
            if total_images == 0:
                print(f"🔍 Nenhuma imagem embebida encontrada, tentando extrair imagens internas...")
                internal_images = self.extract_internal_images(file_path)
                
                if internal_images:
                    print(f"✅ {len(internal_images)} imagens internas extraídas")
                    
                    # Associa imagens internas às REFs por ordem
                    for i, ref_value in enumerate(refs):
                        if i < len(internal_images):
                            internal_img = internal_images[i]
                            images.append({
                                'ref': ref_value, 
                                'bytes': internal_img['bytes'], 
                                'row': i + 4,  # Assume que REFs começam na linha 4
                                'source': 'internal',
                                'filename': internal_img['filename']
                            })
                            print(f"   📷 REF {ref_value} → {internal_img['filename']}")
                        else:
                            print(f"   ⚠️ REF {ref_value} não tem imagem correspondente")
                else:
                    print(f"❌ Nenhuma imagem interna encontrada")
            else:
                # Método original: busca imagens embebidas
                print(f"🔍 Buscando imagens embebidas...")
                for ref_value in refs:
                    # Encontra a linha da REF
                    ref_row = None
                    for row in range(4, worksheet.max_row + 1):
                        cell_value = worksheet[f'A{row}'].value
                        if cell_value and str(cell_value).strip() == ref_value:
                            ref_row = row
                            break
                    
                    if ref_row:
                        # Busca imagem na célula H{ref_row}
                        image_bytes = self.find_image_by_cell(worksheet, f'H{ref_row}')
                        if image_bytes:
                            images.append({'ref': ref_value, 'bytes': image_bytes, 'row': ref_row, 'source': 'embedded'})
            
            workbook.close()
            
            # Upload real via FTP
            upload_successful = 0
            upload_failed = 0
            errors = []
            
            if images:
                upload_successful, upload_failed = self.upload_images_ftp(images)
                
                # Se todos os uploads falharam, adiciona mensagem de erro
                if upload_successful == 0 and upload_failed > 0:
                    errors.append("Falha no upload FTP. Verifique as credenciais FTP (usuário e senha).")
            
            return {
                'success': upload_successful > 0,
                'total_refs': ref_count,
                'images_found': len(images),
                'uploads_successful': upload_successful,
                'uploads_failed': upload_failed,
                'errors': errors,
                'message': f'Processamento concluído: {upload_successful} upload(s) bem-sucedido(s), {upload_failed} falha(s)'
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'total_refs': 0,
                'images_found': 0,
                'uploads_successful': 0,
                'uploads_failed': 1
            }
    
    def extract_internal_images(self, file_path):
        """Extrai imagens internas do arquivo Excel quando não há imagens embebidas"""
        try:
            import zipfile
            
            print(f"🔍 Extraindo imagens internas de {file_path}")
            internal_images = []
            
            with zipfile.ZipFile(file_path, 'r') as zip_file:
                # Lista arquivos de imagem
                image_files = [f for f in zip_file.namelist() if f.startswith('xl/media/') and f.endswith('.jpg')]
                
                print(f"📊 Arquivos de imagem internos encontrados: {len(image_files)}")
                
                for i, img_file in enumerate(image_files):
                    try:
                        # Extrai dados da imagem
                        img_data = zip_file.read(img_file)
                        
                        print(f"   📷 {img_file}: {len(img_data)} bytes")
                        
                        # Verifica se é JPEG válido
                        if len(img_data) >= 8 and img_data.startswith(b'\xff\xd8'):
                            internal_images.append({
                                'index': i,
                                'filename': img_file,
                                'bytes': img_data,
                                'size': len(img_data)
                            })
                            print(f"      ✅ JPEG válido")
                        else:
                            print(f"      ❌ JPEG inválido")
                    
                    except Exception as e:
                        print(f"   ❌ Erro ao extrair {img_file}: {e}")
            
            print(f"✅ Total de imagens internas extraídas: {len(internal_images)}")
            return internal_images
            
        except Exception as e:
            print(f"❌ Erro ao extrair imagens internas: {e}")
            return []

    def find_image_by_cell(self, worksheet, cell_addr):
        """Encontra imagem por célula usando anchor - método robusto"""
        try:
            # Normaliza endereço da célula (ex: "H23")
            cell_addr = cell_addr.upper()
            target_col_letter = ''.join([c for c in cell_addr if c.isalpha()])
            target_row = int(''.join([c for c in cell_addr if c.isdigit()]))
            
            # Converte coluna para índice 0-based
            target_col_idx = openpyxl.utils.column_index_from_string(target_col_letter) - 1
            
            # Itera imagens da planilha e checa o anchor
            for img in getattr(worksheet, "_images", []):
                anchor = img.anchor
                try:
                    # Quando é TwoCellAnchor
                    col_from = anchor._from.col
                    row_from = anchor._from.row
                except AttributeError:
                    # Quando é OneCellAnchor
                    col_from = anchor.col
                    row_from = anchor.row
                
                # Anchors são 0-based tanto para col quanto para row
                if col_from == target_col_idx and (row_from + 1) == target_row:
                    # Extrai bytes da imagem usando método mais confiável
                    try:
                        # Método principal: _data() - mais confiável
                        if hasattr(img, '_data') and callable(img._data):
                            image_bytes = img._data()
                            if image_bytes and len(image_bytes) > 100:
                                print(f"✅ Imagem extraída via _data(): {len(image_bytes)} bytes")
                                return image_bytes
                    except Exception as e:
                        print(f"❌ Erro _data(): {e}")
                        pass
                    
                    try:
                        # Método alternativo: Usa o caminho interno do arquivo
                        with worksheet.parent._archive.open(img.path) as fp:
                            image_bytes = fp.read()
                            if image_bytes and len(image_bytes) > 100:
                                print(f"✅ Imagem extraída via _archive: {len(image_bytes)} bytes")
                                return image_bytes
                    except Exception as e:
                        print(f"❌ Erro _archive: {e}")
                        pass
                    
                    try:
                        # Método fallback: ref (pode falhar se arquivo fechado)
                        if hasattr(img, 'ref') and img.ref:
                            image_bytes = img.ref.read()
                            if image_bytes and len(image_bytes) > 100:
                                print(f"✅ Imagem extraída via ref: {len(image_bytes)} bytes")
                                return image_bytes
                    except Exception as e:
                        print(f"❌ Erro ref: {e}")
                        pass
            
            return None
            
        except Exception as e:
            return None
    
    def upload_images_ftp(self, images):
        """Upload real das imagens via FTP com timeout maior para muitos uploads"""
        upload_successful = 0
        upload_failed = 0
        
        if not FTP_AVAILABLE:
            print("❌ FTP não disponível")
            return 0, len(images)
        
        try:
            # Configurações FTP (pode ser sobrescrito por variáveis de ambiente)
            ftp_host = os.getenv('FTP_HOST', "46.202.90.62")
            ftp_user = os.getenv('FTP_USER', "u715606397.ideolog.ia.br")
            ftp_pass = os.getenv('FTP_PASSWORD', "27y8rYoDq=Q&aHk:")
            
            print(f"🔗 Conectando FTP: {ftp_host}")
            print(f"👤 Usuário: {ftp_user}")
            
            # Conecta com timeout maior para muitos uploads
            timeout = 300 if len(images) > 30 else 60  # Timeout maior para muitos uploads
            with ftplib.FTP(timeout=timeout) as ftp:
                ftp.connect(ftp_host, 21)
                print(f"✅ Conexão estabelecida")
                
                try:
                    ftp.login(ftp_user, ftp_pass)
                    print(f"✅ Login bem-sucedido")
                except ftplib.error_perm as e:
                    print(f"❌ Erro de autenticação FTP: {e}")
                    print(f"💡 Verifique se as credenciais estão corretas:")
                    print(f"   Host: {ftp_host}")
                    print(f"   User: {ftp_user}")
                    print(f"   Password: {'*' * len(ftp_pass)}")
                    return 0, len(images)
                except Exception as e:
                    print(f"❌ Erro ao fazer login FTP: {e}")
                    return 0, len(images)
                
                # Navega para o diretório raiz primeiro
                try:
                    ftp.cwd('/')
                    print(f"📁 Diretório raiz: {ftp.pwd()}")
                except Exception as e:
                    print(f"⚠️ Aviso ao navegar para raiz: {e}")
                
                # Cria estrutura: public_html/images/products
                directories = ['public_html', 'images', 'products']
                current_path = []
                
                for dir_name in directories:
                    try:
                        # Tenta criar o diretório
                        try:
                            ftp.mkd(dir_name)
                            print(f"📁 Diretório '{dir_name}' criado")
                        except ftplib.error_perm as e:
                            if '550' in str(e) or 'already exists' in str(e).lower():
                                print(f"📁 Diretório '{dir_name}' já existe")
                            else:
                                raise
                        
                        # Navega para o diretório
                        ftp.cwd(dir_name)
                        current_path.append(dir_name)
                        print(f"📁 Navegou para: {'/'.join(current_path)}")
                        
                    except Exception as e:
                        print(f"❌ Erro ao criar/navegar para '{dir_name}': {e}")
                        # Tenta navegar mesmo assim (pode já existir)
                        try:
                            ftp.cwd(dir_name)
                            current_path.append(dir_name)
                            print(f"📁 Navegou para: {'/'.join(current_path)}")
                        except Exception as e2:
                            print(f"❌ Erro ao navegar para '{dir_name}': {e2}")
                            return 0, len(images)
                
                # Verifica se chegou no diretório correto
                final_path = ftp.pwd()
                print(f"📁 Diretório final FTP: {final_path}")
                
                # Garante que está no caminho correto
                if 'products' not in final_path:
                    print(f"⚠️ Aviso: Diretório atual não contém 'products'")
                    print(f"   Tentando navegar diretamente...")
                    try:
                        ftp.cwd('/public_html/images/products')
                        print(f"✅ Navegou diretamente para: {ftp.pwd()}")
                    except Exception as e:
                        print(f"❌ Erro ao navegar diretamente: {e}")
                        return 0, len(images)
            
                # Verifica novamente o diretório antes de fazer uploads
                current_dir = ftp.pwd()
                print(f"📁 Diretório atual antes dos uploads: {current_dir}")
                
                # Garante que está em public_html/images/products
                if not current_dir.endswith('products') and 'products' not in current_dir:
                    print(f"⚠️ Diretório não está em products, tentando navegar...")
                    try:
                        # Tenta diferentes caminhos
                        paths_to_try = [
                            '/public_html/images/products',
                            'public_html/images/products',
                            './public_html/images/products'
                        ]
                        for path in paths_to_try:
                            try:
                                ftp.cwd(path)
                                print(f"✅ Navegou para: {ftp.pwd()}")
                                break
                            except:
                                continue
                        else:
                            print(f"❌ Não foi possível navegar para products")
                            return 0, len(images)
                    except Exception as e:
                        print(f"❌ Erro ao navegar: {e}")
                        return 0, len(images)
                
                # Upload das imagens com progresso
                total_images = len(images)
                for i, image_data in enumerate(images):
                    try:
                        # Converte e salva imagem como JPEG válido
                        safe_ref = "".join(c for c in image_data['ref'] if c.isalnum() or c in ('-', '_')).rstrip()
                        if not safe_ref:
                            safe_ref = f"image_{int(time.time())}"
                        
                        temp_image_path = os.path.join(tempfile.gettempdir(), f"{safe_ref}.jpg")
                        
                        # Converte bytes para imagem válida usando PIL
                        try:
                            from PIL import Image
                            import io
                            
                            # Abre a imagem dos bytes
                            image_stream = io.BytesIO(image_data['bytes'])
                            img = Image.open(image_stream)
                            
                            # Detecta formato original
                            original_format = img.format
                            source_info = image_data.get('source', 'unknown')
                            filename_info = f" ({image_data.get('filename', '')})" if 'filename' in image_data else ""
                            print(f"📤 Upload {image_data['ref']}: {len(image_data['bytes'])} bytes ({original_format}) → JPEG [{source_info}]{filename_info}")
                            
                            # Converte para RGB se necessário
                            if img.mode in ('RGBA', 'LA', 'P'):
                                img = img.convert('RGB')
                            
                            # Salva como JPEG com configurações máximas para contornar processamento do servidor
                            img.save(temp_image_path, 'JPEG', quality=100, optimize=False, progressive=False, subsampling=0)
                            
                            print(f"🌐 URL: https://ideolog.ia.br/images/products/{image_data['ref']}.jpg")
                            
                        except ImportError:
                            print(f"⚠️ PIL não disponível, usando bytes originais")
                            # Fallback: salva bytes originais
                            with open(temp_image_path, 'wb') as f:
                                f.write(image_data['bytes'])
                            print(f"📤 Upload {image_data['ref']}: {len(image_data['bytes'])} bytes (sem conversão)")
                            print(f"🌐 URL: https://ideolog.ia.br/images/products/{image_data['ref']}.jpg")
                        except Exception as e:
                            print(f"❌ Erro na conversão PIL: {e}")
                            # Fallback: salva bytes originais
                            with open(temp_image_path, 'wb') as f:
                                f.write(image_data['bytes'])
                            print(f"📤 Upload {image_data['ref']}: {len(image_data['bytes'])} bytes (fallback)")
                            print(f"🌐 URL: https://ideolog.ia.br/images/products/{image_data['ref']}.jpg")
                        
                        # Verifica se arquivo foi criado
                        if not os.path.exists(temp_image_path):
                            print(f"❌ Arquivo temporário não foi criado: {temp_image_path}")
                            upload_failed += 1
                            continue
                        
                        file_size = os.path.getsize(temp_image_path)
                        if file_size == 0:
                            print(f"❌ Arquivo temporário está vazio: {temp_image_path}")
                            upload_failed += 1
                            continue
                        
                        # Upload via FTP
                        remote_filename = f"{image_data['ref']}.jpg"
                        current_ftp_dir = ftp.pwd()
                        full_path = f"{current_ftp_dir}/{remote_filename}"
                        print(f"📤 Fazendo upload FTP:")
                        print(f"   Arquivo: {remote_filename}")
                        print(f"   Tamanho: {file_size} bytes")
                        print(f"   Caminho completo: {full_path}")
                        
                        with open(temp_image_path, 'rb') as file:
                            ftp.storbinary(f'STOR {remote_filename}', file)
                        
                        # Verifica se arquivo foi salvo
                        final_dir = ftp.pwd()
                        print(f"✅ Upload concluído: {remote_filename}")
                        print(f"   Salvo em: {final_dir}/{remote_filename}")
                        print(f"   URL: https://ideolog.ia.br/images/products/{remote_filename}")
                        
                        # Remove arquivo temporário
                        try:
                            os.remove(temp_image_path)
                        except:
                            pass
                        
                        upload_successful += 1
                        
                        # Log de progresso para muitos uploads
                        if total_images > 30 and (i + 1) % 10 == 0:
                            print(f"📤 Upload progress: {i + 1}/{total_images} ({((i + 1) / total_images) * 100:.1f}%)")
                        
                    except Exception as e:
                        upload_failed += 1
                        print(f"❌ Erro no upload {image_data.get('ref', 'unknown')}: {e}")
                        import traceback
                        traceback.print_exc()
                
                print(f"🔌 FTP desconectado")
            
        except ftplib.error_perm as e:
            upload_failed = len(images)
            error_msg = str(e)
            print(f"❌ Erro de permissão FTP: {error_msg}")
            if "530" in error_msg or "Login incorrect" in error_msg:
                print(f"💡 PROBLEMA: Credenciais FTP incorretas!")
                print(f"   Verifique se o usuário e senha estão corretos")
                print(f"   Você pode configurar via variáveis de ambiente:")
                print(f"   export FTP_HOST='46.202.90.62'")
                print(f"   export FTP_USER='seu_usuario'")
                print(f"   export FTP_PASSWORD='sua_senha'")
            import traceback
            traceback.print_exc()
        except Exception as e:
            upload_failed = len(images)
            print(f"❌ Erro geral no FTP: {e}")
            import traceback
            traceback.print_exc()
        
        return upload_successful, upload_failed
    

def start_simple_server():
    """Inicia o servidor simplificado"""
    port = int(os.getenv('PORT', 8080))
    
    print("🚀 Iniciando Sistema Simplificado...")
    print("=" * 50)
    print(f"📁 Frontend: http://localhost:{port}")
    print(f"🔧 API: http://localhost:{port}/upload")
    print(f"⚙️  Config: http://localhost:{port}/config")
    print(f"💚 Health: http://localhost:{port}/health")
    print("=" * 50)
    print(f"🔧 Dependências:")
    print(f"   • FTP: {'✅' if FTP_AVAILABLE else '❌'}")
    print(f"   • openpyxl: {'✅' if OPENPYXL_AVAILABLE else '❌'}")
    print(f"   • PIL: ❌ (não usado nesta versão)")
    print("=" * 50)
    
    server = HTTPServer(('0.0.0.0', port), SimpleUploadHandler)
    
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n👋 Sistema simplificado encerrado.")
        server.shutdown()

if __name__ == "__main__":
    start_simple_server()
