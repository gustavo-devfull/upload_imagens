#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Análise do arquivo nova.xlsx para verificar imagens
"""

import openpyxl
import os

def analisar_nova_xlsx():
    """Analisa o arquivo nova.xlsx para verificar imagens"""
    
    print("🔍 Analisando arquivo nova.xlsx")
    print("=" * 40)
    
    arquivo = "nova.xlsx"
    
    if not os.path.exists(arquivo):
        print(f"❌ Arquivo {arquivo} não encontrado")
        return
    
    try:
        # Carrega o arquivo Excel
        workbook = openpyxl.load_workbook(arquivo, data_only=True, read_only=False)
        worksheet = workbook.active
        
        print(f"📊 Planilha ativa: {worksheet.title}")
        print(f"📏 Dimensões: {worksheet.max_row} linhas x {worksheet.max_column} colunas")
        
        # Verifica se há imagens
        total_imagens = len(worksheet._images)
        print(f"🖼️ Total de imagens encontradas: {total_imagens}")
        
        if total_imagens == 0:
            print("❌ Nenhuma imagem encontrada na planilha")
            workbook.close()
            return
        
        # Analisa cada imagem
        print(f"\n🔍 Análise detalhada das imagens:")
        for i, image in enumerate(worksheet._images):
            print(f"\n📷 Imagem {i+1}:")
            
            # Informações básicas
            print(f"   Tipo: {type(image)}")
            print(f"   Dimensões: {image.width} x {image.height}")
            
            # Posição da imagem
            anchor = image.anchor
            try:
                if hasattr(anchor, '_from'):
                    col_from = anchor._from.col
                    row_from = anchor._from.row
                else:
                    col_from = anchor.col
                    row_from = anchor.row
                
                # Converte para notação Excel
                col_letter = openpyxl.utils.get_column_letter(col_from + 1)
                row_number = row_from + 1
                
                print(f"   Posição: {col_letter}{row_number}")
                
                # Verifica se está na coluna H (PHOTO)
                if col_from == 7:  # H = 7 (0-based)
                    print(f"   ✅ Está na coluna H (PHOTO)")
                else:
                    print(f"   ⚠️ Está na coluna {col_letter}")
                
            except Exception as e:
                print(f"   ❌ Erro ao determinar posição: {e}")
            
            # Verifica dados da imagem
            try:
                if hasattr(image, '_data') and callable(image._data):
                    data_bytes = image._data()
                    print(f"   📊 Dados da imagem: {len(data_bytes)} bytes")
                    
                    # Verifica cabeçalho
                    if len(data_bytes) >= 8:
                        header = data_bytes[:8]
                        print(f"   📋 Cabeçalho: {header.hex()}")
                        
                        # Detecta formato
                        if header.startswith(b'\xff\xd8'):
                            print(f"   ✅ Formato: JPEG")
                        elif header.startswith(b'\x89PNG'):
                            print(f"   ✅ Formato: PNG")
                        elif header.startswith(b'GIF'):
                            print(f"   ✅ Formato: GIF")
                        else:
                            print(f"   ❓ Formato: Desconhecido")
                
            except Exception as e:
                print(f"   ❌ Erro ao obter dados: {e}")
        
        # Verifica REFs na coluna A
        print(f"\n📋 Verificando REFs na coluna A:")
        refs_encontradas = []
        
        for row in range(1, worksheet.max_row + 1):
            cell_value = worksheet[f'A{row}'].value
            if cell_value and str(cell_value).strip():
                ref_value = str(cell_value).strip()
                if ref_value.upper() not in ['TOTAL', 'SUBTOTAL', 'SUM', 'COUNT']:
                    refs_encontradas.append((row, ref_value))
        
        print(f"   📊 Total de REFs encontradas: {len(refs_encontradas)}")
        
        # Mostra algumas REFs
        for i, (row, ref) in enumerate(refs_encontradas[:10]):
            print(f"   Linha {row}: {ref}")
        
        if len(refs_encontradas) > 10:
            print(f"   ... e mais {len(refs_encontradas) - 10} REFs")
        
        # Verifica se há imagens na coluna H correspondentes às REFs
        print(f"\n🔍 Verificando correspondência REF ↔ Imagem:")
        imagens_com_refs = 0
        
        for row, ref in refs_encontradas:
            # Verifica se há imagem na coluna H desta linha
            tem_imagem = False
            for image in worksheet._images:
                try:
                    anchor = image.anchor
                    if hasattr(anchor, '_from'):
                        col_from = anchor._from.col
                        row_from = anchor._from.row
                    else:
                        col_from = anchor.col
                        row_from = anchor.row
                    
                    # Verifica se está na coluna H (7) e linha correspondente
                    if col_from == 7 and (row_from + 1) == row:
                        tem_imagem = True
                        break
                
                except:
                    continue
            
            if tem_imagem:
                imagens_com_refs += 1
                print(f"   ✅ REF {ref} (linha {row}) tem imagem na coluna H")
        
        print(f"\n📊 Resumo:")
        print(f"   • Total de imagens: {total_imagens}")
        print(f"   • Total de REFs: {len(refs_encontradas)}")
        print(f"   • REFs com imagens: {imagens_com_refs}")
        
        if imagens_com_refs > 0:
            print(f"\n✅ Arquivo válido para exportação!")
            print(f"   Pode exportar {imagens_com_refs} imagens")
        else:
            print(f"\n❌ Arquivo não tem imagens na coluna H")
            print(f"   Não pode ser usado para exportação")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Erro ao analisar arquivo: {e}")
        import traceback
        traceback.print_exc()

def testar_extração_imagens():
    """Testa extração de imagens do arquivo nova.xlsx"""
    
    print(f"\n🧪 Testando extração de imagens:")
    
    try:
        workbook = openpyxl.load_workbook("nova.xlsx", data_only=True, read_only=False)
        worksheet = workbook.active
        
        if len(worksheet._images) == 0:
            print("❌ Nenhuma imagem para testar")
            workbook.close()
            return
        
        # Testa extração da primeira imagem
        image = worksheet._images[0]
        
        print(f"🔍 Testando extração da primeira imagem...")
        
        # Método _data()
        try:
            image_bytes = image._data()
            print(f"✅ Método _data(): {len(image_bytes)} bytes")
            
            # Verifica cabeçalho
            if len(image_bytes) >= 8:
                header = image_bytes[:8]
                print(f"📋 Cabeçalho: {header.hex()}")
                
                if header.startswith(b'\xff\xd8'):
                    print(f"✅ Formato JPEG detectado")
                elif header.startswith(b'\x89PNG'):
                    print(f"✅ Formato PNG detectado")
                else:
                    print(f"❓ Formato desconhecido")
            
        except Exception as e:
            print(f"❌ Erro no método _data(): {e}")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Erro: {e}")

if __name__ == "__main__":
    analisar_nova_xlsx()
    testar_extração_imagens()
    
    print(f"\n💡 Próximos passos:")
    print(f"1. Se há imagens válidas, pode usar o sistema de upload")
    print(f"2. Se não há imagens, verifique se estão na coluna H")
    print(f"3. Teste o upload com este arquivo se válido")

