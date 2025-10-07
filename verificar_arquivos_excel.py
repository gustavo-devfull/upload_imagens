#!/usr/bin/env python3
"""
Script para verificar quais arquivos Excel têm imagens na coluna H
"""

import openpyxl
import os

def verificar_imagens_arquivo(arquivo):
    """Verifica se um arquivo Excel tem imagens na coluna H"""
    try:
        workbook = openpyxl.load_workbook(arquivo)
        worksheet = workbook.active
        
        print(f"\n📁 Analisando: {arquivo}")
        print("=" * 50)
        
        # Conta REFs
        refs = []
        for row in range(4, worksheet.max_row + 1):
            cell_value = worksheet[f'A{row}'].value
            if cell_value and str(cell_value).strip() and str(cell_value).upper() not in ['TOTAL', 'SUBTOTAL', '']:
                refs.append(str(cell_value).strip())
        
        print(f"📊 REFs encontradas: {len(refs)}")
        if refs:
            print(f"   Primeiras REFs: {refs[:3]}")
        
        # Conta imagens na coluna H
        images_h = []
        for image in worksheet._images:
            if hasattr(image, 'anchor') and image.anchor:
                anchor = image.anchor
                if hasattr(anchor, '_from') and anchor._from:
                    col_idx = anchor._from.col
                    row_idx = anchor._from.row + 1
                    col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                    
                    if col_letter == 'H' and row_idx >= 4:
                        ref_cell = worksheet[f'A{row_idx}']
                        if ref_cell.value:
                            ref_value = str(ref_cell.value).strip()
                            if ref_value and ref_value.upper() not in ['TOTAL', 'SUBTOTAL', '']:
                                images_h.append({
                                    'ref': ref_value,
                                    'row': row_idx,
                                    'col': col_letter
                                })
        
        print(f"🖼️  Imagens na coluna H: {len(images_h)}")
        if images_h:
            print(f"   Imagens encontradas:")
            for img in images_h:
                print(f"   - REF: {img['ref']} (linha {img['row']})")
        
        # Conta todas as imagens
        total_images = len(worksheet._images)
        print(f"📸 Total de imagens no arquivo: {total_images}")
        
        # Recomendação
        if len(images_h) > 0:
            print("✅ RECOMENDADO: Este arquivo tem imagens na coluna H")
            return True
        else:
            print("❌ NÃO RECOMENDADO: Este arquivo não tem imagens na coluna H")
            return False
            
    except Exception as e:
        print(f"❌ Erro ao processar {arquivo}: {e}")
        return False

def main():
    """Função principal"""
    print("🔍 VERIFICADOR DE ARQUIVOS EXCEL")
    print("=" * 50)
    
    # Lista arquivos Excel
    arquivos_excel = [f for f in os.listdir('.') if f.endswith('.xlsx')]
    
    if not arquivos_excel:
        print("❌ Nenhum arquivo .xlsx encontrado")
        return
    
    print(f"📁 Arquivos Excel encontrados: {len(arquivos_excel)}")
    
    arquivos_recomendados = []
    arquivos_nao_recomendados = []
    
    for arquivo in arquivos_excel:
        if verificar_imagens_arquivo(arquivo):
            arquivos_recomendados.append(arquivo)
        else:
            arquivos_nao_recomendados.append(arquivo)
    
    print("\n" + "=" * 50)
    print("📋 RESUMO FINAL")
    print("=" * 50)
    
    if arquivos_recomendados:
        print("✅ ARQUIVOS RECOMENDADOS (têm imagens na coluna H):")
        for arquivo in arquivos_recomendados:
            print(f"   • {arquivo}")
    
    if arquivos_nao_recomendados:
        print("\n❌ ARQUIVOS NÃO RECOMENDADOS (sem imagens na coluna H):")
        for arquivo in arquivos_nao_recomendados:
            print(f"   • {arquivo}")
    
    print(f"\n💡 Use os arquivos recomendados para testar o sistema de upload!")

if __name__ == "__main__":
    main()
