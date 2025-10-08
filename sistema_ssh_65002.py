#!/usr/bin/env python3
"""
Sistema de Upload de Imagens Excel - Versão SSH Porta 65002
Versão otimizada para SSH na porta 65002 no PythonAnywhere
"""

import os
import sys
import json
import tempfile
import time
import logging
from datetime import datetime
import socket

# Configuração de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Imports condicionais
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.error("❌ openpyxl não disponível")

try:
    import paramiko
    SSH_AVAILABLE = True
except ImportError:
    SSH_AVAILABLE = False
    logger.error("❌ paramiko não disponível")

try:
    from flask import Flask, request, jsonify, render_template_string
    FLASK_AVAILABLE = True
except ImportError:
    FLASK_AVAILABLE = False
    logger.error("❌ Flask não disponível")

# Configurações SSH - Porta 65002
SSH_HOST = "46.202.90.62"
SSH_PORT = 65002
SSH_USER = "u715606397"
SSH_PASS = "]X9CC>t~ihWhdzNq"
SSH_DIR = "/public_html/images/products"

# HTML Template
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>🚀 Sistema Upload Excel - SSH 65002</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 40px; background: #f5f5f5; }
        .container { max-width: 600px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        .header { text-align: center; margin-bottom: 30px; }
        .header h1 { color: #333; margin-bottom: 10px; }
        .upload-area { text-align: center; margin: 30px 0; }
        .file-input { margin: 20px 0; }
        .btn { background: #007bff; color: white; padding: 12px 24px; border: none; border-radius: 5px; cursor: pointer; font-size: 16px; margin: 10px; }
        .btn:hover { background: #0056b3; }
        .btn:disabled { background: #ccc; cursor: not-allowed; }
        .results { margin: 20px 0; padding: 20px; background: #f8f9fa; border-radius: 5px; display: none; }
        .success { background: #d4edda; border: 1px solid #c3e6cb; color: #155724; }
        .error { background: #f8d7da; border: 1px solid #f5c6cb; color: #721c24; }
        .warning { background: #fff3cd; border: 1px solid #ffeaa7; color: #856404; }
        .debug { background: #e2e3e5; border: 1px solid #d6d8db; color: #383d41; font-family: monospace; font-size: 12px; }
        .info { background: #d1ecf1; border: 1px solid #bee5eb; color: #0c5460; }
        .connection-test { background: #f0f8ff; border: 1px solid #87ceeb; color: #0066cc; }
        .progress-bar { width: 100%; background-color: #f0f0f0; border-radius: 10px; overflow: hidden; margin: 20px 0; }
        .progress-fill { height: 30px; background: linear-gradient(90deg, #007bff, #0056b3); width: 0%; transition: width 0.3s ease; display: flex; align-items: center; justify-content: center; color: white; font-weight: bold; }
        .steps { display: flex; justify-content: space-between; margin: 20px 0; }
        .step { flex: 1; text-align: center; padding: 10px; margin: 0 5px; border-radius: 5px; background: #f8f9fa; border: 2px solid #dee2e6; }
        .step.active { background: #007bff; color: white; border-color: #0056b3; }
        .step.completed { background: #28a745; color: white; border-color: #1e7e34; }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🚀 Sistema Upload Excel</h1>
            <p>Versão SSH Porta 65002 - Upload de Imagens Automático</p>
        </div>
        
        <div class="info">
            <h3>📋 Informações do Sistema</h3>
            <p><strong>Plataforma:</strong> PythonAnywhere</p>
            <p><strong>Python:</strong> {{ python_version }}</p>
            <p><strong>Dependências:</strong> Flask ✅ | openpyxl ✅ | SSH ✅</p>
            <p><strong>Status:</strong> Sistema funcionando</p>
        </div>
        
        <div class="connection-test">
            <h3>🔌 Configuração SSH</h3>
            <p><strong>Host:</strong> {{ ssh_host }}</p>
            <p><strong>Porta:</strong> {{ ssh_port }}</p>
            <p><strong>Usuário:</strong> {{ ssh_user }}</p>
            <p><strong>Status:</strong> {{ connection_status }}</p>
        </div>
        
        <div class="upload-area">
            <h2>📁 Selecione seu arquivo Excel</h2>
            <p>Arquivos suportados: .xlsx com imagens na coluna H</p>
            
            <div class="file-input">
                <input type="file" id="excelFile" accept=".xlsx" />
            </div>
            
            <div id="fileName"></div>
            
            <button id="uploadBtn" class="btn" onclick="uploadFile()" disabled>🚀 Fazer Upload</button>
            <button id="resetBtn" class="btn" onclick="resetForm()" disabled>🔄 Resetar</button>
        </div>
        
        <div class="progress-bar" id="progressBar" style="display: none;">
            <div class="progress-fill" id="progressFill">0%</div>
        </div>
        
        <div class="steps" id="steps" style="display: none;">
            <div class="step" id="step1">📁 Carregando Excel</div>
            <div class="step" id="step2">🔍 Detectando Imagens</div>
            <div class="step" id="step3">🔌 Conectando SSH</div>
            <div class="step" id="step4">⬆️ Fazendo Upload</div>
            <div class="step" id="step5">✅ Concluído</div>
        </div>
        
        <div id="results" class="results"></div>
    </div>

    <script>
        let selectedFile = null;
        let startTime = null;
        
        document.getElementById('excelFile').addEventListener('change', function(e) {
            const file = e.target.files[0];
            if (file) {
                selectedFile = file;
                document.getElementById('fileName').innerHTML = `📄 ${file.name} (${(file.size / 1024 / 1024).toFixed(2)} MB)`;
                document.getElementById('uploadBtn').disabled = false;
                document.getElementById('resetBtn').disabled = false;
            }
        });
        
        function updateProgress(percent, step) {
            const progressFill = document.getElementById('progressFill');
            const progressBar = document.getElementById('progressBar');
            const steps = document.getElementById('steps');
            
            progressBar.style.display = 'block';
            steps.style.display = 'flex';
            
            progressFill.style.width = percent + '%';
            progressFill.textContent = percent + '%';
            
            // Atualiza steps
            for (let i = 1; i <= 5; i++) {
                const stepEl = document.getElementById(`step${i}`);
                if (i < step) {
                    stepEl.className = 'step completed';
                } else if (i === step) {
                    stepEl.className = 'step active';
                } else {
                    stepEl.className = 'step';
                }
            }
        }
        
        function uploadFile() {
            if (!selectedFile) {
                alert('Por favor, selecione um arquivo Excel.');
                return;
            }
            
            startTime = Date.now();
            updateProgress(10, 1);
            
            const formData = new FormData();
            formData.append('excel_file', selectedFile);
            
            document.getElementById('uploadBtn').disabled = true;
            document.getElementById('resetBtn').disabled = true;
            document.getElementById('results').style.display = 'block';
            document.getElementById('results').innerHTML = '<div class="debug">🔄 Processando arquivo...</div>';
            
            fetch('/upload', {
                method: 'POST',
                body: formData
            })
            .then(response => response.json())
            .then(data => {
                updateProgress(100, 5);
                showResults(data);
            })
            .catch(error => {
                document.getElementById('results').innerHTML = `<div class="error"><h3>❌ Erro no Upload</h3><p>${error.message}</p></div>`;
                document.getElementById('uploadBtn').disabled = false;
                document.getElementById('resetBtn').disabled = false;
            });
        }
        
        function showResults(data) {
            const resultsDiv = document.getElementById('results');
            const endTime = Date.now();
            const duration = ((endTime - startTime) / 1000).toFixed(2);
            
            let html = '';
            
            if (data.debug_info) {
                html += `<div class="debug"><h3>🔍 Debug Info</h3><pre>${data.debug_info}</pre></div>`;
            }
            
            if (data.success && data.uploads_successful > 0) {
                html += `<div class="success"><h3>✅ Upload Concluído!</h3><p>${data.uploads_successful} imagem(ns) enviada(s) via SSH</p><p>⏱️ Tempo total: ${duration}s</p></div>`;
            } else if (data.error) {
                html += `<div class="warning"><h3>⚠️ Problema</h3><p>${data.error}</p></div>`;
            } else {
                html += `<div class="error"><h3>❌ Erro</h3><p>Não foi possível processar o arquivo</p></div>`;
            }
            
            resultsDiv.innerHTML = html;
            document.getElementById('uploadBtn').disabled = false;
            document.getElementById('resetBtn').disabled = false;
        }
        
        function resetForm() {
            document.getElementById('excelFile').value = '';
            document.getElementById('fileName').innerHTML = '';
            document.getElementById('results').style.display = 'none';
            document.getElementById('progressBar').style.display = 'none';
            document.getElementById('steps').style.display = 'none';
            document.getElementById('uploadBtn').disabled = true;
            document.getElementById('resetBtn').disabled = true;
            selectedFile = null;
        }
    </script>
</body>
</html>
"""

def test_ssh_connection():
    """Testa conectividade SSH na porta 65002"""
    try:
        debug_info = []
        debug_info.append(f"🔌 Testando conexão SSH: {SSH_HOST}:{SSH_PORT}")
        
        # Testa conectividade TCP
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(10)
        result = sock.connect_ex((SSH_HOST, SSH_PORT))
        sock.close()
        
        if result == 0:
            debug_info.append(f"✅ TCP conectado: {SSH_HOST}:{SSH_PORT}")
            
            # Testa SSH
            try:
                ssh = paramiko.SSHClient()
                ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                ssh.connect(
                    SSH_HOST, 
                    SSH_PORT, 
                    SSH_USER, 
                    SSH_PASS,
                    timeout=10
                )
                sftp = ssh.open_sftp()
                sftp.close()
                ssh.close()
                
                debug_info.append(f"✅ SSH funcionando: {SSH_HOST}:{SSH_PORT}")
                return True, debug_info
                
            except Exception as e:
                debug_info.append(f"❌ SSH falhou: {e}")
                return False, debug_info
        else:
            debug_info.append(f"❌ TCP falhou: {SSH_HOST}:{SSH_PORT}")
            return False, debug_info
            
    except Exception as e:
        debug_info.append(f"❌ Erro geral: {e}")
        return False, debug_info

def create_app():
    """Cria aplicação Flask para PythonAnywhere"""
    app = Flask(__name__)
    
    @app.route('/')
    def index():
        python_version = f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"
        
        # Testa conectividade SSH
        ssh_working, connection_debug = test_ssh_connection()
        connection_status = "✅ Conectado" if ssh_working else "❌ Sem conexão"
        
        return render_template_string(HTML_TEMPLATE, 
                                    python_version=python_version,
                                    ssh_host=SSH_HOST,
                                    ssh_port=SSH_PORT,
                                    ssh_user=SSH_USER,
                                    connection_status=connection_status)
    
    @app.route('/health')
    def health():
        """Health check para PythonAnywhere"""
        try:
            # Testa conectividade SSH
            ssh_working, connection_debug = test_ssh_connection()
            
            health_data = {
                "status": "ok",
                "message": "Sistema SSH Porta 65002 funcionando",
                "timestamp": time.time(),
                "platform": "PythonAnywhere",
                "python_version": f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}",
                "ssh_working": ssh_working,
                "ssh_config": {
                    "host": SSH_HOST,
                    "port": SSH_PORT,
                    "user": SSH_USER
                },
                "connection_debug": connection_debug,
                "dependencies": {
                    "ssh": SSH_AVAILABLE,
                    "openpyxl": OPENPYXL_AVAILABLE,
                    "flask": FLASK_AVAILABLE
                }
            }
            
            if not FLASK_AVAILABLE:
                health_data["status"] = "error"
                health_data["message"] = "Flask não disponível"
                return jsonify(health_data), 500
            
            return jsonify(health_data), 200
            
        except Exception as e:
            logger.error(f"❌ Erro no health check: {e}")
            return jsonify({
                "status": "error",
                "message": f"Erro no health check: {str(e)}",
                "timestamp": time.time(),
                "platform": "PythonAnywhere"
            }), 500
    
    @app.route('/upload', methods=['POST'])
    def upload():
        """Upload de arquivo Excel"""
        debug_info = []
        
        try:
            logger.info("=== INÍCIO DO UPLOAD SSH 65002 ===")
            debug_info.append("=== INÍCIO DO UPLOAD SSH 65002 ===")
            
            if not SSH_AVAILABLE or not OPENPYXL_AVAILABLE:
                error_msg = 'Dependências não disponíveis'
                logger.error(error_msg)
                debug_info.append(f"❌ {error_msg}")
                return jsonify({
                    'error': error_msg,
                    'debug_info': '\n'.join(debug_info)
                })
            
            if 'excel_file' not in request.files:
                error_msg = 'Arquivo não encontrado'
                logger.error(error_msg)
                debug_info.append(f"❌ {error_msg}")
                return jsonify({
                    'error': error_msg,
                    'debug_info': '\n'.join(debug_info)
                })
            
            file = request.files['excel_file']
            logger.info(f"📁 Arquivo recebido: {file.filename}")
            debug_info.append(f"📁 Arquivo recebido: {file.filename}")
            
            if file.filename == '':
                error_msg = 'Nenhum arquivo selecionado'
                logger.error(error_msg)
                debug_info.append(f"❌ {error_msg}")
                return jsonify({
                    'error': error_msg,
                    'debug_info': '\n'.join(debug_info)
                })
            
            # Salva arquivo temporariamente
            temp_path = os.path.join('/tmp', file.filename)
            file.save(temp_path)
            logger.info(f"💾 Arquivo salvo em: {temp_path}")
            debug_info.append(f"💾 Arquivo salvo em: {temp_path}")
            
            # Processamento
            stats = process_excel(temp_path, debug_info)
            
            # Remove arquivo temporário
            os.remove(temp_path)
            
            # Adiciona debug info
            stats['debug_info'] = '\n'.join(debug_info)
            return jsonify(stats)
            
        except Exception as e:
            error_msg = str(e)
            logger.error(f"❌ Erro geral: {error_msg}")
            debug_info.append(f"❌ Erro geral: {error_msg}")
            
            return jsonify({
                'error': error_msg,
                'debug_info': '\n'.join(debug_info)
            })
    
    return app

def process_excel(file_path, debug_info):
    """Processamento do Excel"""
    try:
        logger.info(f"📊 Carregando arquivo: {file_path}")
        debug_info.append(f"📊 Carregando arquivo: {file_path}")
        
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        logger.info(f"📋 Planilha: {worksheet.title}")
        debug_info.append(f"📋 Planilha: {worksheet.title}")
        
        # Processa REFs
        refs = []
        for row in range(4, worksheet.max_row + 1):
            cell_value = worksheet[f'A{row}'].value
            if cell_value and str(cell_value).strip() and str(cell_value).upper() not in ['TOTAL', 'SUBTOTAL', '']:
                refs.append(str(cell_value).strip())
        
        logger.info(f"📊 REFs encontradas: {len(refs)}")
        debug_info.append(f"📊 REFs encontradas: {len(refs)}")
        
        # Processa imagens
        images = []
        total_images = len(worksheet._images)
        logger.info(f"🖼️ Total de imagens: {total_images}")
        debug_info.append(f"🖼️ Total de imagens: {total_images}")
        
        for i, image in enumerate(worksheet._images):
            logger.info(f"🖼️ Analisando imagem {i+1}/{total_images}")
            debug_info.append(f"🖼️ Analisando imagem {i+1}/{total_images}")
            
            if hasattr(image, 'anchor') and image.anchor:
                anchor = image.anchor
                if hasattr(anchor, '_from') and anchor._from:
                    col_idx = anchor._from.col
                    row_idx = anchor._from.row + 1
                    col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                    
                    logger.info(f"📍 Posição: {col_letter}{row_idx}")
                    debug_info.append(f"📍 Posição: {col_letter}{row_idx}")
                    
                    if col_letter == 'H' and row_idx >= 4:
                        ref_cell = worksheet[f'A{row_idx}']
                        if ref_cell.value:
                            ref_value = str(ref_cell.value).strip()
                            if ref_value and ref_value.upper() not in ['TOTAL', 'SUBTOTAL', '']:
                                logger.info(f"✅ Imagem válida: REF {ref_value}")
                                debug_info.append(f"✅ Imagem válida: REF {ref_value}")
                                images.append({'image': image, 'ref': ref_value})
        
        logger.info(f"📊 Imagens válidas: {len(images)}")
        debug_info.append(f"📊 Imagens válidas: {len(images)}")
        
        # Upload para SSH
        uploads_successful = 0
        uploads_failed = 0
        
        if images:
            # Testa conectividade SSH
            ssh_working, connection_debug = test_ssh_connection()
            
            if not ssh_working:
                error_msg = "SSH não está funcionando"
                debug_info.extend(connection_debug)
                return {
                    'success': False,
                    'error': error_msg,
                    'total_refs': len(refs),
                    'images_found': len(images),
                    'uploads_successful': 0,
                    'uploads_failed': len(images)
                }
            
            debug_info.extend(connection_debug)
            debug_info.append(f"✅ SSH funcionando: {SSH_HOST}:{SSH_PORT}")
        
        for image_data in images:
            try:
                if upload_image_to_ssh(image_data['image'], image_data['ref'], debug_info):
                    uploads_successful += 1
                    logger.info(f"✅ Upload: {image_data['ref']}")
                    debug_info.append(f"✅ Upload: {image_data['ref']}")
                else:
                    uploads_failed += 1
                    logger.error(f"❌ Falha: {image_data['ref']}")
                    debug_info.append(f"❌ Falha: {image_data['ref']}")
            except Exception as e:
                uploads_failed += 1
                logger.error(f"❌ Erro: {image_data['ref']} - {e}")
                debug_info.append(f"❌ Erro: {image_data['ref']} - {e}")
        
        return {
            'success': True,
            'total_refs': len(refs),
            'images_found': len(images),
            'uploads_successful': uploads_successful,
            'uploads_failed': uploads_failed,
            'message': 'Processamento SSH Porta 65002 concluído'
        }
        
    except Exception as e:
        error_msg = str(e)
        logger.error(f"❌ Erro no processamento: {error_msg}")
        debug_info.append(f"❌ Erro no processamento: {error_msg}")
        
        return {
            'success': False,
            'error': error_msg,
            'total_refs': 0,
            'images_found': 0,
            'uploads_successful': 0,
            'uploads_failed': 1
        }

def upload_image_to_ssh(image, ref_value, debug_info):
    """Upload da imagem para SSH"""
    try:
        logger.info(f"🚀 Upload SSH: {ref_value}")
        debug_info.append(f"🚀 Upload SSH: {ref_value}")
        
        # Salva imagem temporariamente
        temp_image_path = save_image_to_temp(image, ref_value, debug_info)
        
        # Conecta ao SSH
        debug_info.append(f"🔌 Conectando ao SSH: {SSH_HOST}:{SSH_PORT}")
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(
            SSH_HOST, 
            SSH_PORT, 
            SSH_USER, 
            SSH_PASS,
            timeout=30
        )
        sftp = ssh.open_sftp()
        debug_info.append("✅ Conectado ao SSH")
        
        # Cria diretórios
        try:
            sftp.mkdir('/public_html')
            debug_info.append("📁 Diretório /public_html criado")
        except:
            debug_info.append("📁 Diretório /public_html já existe")
        
        try:
            sftp.mkdir('/public_html/images')
            debug_info.append("📁 Diretório /public_html/images criado")
        except:
            debug_info.append("📁 Diretório /public_html/images já existe")
        
        try:
            sftp.mkdir('/public_html/images/products')
            debug_info.append("📁 Diretório /public_html/images/products criado")
        except:
            debug_info.append("📁 Diretório /public_html/images/products já existe")
        
        # Upload
        remote_path = f'/public_html/images/products/{ref_value}.jpg'
        debug_info.append(f"⬆️ Fazendo upload: {remote_path}")
        sftp.put(temp_image_path, remote_path)
        
        sftp.close()
        ssh.close()
        os.remove(temp_image_path)
        debug_info.append("✅ Upload concluído e arquivo temporário removido")
        
        return True
        
    except Exception as e:
        logger.error(f"❌ Erro SSH: {e}")
        debug_info.append(f"❌ Erro SSH: {e}")
        return False

def save_image_to_temp(image, ref_value, debug_info):
    """Salva imagem temporariamente"""
    try:
        safe_ref = "".join(c for c in ref_value if c.isalnum() or c in ('-', '_')).rstrip()
        if not safe_ref:
            safe_ref = f"image_{int(time.time())}"
        
        temp_path = os.path.join('/tmp', f"{safe_ref}.jpg")
        debug_info.append(f"💾 Salvando imagem em: {temp_path}")
        
        # Tenta extrair dados da imagem
        image_data = None
        
        if hasattr(image, '_data') and callable(image._data):
            try:
                image_data = image._data()
                debug_info.append("📊 Dados extraídos via _data()")
            except Exception as e:
                debug_info.append(f"❌ Erro _data(): {e}")
                pass
        
        if not image_data and hasattr(image, 'ref'):
            try:
                image_data = image.ref.read()
                debug_info.append("📊 Dados extraídos via ref.read()")
            except Exception as e:
                debug_info.append(f"❌ Erro ref.read(): {e}")
                pass
        
        if not image_data and hasattr(image, 'data'):
            try:
                image_data = image.data
                debug_info.append("📊 Dados extraídos via data")
            except Exception as e:
                debug_info.append(f"❌ Erro data: {e}")
                pass
        
        if not image_data:
            try:
                image.save(temp_path)
                debug_info.append("📊 Imagem salva via save()")
                return temp_path
            except Exception as e:
                debug_info.append(f"❌ Erro save(): {e}")
                pass
        
        if image_data:
            with open(temp_path, 'wb') as f:
                f.write(image_data)
            debug_info.append("📊 Imagem salva via dados extraídos")
            return temp_path
        
        raise Exception("Não foi possível extrair dados da imagem")
        
    except Exception as e:
        logger.error(f"❌ Erro ao salvar imagem: {e}")
        debug_info.append(f"❌ Erro ao salvar imagem: {e}")
        raise

# Para PythonAnywhere
app = create_app()

if __name__ == "__main__":
    port = int(os.getenv('PORT', 8080))
    
    logger.info("🚀 Iniciando Sistema SSH Porta 65002...")
    print("🚀 Iniciando Sistema SSH Porta 65002...")
    print(f"📁 Frontend: http://localhost:{port}")
    print(f"💚 Health: http://localhost:{port}/health")
    print("=" * 50)
    print("🔧 Dependências:")
    print(f"   • SSH: {'✅' if SSH_AVAILABLE else '❌'}")
    print(f"   • openpyxl: {'✅' if OPENPYXL_AVAILABLE else '❌'}")
    print(f"   • Flask: {'✅' if FLASK_AVAILABLE else '❌'}")
    print(f"   • Host: {SSH_HOST}")
    print(f"   • Porta: {SSH_PORT}")
    print(f"   • Usuário: {SSH_USER}")
    print("=" * 50)
    
    if not FLASK_AVAILABLE:
        logger.error("❌ Flask não disponível")
        sys.exit(1)
    
    try:
        logger.info(f"✅ Servidor iniciado na porta {port}")
        print(f"✅ Servidor iniciado na porta {port}")
        app.run(host='0.0.0.0', port=port, debug=False, threaded=True)
    except Exception as e:
        logger.error(f"❌ Erro ao iniciar servidor: {e}")
        print(f"❌ Erro ao iniciar servidor: {e}")
        sys.exit(1)
