#!/usr/bin/env python3
"""
🔍 INVESTIGADOR AVANÇADO DE IMAGENS EXCEL
Testa diferentes métodos de detecção de imagens para identificar o problema
"""

import openpyxl
import openpyxl.utils
import os
import logging
from typing import List, Dict, Optional

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def investigar_imagens_avancado(arquivo):
    """Investiga imagens usando múltiplos métodos"""
    
    if not os.path.exists(arquivo):
        print(f"❌ Arquivo {arquivo} não existe!")
        return
    
    print(f"\n🔍 INVESTIGAÇÃO AVANÇADA DE IMAGENS")
    print(f"📁 Arquivo: {arquivo}")
    print("=" * 70)
    
    try:
        # Carrega o arquivo
        workbook = openpyxl.load_workbook(arquivo)
        worksheet = workbook.active
        
        print(f"📊 Informações da Planilha:")
        print(f"   • Nome: {worksheet.title}")
        print(f"   • Dimensões: {worksheet.max_row} linhas x {worksheet.max_column} colunas")
        
        # MÉTODO 1: Verificar _images diretamente
        print(f"\n🔍 MÉTODO 1: Verificação direta de _images")
        print(f"   • Total de imagens em worksheet._images: {len(worksheet._images)}")
        
        if len(worksheet._images) > 0:
            print(f"   • Detalhes das imagens:")
            for i, image in enumerate(worksheet._images):
                print(f"     Imagem {i+1}:")
                print(f"       - Tipo: {type(image)}")
                print(f"       - Anchor: {image.anchor}")
                if hasattr(image, 'anchor') and image.anchor:
                    anchor = image.anchor
                    print(f"       - Anchor tipo: {type(anchor)}")
                    if hasattr(anchor, '_from'):
                        print(f"       - _from: {anchor._from}")
                        if anchor._from:
                            print(f"         - col: {anchor._from.col}")
                            print(f"         - row: {anchor._from.row}")
                    if hasattr(anchor, 'col'):
                        print(f"       - col: {anchor.col}")
                    if hasattr(anchor, 'row'):
                        print(f"       - row: {anchor.row}")
        else:
            print(f"   ❌ Nenhuma imagem encontrada em worksheet._images")
        
        # MÉTODO 2: Verificar drawings
        print(f"\n🔍 MÉTODO 2: Verificação de drawings")
        if hasattr(worksheet, '_drawings'):
            print(f"   • Total de drawings: {len(worksheet._drawings) if worksheet._drawings else 0}")
            if worksheet._drawings:
                for i, drawing in enumerate(worksheet._drawings):
                    print(f"     Drawing {i+1}: {type(drawing)}")
                    if hasattr(drawing, '_images'):
                        print(f"       - Imagens no drawing: {len(drawing._images)}")
        else:
            print(f"   ❌ Atributo _drawings não encontrado")
        
        # MÉTODO 3: Verificar todas as planilhas
        print(f"\n🔍 MÉTODO 3: Verificação de todas as planilhas")
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"   • Planilha '{sheet_name}': {len(sheet._images)} imagens")
            if len(sheet._images) > 0:
                print(f"     Detalhes:")
                for i, image in enumerate(sheet._images):
                    anchor = image.anchor
                    if anchor:
                        if hasattr(anchor, '_from') and anchor._from:
                            col_idx = anchor._from.col
                            row_idx = anchor._from.row + 1
                            col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                            print(f"       - Imagem {i+1}: {col_letter}{row_idx}")
                        elif hasattr(anchor, 'col') and hasattr(anchor, 'row'):
                            col_idx = anchor.col
                            row_idx = anchor.row + 1
                            col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                            print(f"       - Imagem {i+1}: {col_letter}{row_idx}")
        
        # MÉTODO 4: Verificar células específicas da coluna H
        print(f"\n🔍 MÉTODO 4: Verificação de células da coluna H")
        for row in range(1, min(worksheet.max_row + 1, 20)):
            cell_h = worksheet[f'H{row}']
            cell_a = worksheet[f'A{row}']
            
            # Verificar se a célula tem conteúdo
            if cell_h.value or cell_a.value:
                print(f"   Linha {row}:")
                print(f"     - H{row}: {cell_h.value}")
                print(f"     - A{row}: {cell_a.value}")
                
                # Verificar se há imagens associadas à célula
                if hasattr(cell_h, '_image') and cell_h._image:
                    print(f"     - ✅ Imagem encontrada na célula H{row}")
                else:
                    print(f"     - ❌ Nenhuma imagem na célula H{row}")
        
        # MÉTODO 5: Verificar workbook._images
        print(f"\n🔍 MÉTODO 5: Verificação de workbook._images")
        if hasattr(workbook, '_images'):
            print(f"   • Total de imagens no workbook: {len(workbook._images) if workbook._images else 0}")
        else:
            print(f"   ❌ Atributo _images não encontrado no workbook")
        
        # MÉTODO 6: Verificar workbook._drawings
        print(f"\n🔍 MÉTODO 6: Verificação de workbook._drawings")
        if hasattr(workbook, '_drawings'):
            print(f"   • Total de drawings no workbook: {len(workbook._drawings) if workbook._drawings else 0}")
        else:
            print(f"   ❌ Atributo _drawings não encontrado no workbook")
        
        # MÉTODO 7: Verificar atributos ocultos
        print(f"\n🔍 MÉTODO 7: Verificação de atributos ocultos")
        print(f"   • Atributos da planilha:")
        for attr in dir(worksheet):
            if 'image' in attr.lower() or 'draw' in attr.lower():
                try:
                    value = getattr(worksheet, attr)
                    if not callable(value):
                        print(f"     - {attr}: {value}")
                except:
                    pass
        
        workbook.close()
        
        # RESULTADO FINAL
        print(f"\n{'='*70}")
        print("📋 DIAGNÓSTICO FINAL:")
        print("=" * 70)
        
        if len(worksheet._images) == 0:
            print("❌ PROBLEMA IDENTIFICADO:")
            print("   • O arquivo não contém imagens inseridas como objetos Excel")
            print("   • As imagens podem estar:")
            print("     1. Inseridas como objetos externos")
            print("     2. Em formato não suportado pelo openpyxl")
            print("     3. Corrompidas durante o upload")
            print("     4. Em uma planilha diferente")
            print()
            print("💡 SOLUÇÕES:")
            print("   1. Re-inserir as imagens diretamente no Excel")
            print("   2. Usar 'Inserir > Imagem > Este Dispositivo'")
            print("   3. Certificar-se que as imagens estão na coluna H")
            print("   4. Salvar como .xlsx (não .xls)")
        else:
            print("✅ IMAGENS ENCONTRADAS!")
            print(f"   • {len(worksheet._images)} imagens detectadas")
            print("   • Problema pode estar na lógica de detecção")
            
    except Exception as e:
        print(f"❌ Erro ao investigar arquivo: {e}")
        import traceback
        traceback.print_exc()

def main():
    """Função principal"""
    
    print("🔍 INVESTIGADOR AVANÇADO DE IMAGENS EXCEL")
    print("=" * 70)
    
    arquivo = "produtos_novos.xlsx"
    investigar_imagens_avancado(arquivo)

if __name__ == "__main__":
    main()

