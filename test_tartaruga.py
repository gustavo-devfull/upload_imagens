#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para testar detecção de imagens no arquivo tartaruga.xlsx
"""

import openpyxl
import os

def test_image_detection():
    """Testa a detecção de imagens no arquivo tartaruga.xlsx"""
    
    file_path = "tartaruga.xlsx"
    
    if not os.path.exists(file_path):
        print(f"❌ Arquivo {file_path} não encontrado!")
        return
    
    try:
        print(f"🔍 Analisando arquivo: {file_path}")
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        
        print(f"📊 Planilha: {worksheet.title}")
        print(f"📏 Dimensões: {worksheet.max_row} linhas x {worksheet.max_column} colunas")
        
        # Verifica imagens inseridas
        print(f"\n🖼️  Total de imagens inseridas: {len(worksheet._images)}")
        for i, image in enumerate(worksheet._images):
            if hasattr(image, 'anchor') and hasattr(image.anchor, '_from'):
                img_row = image.anchor._from.row + 1
                img_col = image.anchor._from.col + 1
                print(f"   Imagem {i+1}: linha {img_row}, coluna {chr(64 + img_col)}")
        
        # Analisa dados das linhas
        print(f"\n📋 Análise das linhas (a partir da linha 4):")
        for row_num in range(4, min(worksheet.max_row + 1, 10)):  # Analisa primeiras 6 linhas
            ref_cell = worksheet[f'A{row_num}']
            photo_cell = worksheet[f'H{row_num}']
            
            print(f"\n   Linha {row_num}:")
            print(f"     REF (A{row_num}): {ref_cell.value}")
            print(f"     Foto (H{row_num}): {photo_cell.value}")
            
            # Verifica células adjacentes
            g_cell = worksheet[f'G{row_num}']
            i_cell = worksheet[f'I{row_num}']
            print(f"     G{row_num}: {g_cell.value}")
            print(f"     I{row_num}: {i_cell.value}")
        
        print(f"\n✅ Análise concluída!")
        
    except Exception as e:
        print(f"❌ Erro ao analisar arquivo: {e}")

if __name__ == "__main__":
    test_image_detection()
