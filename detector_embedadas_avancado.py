#!/usr/bin/env python3
"""
🔍 DETECTOR AVANÇADO DE IMAGENS EMBEDADAS
Tenta diferentes métodos para acessar imagens que estão embedadas no Excel
"""

import openpyxl
import openpyxl.utils
import os
import logging
from typing import List, Dict, Optional
import zipfile
import xml.etree.ElementTree as ET

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def detectar_imagens_embedadas_avancado(arquivo):
    """Detecta imagens usando métodos avançados"""
    
    if not os.path.exists(arquivo):
        print(f"❌ Arquivo {arquivo} não existe!")
        return
    
    print(f"\n🔍 DETECTOR AVANÇADO DE IMAGENS EMBEDADAS")
    print(f"📁 Arquivo: {arquivo}")
    print("=" * 70)
    
    try:
        # MÉTODO 1: Análise direta do arquivo ZIP (Excel é um ZIP)
        print(f"\n🔍 MÉTODO 1: Análise do arquivo como ZIP")
        try:
            with zipfile.ZipFile(arquivo, 'r') as zip_file:
                file_list = zip_file.namelist()
                print(f"   • Total de arquivos no ZIP: {len(file_list)}")
                
                # Procurar por arquivos de imagem
                image_files = [f for f in file_list if any(f.endswith(ext) for ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp'])]
                print(f"   • Arquivos de imagem encontrados: {len(image_files)}")
                if image_files:
                    for img_file in image_files:
                        print(f"     - {img_file}")
                
                # Procurar por drawings
                drawing_files = [f for f in file_list if 'drawing' in f.lower()]
                print(f"   • Arquivos de drawing: {len(drawing_files)}")
                if drawing_files:
                    for draw_file in drawing_files:
                        print(f"     - {draw_file}")
                        
                        # Ler conteúdo do drawing
                        try:
                            content = zip_file.read(draw_file)
                            print(f"       Tamanho: {len(content)} bytes")
                            
                            # Procurar por referências de imagem
                            if b'image' in content.lower():
                                print(f"       ✅ Contém referências a imagens")
                            if b'pic:' in content.lower():
                                print(f"       ✅ Contém elementos de imagem")
                        except Exception as e:
                            print(f"       ❌ Erro ao ler drawing: {e}")
                
        except Exception as e:
            print(f"   ❌ Erro ao analisar como ZIP: {e}")
        
        # MÉTODO 2: Carregar com openpyxl e investigar profundamente
        print(f"\n🔍 MÉTODO 2: Investigação profunda com openpyxl")
        workbook = openpyxl.load_workbook(arquivo)
        worksheet = workbook.active
        
        print(f"   • Planilha: {worksheet.title}")
        print(f"   • Dimensões: {worksheet.max_row} linhas x {worksheet.max_column} colunas")
        
        # Verificar _images
        print(f"   • worksheet._images: {len(worksheet._images)}")
        
        # Verificar _drawing
        if hasattr(worksheet, '_drawing'):
            print(f"   • worksheet._drawing: {worksheet._drawing}")
        
        # Verificar legacy_drawing
        if hasattr(worksheet, 'legacy_drawing'):
            print(f"   • worksheet.legacy_drawing: {worksheet.legacy_drawing}")
        
        # MÉTODO 3: Verificar workbook._images
        print(f"\n🔍 MÉTODO 3: Verificação de workbook._images")
        if hasattr(workbook, '_images'):
            print(f"   • workbook._images: {len(workbook._images) if workbook._images else 0}")
        else:
            print(f"   ❌ workbook não tem atributo _images")
        
        # MÉTODO 4: Verificar workbook._drawings
        print(f"\n🔍 MÉTODO 4: Verificação de workbook._drawings")
        if hasattr(workbook, '_drawings'):
            print(f"   • workbook._drawings: {len(workbook._drawings) if workbook._drawings else 0}")
        else:
            print(f"   ❌ workbook não tem atributo _drawings")
        
        # MÉTODO 5: Verificar todas as planilhas
        print(f"\n🔍 MÉTODO 5: Verificação de todas as planilhas")
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"   • Planilha '{sheet_name}':")
            print(f"     - _images: {len(sheet._images)}")
            if hasattr(sheet, '_drawing'):
                print(f"     - _drawing: {sheet._drawing}")
            if hasattr(sheet, 'legacy_drawing'):
                print(f"     - legacy_drawing: {sheet.legacy_drawing}")
        
        # MÉTODO 6: Verificar células específicas da coluna H
        print(f"\n🔍 MÉTODO 6: Verificação detalhada da coluna H")
        for row in range(1, min(worksheet.max_row + 1, 20)):
            cell_h = worksheet[f'H{row}']
            cell_a = worksheet[f'A{row}']
            
            if cell_h.value or cell_a.value:
                print(f"   Linha {row}:")
                print(f"     - H{row}: '{cell_h.value}'")
                print(f"     - A{row}: '{cell_a.value}'")
                
                # Verificar atributos da célula
                cell_attrs = [attr for attr in dir(cell_h) if not attr.startswith('_') and not callable(getattr(cell_h, attr))]
                print(f"     - Atributos da célula H{row}: {cell_attrs}")
                
                # Verificar se há imagens associadas
                for attr in ['_image', '_images', 'image', 'images']:
                    if hasattr(cell_h, attr):
                        value = getattr(cell_h, attr)
                        print(f"     - {attr}: {value}")
        
        # MÉTODO 7: Verificar workbook._rels
        print(f"\n🔍 MÉTODO 7: Verificação de relacionamentos")
        if hasattr(workbook, '_rels'):
            print(f"   • workbook._rels: {len(workbook._rels) if workbook._rels else 0}")
            if workbook._rels:
                for rel_id, rel in workbook._rels.items():
                    if 'image' in str(rel).lower() or 'drawing' in str(rel).lower():
                        print(f"     - {rel_id}: {rel}")
        
        # MÉTODO 8: Verificar worksheet._rels
        print(f"\n🔍 MÉTODO 8: Verificação de relacionamentos da planilha")
        if hasattr(worksheet, '_rels'):
            print(f"   • worksheet._rels: {len(worksheet._rels) if worksheet._rels else 0}")
            if worksheet._rels:
                for rel_id, rel in worksheet._rels.items():
                    if 'image' in str(rel).lower() or 'drawing' in str(rel).lower():
                        print(f"     - {rel_id}: {rel}")
        
        # MÉTODO 9: Verificar workbook._archive
        print(f"\n🔍 MÉTODO 9: Verificação do arquivo interno")
        if hasattr(workbook, '_archive'):
            print(f"   • workbook._archive: {type(workbook._archive)}")
            if workbook._archive:
                try:
                    file_list = workbook._archive.namelist()
                    print(f"   • Arquivos no archive: {len(file_list)}")
                    
                    # Procurar por imagens
                    image_files = [f for f in file_list if any(f.endswith(ext) for ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp'])]
                    if image_files:
                        print(f"   • Imagens no archive: {image_files}")
                except Exception as e:
                    print(f"   ❌ Erro ao acessar archive: {e}")
        
        workbook.close()
        
        # MÉTODO 10: Análise XML direta
        print(f"\n🔍 MÉTODO 10: Análise XML direta")
        try:
            with zipfile.ZipFile(arquivo, 'r') as zip_file:
                # Procurar por worksheet XML
                worksheet_files = [f for f in zip_file.namelist() if 'worksheet' in f and f.endswith('.xml')]
                print(f"   • Arquivos worksheet XML: {len(worksheet_files)}")
                
                for ws_file in worksheet_files:
                    try:
                        content = zip_file.read(ws_file)
                        print(f"   • Analisando {ws_file}")
                        
                        # Procurar por elementos de imagem
                        if b'drawing' in content:
                            print(f"     ✅ Contém elementos de drawing")
                        if b'pic:' in content:
                            print(f"     ✅ Contém elementos de imagem")
                        if b'image' in content:
                            print(f"     ✅ Contém referências de imagem")
                            
                        # Tentar parsear XML
                        try:
                            root = ET.fromstring(content)
                            # Procurar por elementos de drawing
                            drawings = root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}drawing')
                            if drawings:
                                print(f"     ✅ Encontrados {len(drawings)} elementos de drawing")
                        except Exception as e:
                            print(f"     ❌ Erro ao parsear XML: {e}")
                            
                    except Exception as e:
                        print(f"   ❌ Erro ao ler {ws_file}: {e}")
                        
        except Exception as e:
            print(f"   ❌ Erro na análise XML: {e}")
        
        # RESULTADO FINAL
        print(f"\n{'='*70}")
        print("📋 DIAGNÓSTICO FINAL:")
        print("=" * 70)
        
        # Recarregar para verificar novamente
        workbook = openpyxl.load_workbook(arquivo)
        worksheet = workbook.active
        
        total_images = len(worksheet._images)
        print(f"📊 Total de imagens detectadas: {total_images}")
        
        if total_images > 0:
            print("✅ IMAGENS ENCONTRADAS!")
            print("   • O sistema detectou imagens embedadas")
            print("   • Problema pode estar na lógica de posicionamento")
        else:
            print("❌ NENHUMA IMAGEM DETECTADA")
            print("   • As imagens podem estar em formato não suportado")
            print("   • Ou podem estar corrompidas")
            print("   • Ou podem estar em uma planilha diferente")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Erro geral: {e}")
        import traceback
        traceback.print_exc()

def main():
    """Função principal"""
    
    print("🔍 DETECTOR AVANÇADO DE IMAGENS EMBEDADAS")
    print("=" * 70)
    
    arquivo = "produtos_novos.xlsx"
    detectar_imagens_embedadas_avancado(arquivo)

if __name__ == "__main__":
    main()

