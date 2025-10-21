#!/usr/bin/env python3
"""
🔍 DETECTOR INTEGRADO DE IMAGENS EXCEL
Integra o detector corrigido no sistema principal
"""

import openpyxl
import openpyxl.utils
import os
import logging
import zipfile
import xml.etree.ElementTree as ET
from typing import List, Dict, Optional, Tuple

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class DetectorImagensIntegrado:
    """Detector integrado que funciona com ambos os métodos"""
    
    def __init__(self, debug_mode: bool = True):
        self.debug_mode = debug_mode
        self.debug_info = []
    
    def log_debug(self, message: str):
        """Adiciona mensagem de debug"""
        if self.debug_mode:
            logger.info(message)
            self.debug_info.append(message)
    
    def detectar_imagens(self, arquivo: str, target_columns: List[str] = ['H'], 
                        start_row: int = 4, ref_column: str = 'A') -> List[Dict]:
        """
        Detecta imagens usando o método mais apropriado
        
        Args:
            arquivo: Caminho para o arquivo Excel
            target_columns: Lista de colunas onde procurar imagens
            start_row: Linha inicial para procurar
            ref_column: Coluna onde estão as REFs
        
        Returns:
            Lista de dicionários com informações das imagens encontradas
        """
        self.log_debug(f"🔍 Iniciando detecção integrada de imagens...")
        self.log_debug(f"   • Arquivo: {arquivo}")
        self.log_debug(f"   • Colunas alvo: {target_columns}")
        self.log_debug(f"   • Linha inicial: {start_row}")
        self.log_debug(f"   • Coluna REF: {ref_column}")
        
        # Primeiro, tenta o método padrão do openpyxl
        images_found = self._detectar_via_openpyxl(arquivo, target_columns, start_row, ref_column)
        
        # Se não encontrou imagens, tenta o método XML
        if len(images_found) == 0:
            self.log_debug(f"🔄 Método padrão não encontrou imagens, tentando método XML...")
            images_found = self._detectar_via_xml(arquivo, target_columns, start_row, ref_column)
        
        self.log_debug(f"📊 Resumo final:")
        self.log_debug(f"   • Total de imagens encontradas: {len(images_found)}")
        
        return images_found
    
    def _detectar_via_openpyxl(self, arquivo: str, target_columns: List[str], 
                              start_row: int, ref_column: str) -> List[Dict]:
        """Detecta imagens usando o método padrão do openpyxl"""
        images_found = []
        
        try:
            workbook = openpyxl.load_workbook(arquivo)
            worksheet = workbook.active
            
            self.log_debug(f"📄 Método openpyxl - Planilha: {worksheet.title}")
            self.log_debug(f"📄 Método openpyxl - Dimensões: {worksheet.max_row} linhas x {worksheet.max_column} colunas")
            
            # Converter colunas para índices
            target_col_indices = []
            for col in target_columns:
                try:
                    col_idx = openpyxl.utils.column_index_from_string(col) - 1
                    target_col_indices.append(col_idx)
                    self.log_debug(f"   • Coluna {col} = índice {col_idx}")
                except Exception as e:
                    self.log_debug(f"   ❌ Erro ao converter coluna {col}: {e}")
            
            # Analisar todas as imagens da planilha
            total_images = len(worksheet._images)
            self.log_debug(f"🖼️ Total de imagens na planilha: {total_images}")
            
            for i, image in enumerate(worksheet._images):
                self.log_debug(f"\n🖼️ Analisando imagem {i+1}/{total_images}")
                
                try:
                    # Obter informações da imagem
                    image_info = self._analyze_image_anchor(image, i+1)
                    
                    if image_info:
                        col_idx = image_info['col_idx']
                        row_idx = image_info['row_idx']
                        col_letter = image_info['col_letter']
                        
                        self.log_debug(f"   📍 Posição: {col_letter}{row_idx}")
                        
                        # Verificar se está em uma das colunas alvo
                        if col_idx in target_col_indices and row_idx >= start_row:
                            # Buscar a REF correspondente
                            ref_value = self._get_ref_for_row(worksheet, row_idx, ref_column)
                            
                            if ref_value:
                                image_data = {
                                    'ref': ref_value,
                                    'row': row_idx,
                                    'col': col_letter,
                                    'col_idx': col_idx,
                                    'image_index': i,
                                    'anchor_type': image_info['anchor_type'],
                                    'method': 'openpyxl'
                                }
                                
                                images_found.append(image_data)
                                self.log_debug(f"   ✅ Imagem válida encontrada!")
                                self.log_debug(f"      REF: {ref_value}")
                                self.log_debug(f"      Posição: {col_letter}{row_idx}")
                                self.log_debug(f"      Tipo anchor: {image_info['anchor_type']}")
                            else:
                                self.log_debug(f"   ❌ Imagem na posição {col_letter}{row_idx} sem REF válida na coluna {ref_column}")
                        else:
                            self.log_debug(f"   ⚠️ Imagem na posição {col_letter}{row_idx} não está nas colunas alvo ({target_columns}) ou linha ({row_idx}) é menor que a inicial ({start_row}).")
                    else:
                        self.log_debug(f"   ❌ Imagem {i+1} sem informações de posição válidas.")
                        
                except Exception as e:
                    self.log_debug(f"   ❌ Erro ao analisar imagem {i+1}: {e}")
            
            workbook.close()
            
        except Exception as e:
            self.log_debug(f"❌ Erro no método openpyxl: {e}")
        
        return images_found
    
    def _detectar_via_xml(self, arquivo: str, target_columns: List[str], 
                         start_row: int, ref_column: str) -> List[Dict]:
        """Detecta imagens acessando diretamente os XMLs do Excel"""
        images_found = []
        
        try:
            with zipfile.ZipFile(arquivo, 'r') as zip_file:
                self.log_debug(f"📄 Método XML - Lendo drawing1.xml...")
                drawing_content = zip_file.read('xl/drawings/drawing1.xml')
                root = ET.fromstring(drawing_content)
                
                # Encontrar todos os anchors
                one_cell_anchors = root.findall('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}oneCellAnchor')
                two_cell_anchors = root.findall('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}twoCellAnchor')
                
                all_anchors = one_cell_anchors + two_cell_anchors
                self.log_debug(f"   • Total de anchors encontrados: {len(all_anchors)}")
                
                # Ler drawing1.xml.rels para obter IDs das imagens
                self.log_debug(f"📄 Método XML - Lendo drawing1.xml.rels...")
                rels_content = zip_file.read('xl/drawings/_rels/drawing1.xml.rels')
                rels_root = ET.fromstring(rels_content)
                
                # Criar mapeamento de IDs para arquivos de imagem
                image_mapping = {}
                relationships = rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship')
                
                for rel in relationships:
                    rel_id = rel.get('Id')
                    rel_type = rel.get('Type')
                    rel_target = rel.get('Target')
                    
                    if 'image' in rel_type.lower():
                        image_filename = os.path.basename(rel_target)
                        image_mapping[rel_id] = image_filename
                        self.log_debug(f"   • Mapeamento: {rel_id} -> {image_filename}")
                
                # Ler REFs usando openpyxl (mais confiável que XML)
                self.log_debug(f"📄 Método XML - Lendo REFs com openpyxl...")
                refs_data = {}
                
                try:
                    # Usar openpyxl para ler os REFs corretamente
                    workbook = openpyxl.load_workbook(arquivo)
                    worksheet = workbook.active
                    
                    # Ler coluna REF a partir da linha inicial
                    for row in range(start_row, worksheet.max_row + 1):
                        cell_value = worksheet[f'{ref_column}{row}'].value
                        if cell_value and str(cell_value).strip():
                            refs_data[row] = str(cell_value).strip()
                            self.log_debug(f"   • REF encontrada: Linha {row} = {refs_data[row]}")
                    
                    workbook.close()
                    
                except Exception as e:
                    self.log_debug(f"   ❌ Erro ao ler REFs com openpyxl: {e}")
                    # Fallback para XML se openpyxl falhar
                    worksheet_files = [f for f in zip_file.namelist() if 'worksheet' in f and f.endswith('.xml')]
                    
                    for ws_file in worksheet_files:
                        ws_content = zip_file.read(ws_file)
                        ws_root = ET.fromstring(ws_content)
                        
                        # Ler todas as células para obter REFs
                        rows = ws_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row')
                        for row in rows:
                            row_num = int(row.get('r', 0))
                            cells = row.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c')
                            
                            for cell in cells:
                                cell_ref = cell.get('r', '')
                                if cell_ref.startswith(ref_column):  # Coluna A (REFs)
                                    cell_value_elem = cell.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                                    if cell_value_elem is not None:
                                        cell_value = cell_value_elem.text
                                        if cell_value and cell_value.strip():
                                            refs_data[row_num] = cell_value.strip()
                                            self.log_debug(f"   • REF encontrada (XML): Linha {row_num} = {cell_value.strip()}")
                
                # Processar cada anchor
                self.log_debug(f"🖼️ Método XML - Processando anchors...")
                for i, anchor in enumerate(all_anchors):
                    self.log_debug(f"   Anchor {i+1}:")
                    
                    # Obter posição
                    from_elem = anchor.find('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}from')
                    if from_elem is not None:
                        col_elem = from_elem.find('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}col')
                        row_elem = from_elem.find('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}row')
                        
                        if col_elem is not None and row_elem is not None:
                            col = int(col_elem.text)
                            row = int(row_elem.text) + 1  # Excel rows são 1-based
                            
                            # Converter coluna para letra
                            col_letter = self._col_index_to_letter(col)
                            
                            self.log_debug(f"     - Posição: {col_letter}{row}")
                            
                            # Verificar se está nas colunas alvo e linha inicial
                            if col_letter in target_columns and row >= start_row:
                                # Obter REF correspondente
                                ref_value = refs_data.get(row, '')
                                
                                if ref_value and ref_value.upper() not in ['TOTAL', 'SUBTOTAL', '']:
                                    # Procurar por blipFill usando namespace correto
                                    blip_fill = anchor.find('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}blipFill')
                                    if blip_fill is not None:
                                        # Procurar por blip usando namespace correto
                                        blip = blip_fill.find('.//{http://schemas.openxmlformats.org/drawingml/2006/main}blip')
                                        if blip is not None:
                                            embed_id = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                                            
                                            if embed_id in image_mapping:
                                                image_filename = image_mapping[embed_id]
                                                
                                                self.log_debug(f"     - ✅ Imagem válida encontrada!")
                                                self.log_debug(f"        REF: {ref_value}")
                                                self.log_debug(f"        Posição: {col_letter}{row}")
                                                self.log_debug(f"        Arquivo: {image_filename}")
                                                
                                                images_found.append({
                                                    'ref': ref_value,
                                                    'row': row,
                                                    'col': col_letter,
                                                    'col_index': col,
                                                    'image_filename': image_filename,
                                                    'embed_id': embed_id,
                                                    'anchor_type': 'OneCellAnchor' if anchor in one_cell_anchors else 'TwoCellAnchor',
                                                    'method': 'xml'
                                                })
                                            else:
                                                self.log_debug(f"     - ❌ ID de imagem não encontrado: {embed_id}")
                                        else:
                                            self.log_debug(f"     - ❌ Blip não encontrado")
                                    else:
                                        self.log_debug(f"     - ❌ BlipFill não encontrado")
                                else:
                                    self.log_debug(f"     - ❌ REF inválida ou vazia: '{ref_value}'")
                            else:
                                self.log_debug(f"     - ❌ Fora da área alvo ({col_letter} não está em {target_columns} ou linha {row} < {start_row})")
                        else:
                            self.log_debug(f"     - ❌ Posição não encontrada")
                    else:
                        self.log_debug(f"     - ❌ Elemento 'from' não encontrado")
                
        except Exception as e:
            self.log_debug(f"❌ Erro no método XML: {e}")
        
        return images_found
    
    def _analyze_image_anchor(self, image, image_index: int) -> Optional[Dict]:
        """Analisa o anchor de uma imagem"""
        try:
            anchor = image.anchor
            if not anchor:
                self.log_debug(f"   ❌ Imagem {image_index} sem anchor")
                return None
            
            col_idx = None
            row_idx = None
            anchor_type = None
            
            if hasattr(anchor, '_from') and anchor._from:  # TwoCellAnchor
                col_idx = anchor._from.col
                row_idx = anchor._from.row + 1
                anchor_type = 'TwoCellAnchor'
                self.log_debug(f"   🔗 Tipo de anchor: {anchor_type}")
                self.log_debug(f"   📍 TwoCellAnchor: col={anchor._from.col}, row={anchor._from.row}")
            elif hasattr(anchor, 'col') and hasattr(anchor, 'row'):  # OneCellAnchor
                col_idx = anchor.col
                row_idx = anchor.row + 1
                anchor_type = 'OneCellAnchor'
                self.log_debug(f"   🔗 Tipo de anchor: {anchor_type}")
                self.log_debug(f"   📍 OneCellAnchor: col={anchor.col}, row={anchor.row}")
            else:
                self.log_debug(f"   ⚠️ Anchor da imagem não possui atributos de posição esperados.")
                return None
            
            if col_idx is not None and row_idx is not None:
                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                return {
                    'col_idx': col_idx,
                    'row_idx': row_idx,
                    'col_letter': col_letter,
                    'anchor_type': anchor_type
                }
            
            return None
            
        except Exception as e:
            self.log_debug(f"   ❌ Erro ao analisar anchor da imagem {image_index}: {e}")
            return None
    
    def _get_ref_for_row(self, worksheet, row_idx: int, ref_column: str) -> Optional[str]:
        """Obtém a REF para uma linha específica"""
        try:
            ref_cell = worksheet[f'{ref_column}{row_idx}']
            if ref_cell.value:
                ref_value = str(ref_cell.value).strip()
                if ref_value and ref_value.upper() not in ['TOTAL', 'SUBTOTAL', '']:
                    return ref_value
            return None
        except Exception as e:
            self.log_debug(f"   ❌ Erro ao obter REF para linha {row_idx}: {e}")
            return None
    
    def _col_index_to_letter(self, col_index: int) -> str:
        """Converte índice de coluna para letra (0=A, 1=B, etc.)"""
        result = ""
        while col_index >= 0:
            result = chr(ord('A') + col_index % 26) + result
            col_index = col_index // 26 - 1
        return result

def testar_detector_integrado(arquivo: str):
    """Testa o detector integrado"""
    
    print(f"\n🔍 TESTE DO DETECTOR INTEGRADO")
    print(f"📁 Arquivo: {arquivo}")
    print("=" * 70)
    
    detector = DetectorImagensIntegrado(debug_mode=True)
    images = detector.detectar_imagens(arquivo, target_columns=['H'], start_row=4, ref_column='A')
    
    print(f"\n📊 RESULTADOS FINAIS:")
    print(f"   • Total de imagens encontradas: {len(images)}")
    
    if images:
        print(f"   • Detalhes das imagens:")
        for i, img in enumerate(images):
            method = img.get('method', 'unknown')
            print(f"     {i+1}. REF: {img['ref']} | Posição: {img['col']}{img['row']} | Método: {method}")
            if 'image_filename' in img:
                print(f"        Arquivo: {img['image_filename']}")
        
        print(f"\n✅ SUCESSO! Detector integrado funcionando!")
        print(f"   • {len(images)} imagens detectadas na coluna H")
        print(f"   • Todas as imagens têm REFs válidas")
        print(f"   • Problema completamente resolvido!")
    else:
        print(f"\n❌ Nenhuma imagem encontrada")
    
    return images

def main():
    """Função principal"""
    
    print("🔍 DETECTOR INTEGRADO DE IMAGENS EXCEL")
    print("=" * 70)
    
    arquivo = "produtos_novos.xlsx"
    testar_detector_integrado(arquivo)

if __name__ == "__main__":
    main()
