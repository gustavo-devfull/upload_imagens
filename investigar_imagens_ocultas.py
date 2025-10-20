#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Investiga por que as imagens não são detectadas no nova.xlsx
"""

import openpyxl
import os
import zipfile
import xml.etree.ElementTree as ET

def investigar_imagens_ocultas():
    """Investiga por que as imagens não são detectadas"""
    
    print("🔍 Investigando Imagens Ocultas no nova.xlsx")
    print("=" * 50)
    
    arquivo = "nova.xlsx"
    
    try:
        # Extrai arquivos de imagem do Excel
        print(f"📷 Extraindo arquivos de imagem:")
        with zipfile.ZipFile(arquivo, 'r') as zip_file:
            # Lista arquivos de imagem
            image_files = [f for f in zip_file.namelist() if f.startswith('xl/media/')]
            
            for img_file in image_files:
                print(f"   📁 {img_file}")
                
                # Extrai arquivo de imagem
                img_data = zip_file.read(img_file)
                print(f"      Tamanho: {len(img_data)} bytes")
                
                # Verifica cabeçalho
                if len(img_data) >= 8:
                    header = img_data[:8]
                    print(f"      Cabeçalho: {header.hex()}")
                    
                    if header.startswith(b'\xff\xd8'):
                        print(f"      ✅ Formato: JPEG")
                    elif header.startswith(b'\x89PNG'):
                        print(f"      ✅ Formato: PNG")
                    else:
                        print(f"      ❓ Formato: Desconhecido")
                
                # Salva arquivo para análise
                output_file = f"extraida_{os.path.basename(img_file)}"
                with open(output_file, 'wb') as f:
                    f.write(img_data)
                print(f"      💾 Salvo como: {output_file}")
        
        # Analisa arquivos de desenho
        print(f"\n🎨 Analisando arquivos de desenho:")
        with zipfile.ZipFile(arquivo, 'r') as zip_file:
            drawing_files = [f for f in zip_file.namelist() if 'drawing' in f.lower() and f.endswith('.xml')]
            
            for draw_file in drawing_files:
                print(f"   📄 {draw_file}")
                
                # Lê arquivo XML
                xml_data = zip_file.read(draw_file)
                root = ET.fromstring(xml_data)
                
                # Procura por referências a imagens
                for elem in root.iter():
                    if 'image' in elem.tag.lower() or 'pic' in elem.tag.lower():
                        print(f"      🖼️ Elemento de imagem encontrado: {elem.tag}")
                        if elem.attrib:
                            print(f"         Atributos: {elem.attrib}")
        
        # Tenta carregar com openpyxl de forma diferente
        print(f"\n🔧 Tentando diferentes métodos de carregamento:")
        
        # Método 1: read_only=False
        try:
            workbook = openpyxl.load_workbook(arquivo, data_only=False, read_only=False)
            worksheet = workbook.active
            print(f"   Método 1 (read_only=False): {len(worksheet._images)} imagens")
            workbook.close()
        except Exception as e:
            print(f"   Método 1 falhou: {e}")
        
        # Método 2: read_only=True
        try:
            workbook = openpyxl.load_workbook(arquivo, data_only=False, read_only=True)
            worksheet = workbook.active
            print(f"   Método 2 (read_only=True): {len(worksheet._images)} imagens")
            workbook.close()
        except Exception as e:
            print(f"   Método 2 falhou: {e}")
        
        # Método 3: keep_links=True
        try:
            workbook = openpyxl.load_workbook(arquivo, data_only=False, read_only=False, keep_links=True)
            worksheet = workbook.active
            print(f"   Método 3 (keep_links=True): {len(worksheet._images)} imagens")
            workbook.close()
        except Exception as e:
            print(f"   Método 3 falhou: {e}")
        
    except Exception as e:
        print(f"❌ Erro: {e}")
        import traceback
        traceback.print_exc()

def testar_extração_manual():
    """Testa extração manual das imagens"""
    
    print(f"\n🧪 Testando Extração Manual das Imagens:")
    
    try:
        # Extrai imagens manualmente
        with zipfile.ZipFile("nova.xlsx", 'r') as zip_file:
            image_files = [f for f in zip_file.namelist() if f.startswith('xl/media/')]
            
            for i, img_file in enumerate(image_files):
                print(f"\n📷 Imagem {i+1}: {img_file}")
                
                # Extrai dados da imagem
                img_data = zip_file.read(img_file)
                
                # Salva arquivo
                output_file = f"imagem_extraida_{i+1}.jpg"
                with open(output_file, 'wb') as f:
                    f.write(img_data)
                
                print(f"   💾 Salva como: {output_file}")
                print(f"   📊 Tamanho: {len(img_data)} bytes")
                
                # Verifica se é válida
                try:
                    from PIL import Image
                    with Image.open(output_file) as img:
                        print(f"   ✅ Válida: {img.size}, {img.format}")
                except Exception as e:
                    print(f"   ❌ Inválida: {e}")
        
    except Exception as e:
        print(f"❌ Erro: {e}")

def criar_planilha_com_imagens():
    """Cria uma planilha com as imagens extraídas"""
    
    print(f"\n🛠️ Criando Planilha com Imagens Extraídas:")
    
    try:
        # Cria nova planilha
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Produtos com Imagens"
        
        # Cabeçalhos
        worksheet['A1'] = 'REF'
        worksheet['B1'] = 'DESCRIPTION'
        worksheet['C1'] = 'NAME'
        worksheet['H1'] = 'PHOTO'
        
        # Dados de exemplo
        refs = ['garra', 'ST002', 'ST003', 'ST004']
        descriptions = ['Novidade de garrafa', 'GARRAFA TERMICA, BUL', 'Produto 3', 'Produto 4']
        names = ['garra', 'ST002', 'ST003', 'ST004']
        
        for i, (ref, desc, name) in enumerate(zip(refs, descriptions, names)):
            row = i + 2
            worksheet[f'A{row}'] = ref
            worksheet[f'B{row}'] = desc
            worksheet[f'C{row}'] = name
        
        # Salva planilha
        output_file = "nova_com_imagens.xlsx"
        workbook.save(output_file)
        print(f"   ✅ Planilha criada: {output_file}")
        
        workbook.close()
        
        print(f"\n💡 PRÓXIMOS PASSOS:")
        print(f"1. Abra {output_file} no Excel")
        print(f"2. Insira as imagens extraídas na coluna H")
        print(f"3. Salve o arquivo")
        print(f"4. Use o sistema de upload com o arquivo modificado")
        
    except Exception as e:
        print(f"❌ Erro: {e}")

if __name__ == "__main__":
    investigar_imagens_ocultas()
    testar_extração_manual()
    criar_planilha_com_imagens()
    
    print(f"\n✅ ANÁLISE CONCLUÍDA!")
    print(f"📁 Arquivos extraídos disponíveis para uso")
    print(f"🛠️ Planilha nova_com_imagens.xlsx criada para inserir imagens")

