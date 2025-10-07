#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de diagnóstico para entender por que não está encontrando imagens
"""

import openpyxl
import os
import logging

# Configuração de logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def diagnosticar_excel(file_path):
    """Diagnostica um arquivo Excel para entender a estrutura"""
    print(f"🔍 Diagnosticando arquivo: {file_path}")
    print("=" * 60)
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        
        print(f"📊 Planilha ativa: {worksheet.title}")
        print(f"📏 Dimensões: {worksheet.max_row} linhas x {worksheet.max_column} colunas")
        print()
        
        # Analisa as primeiras linhas
        print("📋 Primeiras 10 linhas:")
        for row in range(1, min(11, worksheet.max_row + 1)):
            row_data = []
            for col in range(1, min(6, worksheet.max_column + 1)):  # Primeiras 5 colunas
                cell = worksheet.cell(row=row, column=col)
                value = cell.value
                if value is None:
                    value = ""
                else:
                    value = str(value)[:20]  # Limita tamanho
                row_data.append(f"{openpyxl.utils.get_column_letter(col)}{row}: {value}")
            print(f"  Linha {row}: {' | '.join(row_data)}")
        
        print()
        
        # Analisa coluna REF (A)
        print("🔍 Análise da coluna REF (A):")
        refs_validas = []
        refs_invalidas = []
        
        for row in range(4, worksheet.max_row + 1):
            cell_value = worksheet[f'A{row}'].value
            if cell_value:
                ref_str = str(cell_value).strip()
                if ref_str.upper() in ['TOTAL', 'SUBTOTAL', 'SUM', 'COUNT', 'AVERAGE', 'MAX', 'MIN']:
                    refs_invalidas.append((row, ref_str))
                else:
                    refs_validas.append((row, ref_str))
        
        print(f"  ✅ REFs válidas encontradas: {len(refs_validas)}")
        for row, ref in refs_validas[:5]:  # Mostra apenas as primeiras 5
            print(f"    Linha {row}: {ref}")
        if len(refs_validas) > 5:
            print(f"    ... e mais {len(refs_validas) - 5} REFs")
        
        print(f"  ❌ REFs inválidas encontradas: {len(refs_invalidas)}")
        for row, ref in refs_invalidas[:3]:  # Mostra apenas as primeiras 3
            print(f"    Linha {row}: {ref}")
        
        print()
        
        # Analisa imagens
        print("🖼️ Análise de imagens:")
        print(f"  Total de imagens na planilha: {len(worksheet._images)}")
        
        if len(worksheet._images) == 0:
            print("  ❌ Nenhuma imagem encontrada!")
            print("  💡 Possíveis causas:")
            print("    - As imagens não estão inseridas como objetos de imagem")
            print("    - As imagens estão em formato não suportado")
            print("    - As imagens estão em outras planilhas")
        else:
            print("  📍 Detalhes das imagens:")
            for i, image in enumerate(worksheet._images):
                anchor = image.anchor
                print(f"    Imagem {i+1}:")
                
                # Tenta determinar posição
                row = None
                col = None
                
                if hasattr(anchor, '_from') and hasattr(anchor._from, 'row') and hasattr(anchor._from, 'col'):
                    row = anchor._from.row + 1
                    col = anchor._from.col + 1
                elif hasattr(anchor, '_from') and isinstance(anchor._from, tuple) and len(anchor._from) >= 2:
                    row = anchor._from[0] + 1
                    col = anchor._from[1] + 1
                elif hasattr(anchor, 'row') and hasattr(anchor, 'col'):
                    row = anchor.row + 1
                    col = anchor.col + 1
                
                if row and col:
                    col_letter = openpyxl.utils.get_column_letter(col)
                    print(f"      Posição: {col_letter}{row}")
                    print(f"      Tipo: {type(image)}")
                    
                    # Verifica se está na coluna H (PHOTO)
                    if col == 8:  # H = 8
                        print(f"      ✅ Está na coluna PHOTO (H)")
                    else:
                        print(f"      ❌ Está na coluna {col_letter}, não na PHOTO (H)")
                    
                    # Verifica se está a partir da linha 4
                    if row >= 4:
                        print(f"      ✅ Está a partir da linha 4")
                    else:
                        print(f"      ❌ Está antes da linha 4")
                else:
                    print(f"      ❌ Não foi possível determinar a posição")
        
        print()
        
        # Analisa coluna PHOTO (H)
        print("📸 Análise da coluna PHOTO (H):")
        photo_data = []
        for row in range(4, min(15, worksheet.max_row + 1)):  # Primeiras 10 linhas a partir da linha 4
            cell_value = worksheet[f'H{row}'].value
            if cell_value:
                photo_data.append((row, str(cell_value)[:50]))
        
        if photo_data:
            print(f"  Dados encontrados na coluna H:")
            for row, value in photo_data:
                print(f"    Linha {row}: {value}")
        else:
            print("  ❌ Nenhum dado encontrado na coluna H")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Erro ao diagnosticar arquivo: {e}")
        logger.error(f"Erro no diagnóstico: {e}")

def main():
    """Função principal"""
    # Procura por arquivos Excel no diretório atual
    excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
    
    if not excel_files:
        print("❌ Nenhum arquivo Excel encontrado no diretório atual")
        return
    
    print(f"📁 Arquivos Excel encontrados: {excel_files}")
    print()
    
    # Diagnostica cada arquivo
    for excel_file in excel_files:
        diagnosticar_excel(excel_file)
        print("\n" + "="*60 + "\n")

if __name__ == "__main__":
    main()
