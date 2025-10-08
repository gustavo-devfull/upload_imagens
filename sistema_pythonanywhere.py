#!/usr/bin/env python3
"""
Sistema de Upload de Imagens Excel - Vers√£o PythonAnywhere
Vers√£o otimizada para PythonAnywhere com configura√ß√µes espec√≠ficas
"""

import os
import sys
import json
import tempfile
import time
import logging
from datetime import datetime

# Configura√ß√£o de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Imports condicionais
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.error("‚ùå openpyxl n√£o dispon√≠vel")

try:
    import ftplib
    FTP_AVAILABLE = True
except ImportError:
    FTP_AVAILABLE = False
    logger.error("‚ùå ftplib n√£o dispon√≠vel")

try:
    from flask import Flask, request, jsonify, render_template_string
    FLASK_AVAILABLE = True
except ImportError:
    FLASK_AVAILABLE = False
    logger.error("‚ùå Flask n√£o dispon√≠vel")

# Configura√ß√µes FTP
FTP_HOST = "46.202.90.62"
FTP_USER = "u715606397.ideolog.ia.br"
FTP_PASS = "]X9CC>t~ihWhdzNq"
FTP_DIR = "public_html/images/products"

# HTML Template para PythonAnywhere
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>üöÄ Sistema Upload Excel - PythonAnywhere</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 40px; background: #f5f5f5; }
        .container { max-width: 600px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        .header { text-align: center; margin-bottom: 30px; }
        .header h1 { color: #333; margin-bottom: 10px; }
        .upload-area { text-align: center; margin: 30px 0; }
        .file-input { margin: 20px 0; }
        .btn { background: #007bff; color: white; padding: 12px 24px; border: none; border-radius: 5px; cursor: pointer; font-size: 16px; margin: 10px; }
        .btn:hover { background: #0056b3; }
        .btn:disabled { background: #ccc; cursor: not-allowed; }
        .results { margin: 20px 0; padding: 20px; background: #f8f9fa; border-radius: 5px; display: none; }
        .success { background: #d4edda; border: 1px solid #c3e6cb; color: #155724; }
        .error { background: #f8d7da; border: 1px solid #f5c6cb; color: #721c24; }
        .warning { background: #fff3cd; border: 1px solid #ffeaa7; color: #856404; }
        .debug { background: #e2e3e5; border: 1px solid #d6d8db; color: #383d41; font-family: monospace; font-size: 12px; }
        .info { background: #d1ecf1; border: 1px solid #bee5eb; color: #0c5460; }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>üöÄ Sistema Upload Excel</h1>
            <p>Vers√£o PythonAnywhere - Upload de Imagens Autom√°tico</p>
        </div>
        
        <div class="info">
            <h3>üìã Informa√ß√µes do Sistema</h3>
            <p><strong>Plataforma:</strong> PythonAnywhere</p>
            <p><strong>Python:</strong> {{ python_version }}</p>
            <p><strong>Depend√™ncias:</strong> Flask ‚úÖ | openpyxl ‚úÖ | FTP ‚úÖ</p>
            <p><strong>Status:</strong> Sistema funcionando</p>
        </div>
        
        <div class="upload-area">
            <h2>üìÅ Selecione seu arquivo Excel</h2>
            <p>Arquivos suportados: .xlsx com imagens na coluna H</p>
            
            <div class="file-input">
                <input type="file" id="excelFile" accept=".xlsx" />
            </div>
            
            <div id="fileName"></div>
            
            <button id="uploadBtn" class="btn" onclick="uploadFile()" disabled>üöÄ Fazer Upload</button>
            <button id="resetBtn" class="btn" onclick="resetForm()" disabled>üîÑ Resetar</button>
        </div>
        
        <div id="results" class="results"></div>
    </div>

    <script>
        let selectedFile = null;
        
        document.getElementById('excelFile').addEventListener('change', function(e) {
            const file = e.target.files[0];
            if (file) {
                selectedFile = file;
                document.getElementById('fileName').innerHTML = `üìÑ ${file.name} (${(file.size / 1024 / 1024).toFixed(2)} MB)`;
                document.getElementById('uploadBtn').disabled = false;
                document.getElementById('resetBtn').disabled = false;
            }
        });
        
        function uploadFile() {
            if (!selectedFile) {
                alert('Por favor, selecione um arquivo Excel.');
                return;
            }
            
            const formData = new FormData();
            formData.append('excel_file', selectedFile);
            
            document.getElementById('uploadBtn').disabled = true;
            document.getElementById('resetBtn').disabled = true;
            document.getElementById('results').style.display = 'block';
            document.getElementById('results').innerHTML = '<div class="debug">üîÑ Processando arquivo...</div>';
            
            fetch('/upload', {
                method: 'POST',
                body: formData
            })
            .then(response => response.json())
            .then(data => {
                showResults(data);
            })
            .catch(error => {
                document.getElementById('results').innerHTML = `<div class="error"><h3>‚ùå Erro no Upload</h3><p>${error.message}</p></div>`;
                document.getElementById('uploadBtn').disabled = false;
                document.getElementById('resetBtn').disabled = false;
            });
        }
        
        function showResults(data) {
            const resultsDiv = document.getElementById('results');
            
            let html = '';
            
            if (data.debug_info) {
                html += `<div class="debug"><h3>üîç Debug Info</h3><pre>${data.debug_info}</pre></div>`;
            }
            
            if (data.success && data.uploads_successful > 0) {
                html += `<div class="success"><h3>‚úÖ Upload Conclu√≠do!</h3><p>${data.uploads_successful} imagem(ns) enviada(s)</p></div>`;
            } else if (data.error) {
                html += `<div class="warning"><h3>‚ö†Ô∏è Problema</h3><p>${data.error}</p></div>`;
            } else {
                html += `<div class="error"><h3>‚ùå Erro</h3><p>N√£o foi poss√≠vel processar o arquivo</p></div>`;
            }
            
            resultsDiv.innerHTML = html;
            document.getElementById('uploadBtn').disabled = false;
            document.getElementById('resetBtn').disabled = false;
        }
        
        function resetForm() {
            document.getElementById('excelFile').value = '';
            document.getElementById('fileName').innerHTML = '';
            document.getElementById('results').style.display = 'none';
            document.getElementById('uploadBtn').disabled = true;
            document.getElementById('resetBtn').disabled = true;
            selectedFile = null;
        }
    </script>
</body>
</html>
"""

def create_app():
    """Cria aplica√ß√£o Flask para PythonAnywhere"""
    app = Flask(__name__)
    
    @app.route('/')
    def index():
        python_version = f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"
        return render_template_string(HTML_TEMPLATE, python_version=python_version)
    
    @app.route('/health')
    def health():
        """Health check para PythonAnywhere"""
        try:
            health_data = {
                "status": "ok",
                "message": "Sistema PythonAnywhere funcionando",
                "timestamp": time.time(),
                "platform": "PythonAnywhere",
                "python_version": f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}",
                "dependencies": {
                    "ftp": FTP_AVAILABLE,
                    "openpyxl": OPENPYXL_AVAILABLE,
                    "flask": FLASK_AVAILABLE
                }
            }
            
            # Se depend√™ncias cr√≠ticas n√£o est√£o dispon√≠veis, retorna erro
            if not FLASK_AVAILABLE:
                health_data["status"] = "error"
                health_data["message"] = "Flask n√£o dispon√≠vel"
                return jsonify(health_data), 500
            
            return jsonify(health_data), 200
            
        except Exception as e:
            logger.error(f"‚ùå Erro no health check: {e}")
            return jsonify({
                "status": "error",
                "message": f"Erro no health check: {str(e)}",
                "timestamp": time.time(),
                "platform": "PythonAnywhere"
            }), 500
    
    @app.route('/upload', methods=['POST'])
    def upload():
        """Upload de arquivo Excel"""
        debug_info = []
        
        try:
            logger.info("=== IN√çCIO DO UPLOAD PYTHONANYWHERE ===")
            debug_info.append("=== IN√çCIO DO UPLOAD PYTHONANYWHERE ===")
            
            if not FTP_AVAILABLE or not OPENPYXL_AVAILABLE:
                error_msg = 'Depend√™ncias n√£o dispon√≠veis'
                logger.error(error_msg)
                debug_info.append(f"‚ùå {error_msg}")
                return jsonify({
                    'error': error_msg,
                    'debug_info': '\n'.join(debug_info)
                })
            
            if 'excel_file' not in request.files:
                error_msg = 'Arquivo n√£o encontrado'
                logger.error(error_msg)
                debug_info.append(f"‚ùå {error_msg}")
                return jsonify({
                    'error': error_msg,
                    'debug_info': '\n'.join(debug_info)
                })
            
            file = request.files['excel_file']
            logger.info(f"üìÅ Arquivo recebido: {file.filename}")
            debug_info.append(f"üìÅ Arquivo recebido: {file.filename}")
            
            if file.filename == '':
                error_msg = 'Nenhum arquivo selecionado'
                logger.error(error_msg)
                debug_info.append(f"‚ùå {error_msg}")
                return jsonify({
                    'error': error_msg,
                    'debug_info': '\n'.join(debug_info)
                })
            
            # Salva arquivo temporariamente (PythonAnywhere usa /tmp)
            temp_path = os.path.join('/tmp', file.filename)
            file.save(temp_path)
            logger.info(f"üíæ Arquivo salvo em: {temp_path}")
            debug_info.append(f"üíæ Arquivo salvo em: {temp_path}")
            
            # Processamento
            stats = process_excel(temp_path, debug_info)
            
            # Remove arquivo tempor√°rio
            os.remove(temp_path)
            
            # Adiciona debug info
            stats['debug_info'] = '\n'.join(debug_info)
            return jsonify(stats)
            
        except Exception as e:
            error_msg = str(e)
            logger.error(f"‚ùå Erro geral: {error_msg}")
            debug_info.append(f"‚ùå Erro geral: {error_msg}")
            
            return jsonify({
                'error': error_msg,
                'debug_info': '\n'.join(debug_info)
            })
    
    return app

def process_excel(file_path, debug_info):
    """Processamento do Excel"""
    try:
        logger.info(f"üìä Carregando arquivo: {file_path}")
        debug_info.append(f"üìä Carregando arquivo: {file_path}")
        
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        logger.info(f"üìã Planilha: {worksheet.title}")
        debug_info.append(f"üìã Planilha: {worksheet.title}")
        
        # Processa REFs
        refs = []
        for row in range(4, worksheet.max_row + 1):
            cell_value = worksheet[f'A{row}'].value
            if cell_value and str(cell_value).strip() and str(cell_value).upper() not in ['TOTAL', 'SUBTOTAL', '']:
                refs.append(str(cell_value).strip())
        
        logger.info(f"üìä REFs encontradas: {len(refs)}")
        debug_info.append(f"üìä REFs encontradas: {len(refs)}")
        
        # Processa imagens
        images = []
        total_images = len(worksheet._images)
        logger.info(f"üñºÔ∏è Total de imagens: {total_images}")
        debug_info.append(f"üñºÔ∏è Total de imagens: {total_images}")
        
        for i, image in enumerate(worksheet._images):
            logger.info(f"üñºÔ∏è Analisando imagem {i+1}/{total_images}")
            debug_info.append(f"üñºÔ∏è Analisando imagem {i+1}/{total_images}")
            
            if hasattr(image, 'anchor') and image.anchor:
                anchor = image.anchor
                if hasattr(anchor, '_from') and anchor._from:
                    col_idx = anchor._from.col
                    row_idx = anchor._from.row + 1
                    col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                    
                    logger.info(f"üìç Posi√ß√£o: {col_letter}{row_idx}")
                    debug_info.append(f"üìç Posi√ß√£o: {col_letter}{row_idx}")
                    
                    if col_letter == 'H' and row_idx >= 4:
                        ref_cell = worksheet[f'A{row_idx}']
                        if ref_cell.value:
                            ref_value = str(ref_cell.value).strip()
                            if ref_value and ref_value.upper() not in ['TOTAL', 'SUBTOTAL', '']:
                                logger.info(f"‚úÖ Imagem v√°lida: REF {ref_value}")
                                debug_info.append(f"‚úÖ Imagem v√°lida: REF {ref_value}")
                                images.append({'image': image, 'ref': ref_value})
        
        logger.info(f"üìä Imagens v√°lidas: {len(images)}")
        debug_info.append(f"üìä Imagens v√°lidas: {len(images)}")
        
        # Upload
        uploads_successful = 0
        uploads_failed = 0
        
        for image_data in images:
            try:
                if upload_image_to_ftp(image_data['image'], image_data['ref']):
                    uploads_successful += 1
                    logger.info(f"‚úÖ Upload: {image_data['ref']}")
                    debug_info.append(f"‚úÖ Upload: {image_data['ref']}")
                else:
                    uploads_failed += 1
                    logger.error(f"‚ùå Falha: {image_data['ref']}")
                    debug_info.append(f"‚ùå Falha: {image_data['ref']}")
            except Exception as e:
                uploads_failed += 1
                logger.error(f"‚ùå Erro: {image_data['ref']} - {e}")
                debug_info.append(f"‚ùå Erro: {image_data['ref']} - {e}")
        
        return {
            'success': True,
            'total_refs': len(refs),
            'images_found': len(images),
            'uploads_successful': uploads_successful,
            'uploads_failed': uploads_failed,
            'message': 'Processamento PythonAnywhere conclu√≠do'
        }
        
    except Exception as e:
        error_msg = str(e)
        logger.error(f"‚ùå Erro no processamento: {error_msg}")
        debug_info.append(f"‚ùå Erro no processamento: {error_msg}")
        
        return {
            'success': False,
            'error': error_msg,
            'total_refs': 0,
            'images_found': 0,
            'uploads_successful': 0,
            'uploads_failed': 1
        }

def upload_image_to_ftp(image, ref_value):
    """Upload da imagem para FTP"""
    try:
        logger.info(f"üöÄ Upload FTP: {ref_value}")
        
        # Salva imagem temporariamente (PythonAnywhere usa /tmp)
        temp_image_path = save_image_to_temp(image, ref_value)
        
        # Conecta ao FTP
        ftp = ftplib.FTP()
        ftp.connect(FTP_HOST, 21, timeout=300)
        ftp.login(FTP_USER, FTP_PASS)
        
        # Cria diret√≥rios
        try:
            ftp.cwd('public_html')
        except:
            ftp.mkd('public_html')
            ftp.cwd('public_html')
        
        try:
            ftp.cwd('images')
        except:
            ftp.mkd('images')
            ftp.cwd('images')
        
        try:
            ftp.cwd('products')
        except:
            ftp.mkd('products')
            ftp.cwd('products')
        
        # Upload
        with open(temp_image_path, 'rb') as file:
            ftp.storbinary(f'STOR {ref_value}.jpg', file)
        
        ftp.quit()
        os.remove(temp_image_path)
        
        return True
        
    except Exception as e:
        logger.error(f"‚ùå Erro FTP: {e}")
        return False

def save_image_to_temp(image, ref_value):
    """Salva imagem temporariamente"""
    try:
        safe_ref = "".join(c for c in ref_value if c.isalnum() or c in ('-', '_')).rstrip()
        if not safe_ref:
            safe_ref = f"image_{int(time.time())}"
        
        # PythonAnywhere usa /tmp
        temp_path = os.path.join('/tmp', f"{safe_ref}.jpg")
        
        # Tenta extrair dados da imagem
        image_data = None
        
        if hasattr(image, '_data') and callable(image._data):
            try:
                image_data = image._data()
            except:
                pass
        
        if not image_data and hasattr(image, 'ref'):
            try:
                image_data = image.ref.read()
            except:
                pass
        
        if not image_data and hasattr(image, 'data'):
            try:
                image_data = image.data
            except:
                pass
        
        if not image_data:
            try:
                image.save(temp_path)
                return temp_path
            except:
                pass
        
        if image_data:
            with open(temp_path, 'wb') as f:
                f.write(image_data)
            return temp_path
        
        raise Exception("N√£o foi poss√≠vel extrair dados da imagem")
        
    except Exception as e:
        logger.error(f"‚ùå Erro ao salvar imagem: {e}")
        raise

# Para PythonAnywhere, n√£o precisamos de main() pois eles gerenciam o servidor
app = create_app()

if __name__ == "__main__":
    # Apenas para teste local
    port = int(os.getenv('PORT', 8080))
    
    logger.info("üöÄ Iniciando Sistema PythonAnywhere...")
    print("üöÄ Iniciando Sistema PythonAnywhere...")
    print(f"üìÅ Frontend: http://localhost:{port}")
    print(f"üíö Health: http://localhost:{port}/health")
    print("=" * 50)
    print("üîß Depend√™ncias:")
    print(f"   ‚Ä¢ FTP: {'‚úÖ' if FTP_AVAILABLE else '‚ùå'}")
    print(f"   ‚Ä¢ openpyxl: {'‚úÖ' if OPENPYXL_AVAILABLE else '‚ùå'}")
    print(f"   ‚Ä¢ Flask: {'‚úÖ' if FLASK_AVAILABLE else '‚ùå'}")
    print("=" * 50)
    
    if not FLASK_AVAILABLE:
        logger.error("‚ùå Flask n√£o dispon√≠vel")
        sys.exit(1)
    
    try:
        logger.info(f"‚úÖ Servidor iniciado na porta {port}")
        print(f"‚úÖ Servidor iniciado na porta {port}")
        app.run(host='0.0.0.0', port=port, debug=False, threaded=True)
    except Exception as e:
        logger.error(f"‚ùå Erro ao iniciar servidor: {e}")
        print(f"‚ùå Erro ao iniciar servidor: {e}")
        sys.exit(1)
