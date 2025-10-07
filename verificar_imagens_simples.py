#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script simples para verificar imagens
"""

import openpyxl
import os

def verificar_imagens_simples(file_path):
    """Verificação simples das imagens"""
    print(f"🔍 Verificando: {file_path}")
    print("-" * 40)
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        
        print(f"📊 Planilha: {worksheet.title}")
        print(f"📏 Dimensões: {worksheet.max_row} linhas x {worksheet.max_column} colunas")
        
        # Conta imagens
        total_imagens = len(worksheet._images)
        print(f"🖼️ Total de imagens: {total_imagens}")
        
        if total_imagens > 0:
            print("✅ IMAGENS ENCONTRADAS!")
            
            # Analisa cada imagem
            for i, image in enumerate(worksheet._images):
                print(f"\n🖼️ Imagem {i+1}:")
                print(f"  Tipo: {type(image)}")
                
                # Tenta obter posição
                anchor = image.anchor
                print(f"  Anchor: {anchor}")
                
                # Método simples para obter posição
                try:
                    if hasattr(anchor, '_from'):
                        if hasattr(anchor._from, 'row') and hasattr(anchor._from, 'col'):
                            row = anchor._from.row + 1
                            col = anchor._from.col + 1
                            col_letter = openpyxl.utils.get_column_letter(col)
                            print(f"  📍 Posição: {col_letter}{row}")
                            
                            # Verifica se está na coluna H
                            if col == 8:
                                print(f"  ✅ Está na coluna H (PHOTO)")
                            else:
                                print(f"  ❌ Está na coluna {col_letter}, não na H")
                            
                            # Verifica se está a partir da linha 4
                            if row >= 4:
                                print(f"  ✅ Está a partir da linha 4")
                            else:
                                print(f"  ❌ Está antes da linha 4")
                            
                            # Verifica REF correspondente
                            ref_value = worksheet[f'A{row}'].value
                            if ref_value:
                                print(f"  🔍 REF: {ref_value}")
                            else:
                                print(f"  ❌ Sem REF na linha {row}")
                        else:
                            print(f"  ❌ Anchor._from não tem row/col")
                    else:
                        print(f"  ❌ Anchor não tem _from")
                        
                except Exception as e:
                    print(f"  ❌ Erro ao obter posição: {e}")
        else:
            print("❌ NENHUMA IMAGEM ENCONTRADA!")
            
            # Verifica outras planilhas
            print(f"\n📋 Verificando outras planilhas:")
            for sheet in workbook.worksheets:
                if sheet != worksheet:
                    print(f"  {sheet.title}: {len(sheet._images)} imagens")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Erro: {e}")

def main():
    """Função principal"""
    excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
    
    for file in excel_files:
        verificar_imagens_simples(file)
        print("\n" + "="*50 + "\n")

if __name__ == "__main__":
    main()
