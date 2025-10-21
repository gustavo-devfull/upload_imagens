#!/usr/bin/env python3
"""
Sistema MELHORADO de processamento Excel com detecção robusta de imagens
Integra a classe ExcelImageDetector melhorada
"""

import openpyxl
import openpyxl.utils
import os
import logging
from typing import List, Dict, Optional
from detector_imagens_melhorado import ExcelImageDetector

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class SistemaExcelMelhorado:
    """Sistema melhorado para processamento de arquivos Excel com imagens"""
    
    def __init__(self, debug_mode: bool = True):
        self.debug_mode = debug_mode
        self.detector = ExcelImageDetector(debug_mode=debug_mode)
        self.debug_info = []
    
    def log_debug(self, message: str):
        """Adiciona mensagem de debug"""
        if self.debug_mode:
            logger.info(message)
            self.debug_info.append(message)
    
    def process_excel_melhorado(self, file_path: str, target_columns: List[str] = ['H'], 
                              start_row: int = 4, ref_column: str = 'A') -> Dict:
        """
        Processa arquivo Excel com detecção melhorada de imagens
        
        Args:
            file_path: Caminho para o arquivo Excel
            target_columns: Colunas onde procurar imagens
            start_row: Linha inicial para processar
            ref_column: Coluna onde estão as REFs
        
        Returns:
            Dicionário com estatísticas do processamento
        """
        self.debug_info = []
        
        try:
            self.log_debug(f"🚀 Iniciando processamento melhorado...")
            self.log_debug(f"📁 Arquivo: {file_path}")
            self.log_debug(f"🎯 Colunas alvo: {target_columns}")
            self.log_debug(f"📍 Linha inicial: {start_row}")
            self.log_debug(f"🔍 Coluna REF: {ref_column}")
            
            # Carrega o arquivo
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")
            
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            
            self.log_debug(f"✅ Arquivo carregado com sucesso")
            self.log_debug(f"📊 Planilha: {worksheet.title}")
            self.log_debug(f"📏 Dimensões: {worksheet.max_row} linhas x {worksheet.max_column} colunas")
            
            # Detecta imagens usando o sistema melhorado
            images_found = self.detector.detect_images_in_worksheet(
                worksheet=worksheet,
                target_columns=target_columns,
                start_row=start_row,
                ref_column=ref_column
            )
            
            self.log_debug(f"🖼️ Imagens detectadas: {len(images_found)}")
            
            # Obtém estatísticas detalhadas
            stats = self.detector.get_detailed_image_info(worksheet)
            
            # Prepara resultado
            resultado = {
                'success': True,
                'file_path': file_path,
                'total_images': stats['total_images'],
                'images_found': len(images_found),
                'images_by_column': stats['images_by_column'],
                'images_with_refs': stats['images_with_refs'],
                'images_without_refs': stats['images_without_refs'],
                'anchor_types': stats['anchor_types'],
                'images_data': images_found,
                'debug_info': self.debug_info,
                'recommendations': self._generate_recommendations(stats, images_found)
            }
            
            workbook.close()
            
            self.log_debug(f"✅ Processamento concluído com sucesso!")
            return resultado
            
        except Exception as e:
            error_msg = f"❌ Erro no processamento: {e}"
            self.log_debug(error_msg)
            
            return {
                'success': False,
                'error': str(e),
                'file_path': file_path,
                'total_images': 0,
                'images_found': 0,
                'debug_info': self.debug_info
            }
    
    def _generate_recommendations(self, stats: Dict, images_found: List) -> List[str]:
        """Gera recomendações baseadas na análise"""
        recommendations = []
        
        if stats['total_images'] == 0:
            recommendations.append("❌ Arquivo não contém imagens. Insira imagens nas células.")
        elif len(images_found) == 0:
            recommendations.append("❌ Nenhuma imagem válida encontrada na coluna H.")
            recommendations.append("💡 Verifique se as imagens estão na coluna H e têm REFs correspondentes.")
        elif len(images_found) < stats['total_images']:
            recommendations.append(f"⚠️ Apenas {len(images_found)} de {stats['total_images']} imagens são válidas.")
            recommendations.append("💡 Verifique se todas as imagens têm REFs correspondentes.")
        else:
            recommendations.append(f"✅ Todas as {len(images_found)} imagens são válidas!")
        
        # Recomendações sobre colunas
        if 'H' not in stats['images_by_column']:
            recommendations.append("💡 Considere mover imagens para a coluna H.")
        
        # Recomendações sobre tipos de anchor
        if stats['anchor_types']:
            anchor_types = list(stats['anchor_types'].keys())
            if len(anchor_types) > 1:
                recommendations.append(f"ℹ️ Arquivo contém múltiplos tipos de anchor: {', '.join(anchor_types)}")
        
        return recommendations
    
    def validate_file_for_upload(self, file_path: str) -> Dict:
        """
        Valida se um arquivo é adequado para upload
        
        Args:
            file_path: Caminho para o arquivo Excel
        
        Returns:
            Dicionário com resultado da validação
        """
        resultado = self.process_excel_melhorado(file_path)
        
        if not resultado['success']:
            return {
                'valid': False,
                'error': resultado['error'],
                'message': f"❌ Erro ao processar arquivo: {resultado['error']}"
            }
        
        if resultado['images_found'] == 0:
            return {
                'valid': False,
                'error': 'no_images',
                'message': f"❌ Arquivo {os.path.basename(file_path)} não contém imagens válidas na coluna H.",
                'suggestion': 'Use um arquivo que tenha imagens inseridas na coluna H com REFs correspondentes.',
                'recommended_files': ['fabrica_com_imagens.xlsx', 'tartaruga.xlsx', 'carrinho.xlsx']
            }
        
        return {
            'valid': True,
            'message': f"✅ Arquivo válido! Pode processar {resultado['images_found']} imagens.",
            'images_count': resultado['images_found'],
            'total_images': resultado['total_images']
        }

def testar_sistema_melhorado():
    """Testa o sistema melhorado com diferentes arquivos"""
    
    print("🚀 TESTE DO SISTEMA MELHORADO")
    print("=" * 60)
    
    sistema = SistemaExcelMelhorado(debug_mode=True)
    
    # Lista de arquivos para testar
    arquivos_teste = [
        "fabrica_com_imagens.xlsx",
        "tartaruga.xlsx",
        "carrinho.xlsx", 
        "nova.xlsx",
        "produto.xlsx"
    ]
    
    resultados = {}
    
    for arquivo in arquivos_teste:
        print(f"\n{'='*60}")
        print(f"📁 Testando: {arquivo}")
        print("=" * 60)
        
        # Valida o arquivo
        validacao = sistema.validate_file_for_upload(arquivo)
        
        if validacao['valid']:
            print(f"✅ {validacao['message']}")
            print(f"   • Imagens válidas: {validacao['images_count']}")
            print(f"   • Total de imagens: {validacao['total_images']}")
        else:
            print(f"❌ {validacao['message']}")
            if 'suggestion' in validacao:
                print(f"💡 {validacao['suggestion']}")
            if 'recommended_files' in validacao:
                print(f"📋 Arquivos recomendados: {', '.join(validacao['recommended_files'])}")
        
        resultados[arquivo] = validacao['valid']
    
    # Resumo final
    print(f"\n{'='*60}")
    print("📋 RESUMO FINAL")
    print("=" * 60)
    
    arquivos_validos = [arquivo for arquivo, valido in resultados.items() if valido]
    arquivos_invalidos = [arquivo for arquivo, valido in resultados.items() if not valido]
    
    if arquivos_validos:
        print("✅ ARQUIVOS VÁLIDOS PARA UPLOAD:")
        for arquivo in arquivos_validos:
            print(f"   • {arquivo}")
    
    if arquivos_invalidos:
        print("❌ ARQUIVOS INVÁLIDOS:")
        for arquivo in arquivos_invalidos:
            print(f"   • {arquivo}")
    
    print(f"\n💡 Use os arquivos válidos para fazer upload!")

if __name__ == "__main__":
    testar_sistema_melhorado()
