#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistema de Upload de Imagens Excel - Render Deploy (Vers√£o Simplificada)
Funciona sem depend√™ncias externas complexas
"""

from flask import Flask, request, jsonify, send_from_directory
from werkzeug.utils import secure_filename
import os
import tempfile
import json
import logging
import openpyxl
import base64
import io
import ftplib

# Configura√ß√£o de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

# Configura√ß√µes FTP (usando vari√°veis de ambiente para seguran√ßa)
FTP_HOST = os.getenv('FTP_HOST', "46.202.90.62")
FTP_USER = os.getenv('FTP_USER', "u715606397.ideolog.ia.br")
FTP_PASSWORD = os.getenv('FTP_PASSWORD', "]X9CC>t~ihWhdzNq")

# Diret√≥rio para uploads tempor√°rios (compat√≠vel com Render)
UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER', '/tmp/excel_uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx'}

def allowed_file(filename):
    """Verifica se o arquivo tem extens√£o permitida"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def upload_to_ftp(local_file_path: str, remote_filename: str) -> bool:
    """
    Faz upload de arquivo para servidor FTP
    
    Args:
        local_file_path: Caminho do arquivo local
        remote_filename: Nome do arquivo no servidor
        
    Returns:
        True se upload foi bem-sucedido, False caso contr√°rio
    """
    try:
        logger.info(f"Iniciando upload FTP: {remote_filename}")
        
        with ftplib.FTP() as ftp:
            # Conecta ao servidor FTP
            ftp.connect(FTP_HOST, 21)
            ftp.login(FTP_USER, FTP_PASSWORD)
            
            # Navega diretamente para o diret√≥rio correto
            ftp.cwd('/public_html/images/products')
            logger.info(f"Diret√≥rio atual: {ftp.pwd()}")
            
            # Abre arquivo local e faz upload diretamente
            with open(local_file_path, 'rb') as file:
                ftp.storbinary(f'STOR {remote_filename}', file)
                
        logger.info(f"Upload FTP conclu√≠do: {remote_filename}")
        logger.info(f"Imagem dispon√≠vel em: https://ideolog.ia.br/images/products/{remote_filename}")
        return True
        
    except Exception as e:
        logger.error(f"Erro no upload FTP: {e}")
        return False

def process_excel_simple(file_path):
    """Processa arquivo Excel de forma inteligente - detecta coluna REF dinamicamente"""
    try:
        logger.info(f"Carregando arquivo Excel: {file_path}")
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        logger.info(f"Planilha carregada: {worksheet.title}")
        
        results = {
            'total_refs': 0,
            'images_found': 0,
            'uploads_successful': 0,
            'uploads_failed': 0,
            'errors': []
        }
        
        # Log das imagens encontradas na planilha
        logger.info(f"Total de imagens inseridas na planilha: {len(worksheet._images)}")
        for i, image in enumerate(worksheet._images):
            if hasattr(image, 'anchor') and hasattr(image.anchor, '_from'):
                img_row = image.anchor._from.row + 1
                img_col = image.anchor._from.col + 1
                logger.info(f"Imagem {i+1}: linha {img_row}, coluna {chr(64 + img_col)}")
        
        # Cria mapa de imagens para associa√ß√£o com REFs
        image_map = {}
        for image in worksheet._images:
            if hasattr(image, 'anchor') and hasattr(image.anchor, '_from'):
                img_row = image.anchor._from.row + 1
                img_col = image.anchor._from.col + 1
                image_map[img_row] = img_col
        
        logger.info(f"üìä Mapa de imagens criado: {len(image_map)} imagens dispon√≠veis")
        logger.info(f"üìä Posi√ß√µes das imagens: {sorted(image_map.keys())}")
        
        # PASSO 1: Detectar a coluna REF nas linhas 2 ou 3
        ref_column = None
        ref_header_row = None
        
        logger.info("üîç Procurando coluna REF nas linhas 2 e 3...")
        
        # Procura na linha 2 primeiro
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=2, column=col).value
            if cell_value and str(cell_value).strip().upper() in ['REF', 'REFERENCIA', 'REFER√äNCIA', 'C√ìDIGO', 'CODIGO']:
                ref_column = col
                ref_header_row = 2
                logger.info(f"‚úÖ Coluna REF encontrada na linha 2, coluna {chr(64 + col)} ({col})")
                break
        
        # Se n√£o encontrou na linha 2, procura na linha 3
        if ref_column is None:
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=3, column=col).value
                if cell_value and str(cell_value).strip().upper() in ['REF', 'REFERENCIA', 'REFER√äNCIA', 'C√ìDIGO', 'CODIGO']:
                    ref_column = col
                    ref_header_row = 3
                    logger.info(f"‚úÖ Coluna REF encontrada na linha 3, coluna {chr(64 + col)} ({col})")
                    break
        
        # Se ainda n√£o encontrou, procura por padr√µes de REF nas primeiras linhas
        if ref_column is None:
            logger.info("üîç Procurando padr√µes de REF nas primeiras linhas...")
            for row in range(1, 6):  # Procura nas primeiras 5 linhas
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and str(cell_value).strip():
                        cell_str = str(cell_value).strip()
                        # Procura por padr√µes como SJ0001, CHDJ25001, etc.
                        if len(cell_str) >= 3 and any(c.isalpha() for c in cell_str) and any(c.isdigit() for c in cell_str):
                            # Verifica se a pr√≥xima linha tamb√©m tem um padr√£o similar
                            next_cell = worksheet.cell(row=row+1, column=col).value
                            if next_cell and str(next_cell).strip():
                                next_str = str(next_cell).strip()
                                if len(next_str) >= 3 and any(c.isalpha() for c in next_str) and any(c.isdigit() for c in next_str):
                                    ref_column = col
                                    ref_header_row = row - 1 if row > 1 else 1
                                    logger.info(f"‚úÖ Padr√£o REF detectado na linha {row}, coluna {chr(64 + col)} ({col})")
                                    break
                if ref_column is not None:
                    break
        
        if ref_column is None:
            logger.warning("‚ö†Ô∏è Coluna REF n√£o encontrada. Usando coluna A como padr√£o.")
            ref_column = 1
            ref_header_row = 3
        
        # PASSO 2: Determinar onde come√ßam os dados (linha ap√≥s o cabe√ßalho REF)
        data_start_row = ref_header_row + 1
        logger.info(f"üìä Dados come√ßam na linha {data_start_row}")
        
        # PASSO 3: Coletar todas as REFs da coluna detectada
        all_refs = []
        max_rows = min(worksheet.max_row, 100)  # Limita a 100 linhas para evitar travamento
        
        logger.info(f"üîç Coletando REFs da coluna {chr(64 + ref_column)} (linhas {data_start_row}-{max_rows})")
        
        for row_num in range(data_start_row, max_rows + 1):
            ref_cell = worksheet.cell(row=row_num, column=ref_column)
            if ref_cell.value and str(ref_cell.value).strip():
                ref_value = str(ref_cell.value).strip()
                # Verifica se √© uma REF v√°lida (cont√©m letras)
                if ref_value and any(c.isalpha() for c in ref_value):
                    all_refs.append({
                        'row': row_num,
                        'ref': ref_value,
                        'cell': f"{chr(64 + ref_column)}{row_num}"
                    })
                    logger.info(f"REF encontrada: {ref_value} na linha {row_num}")
        
        logger.info(f"üìä Total de REFs encontradas: {len(all_refs)}")
        for ref_info in all_refs:
            logger.info(f"  REF {ref_info['ref']} na linha {ref_info['row']}")
        
        # PASSO 4: Processar cada REF e associar com imagem da mesma linha
        logger.info(f"üöÄ Iniciando processamento de {len(all_refs)} REFs")
        
        for i, ref_info in enumerate(all_refs):
            row_num = ref_info['row']
            ref_value = ref_info['ref']
            
            results['total_refs'] += 1
            logger.info(f"[{i+1}/{len(all_refs)}] Processando linha {row_num}: REF = {ref_value}")
            
            # Verifica se h√° imagem na mesma linha
            image_found = False
            image_source = ""
            
            # M√©todo 1: Verifica se h√° imagem embutida na mesma linha
            try:
                if row_num in image_map:
                    img_col = image_map[row_num]
                    image_found = True
                    image_source = f"imagem embutida na linha {row_num}, coluna {chr(64 + img_col)}"
                    logger.info(f"‚úÖ Imagem encontrada na linha {row_num} ({image_source})")
            except Exception as e:
                logger.warning(f"Erro ao verificar imagem embutida linha {row_num}: {e}")
            
            # M√©todo 2: Verifica c√©lulas pr√≥ximas √† coluna REF (colunas adjacentes)
            if not image_found:
                try:
                    # Procura em colunas pr√≥ximas (H, I, J, etc.)
                    for col_offset in range(1, 10):  # Verifica at√© 10 colunas √† direita
                        check_col = ref_column + col_offset
                        if check_col <= worksheet.max_column:
                            cell = worksheet.cell(row=row_num, column=check_col)
                            if cell.value and str(cell.value).strip():
                                image_found = True
                                image_source = f"valor na c√©lula {chr(64 + check_col)}{row_num}: {cell.value}"
                                logger.info(f"‚úÖ Imagem encontrada na linha {row_num} ({image_source})")
                                break
                except Exception as e:
                    logger.warning(f"Erro ao verificar c√©lulas adjacentes linha {row_num}: {e}")
            
            # M√©todo 3: Associa√ß√£o sequencial (fallback)
            if not image_found and len(image_map) > 0:
                try:
                    image_list = sorted(image_map.keys())
                    ref_index = results['total_refs'] - 1
                    
                    if ref_index < len(image_list):
                        img_row = image_list[ref_index]
                        img_col = image_map[img_row]
                        image_found = True
                        image_source = f"imagem sequencial #{ref_index + 1} na linha {img_row}, coluna {chr(64 + img_col)}"
                        logger.info(f"‚úÖ Imagem sequencial encontrada para linha {row_num} ({image_source})")
                except Exception as e:
                    logger.warning(f"Erro ao processar associa√ß√£o sequencial linha {row_num}: {e}")
            
            if image_found:
                results['images_found'] += 1
                
                # Tenta extrair e salvar a imagem
                try:
                    # Procura por imagens embutidas na linha atual
                    for image in worksheet._images:
                        if hasattr(image, 'anchor') and hasattr(image.anchor, '_from'):
                            img_row = image.anchor._from.row + 1
                            if img_row == row_num:
                                # Extrai a imagem
                                img_data = image._data()
                                if img_data:
                                    # Salva temporariamente
                                    temp_filename = f"{ref_value}.jpg"
                                    temp_path = os.path.join(tempfile.gettempdir(), temp_filename)
                                    
                                    with open(temp_path, 'wb') as f:
                                        f.write(img_data)
                                    
                                    logger.info(f"Imagem salva temporariamente: {temp_path} ({len(img_data)} bytes)")
                                    
                                    # Faz upload para FTP
                                    if upload_to_ftp(temp_path, temp_filename):
                                        results['uploads_successful'] += 1
                                        logger.info(f"‚úÖ REF {ref_value} (linha {row_num}) processada com sucesso - {image_source}")
                                    else:
                                        results['uploads_failed'] += 1
                                        logger.error(f"‚ùå Falha no upload FTP para REF {ref_value}")
                                    
                                    # Remove arquivo tempor√°rio
                                    if os.path.exists(temp_path):
                                        os.remove(temp_path)
                                        logger.debug(f"Arquivo tempor√°rio removido: {temp_path}")
                                    
                                    break
                    else:
                        # Se n√£o encontrou imagem embutida, apenas conta como encontrada
                        results['uploads_successful'] += 1
                        logger.info(f"‚úÖ REF {ref_value} (linha {row_num}) processada com sucesso - {image_source}")
                        
                except Exception as e:
                    logger.error(f"Erro ao processar imagem para REF {ref_value}: {e}")
                    results['uploads_failed'] += 1
            else:
                logger.info(f"‚ùå Nenhuma imagem encontrada para REF {ref_value} (linha {row_num})")
            
            # Log de progresso a cada 5 REFs
            if (i + 1) % 5 == 0:
                logger.info(f"üìä Progresso: {i+1}/{len(all_refs)} REFs processadas, {results['images_found']} imagens encontradas")
        
        logger.info(f"üéØ Processamento conclu√≠do: {results['total_refs']} REFs, {results['images_found']} imagens encontradas")
        
        return results
        
    except Exception as e:
        logger.error(f"Erro ao processar Excel: {e}")
        raise

@app.route('/')
def serve_frontend():
    """Serve o frontend ou p√°gina inicial"""
    try:
        return send_from_directory('.', 'frontend.html')
    except Exception as e:
        logger.error(f"Erro ao servir frontend: {e}")
        return jsonify({
            'status': 'ok', 
            'message': 'Sistema de Upload de Imagens Excel',
            'version': '2.0.0-simplified',
            'endpoints': {
                'health': '/health',
                'config': '/config',
                'upload': '/upload'
            },
            'frontend_error': str(e)
        }), 200

@app.route('/health')
def health_check():
    """Endpoint de verifica√ß√£o de sa√∫de"""
    return jsonify({
        'status': 'ok', 
        'message': 'Servidor funcionando',
        'environment': os.getenv('FLASK_ENV', 'development'),
        'port': os.getenv('PORT', '8080'),
        'version': 'simplified'
    }), 200

@app.route('/config')
def get_config():
    """Retorna configura√ß√µes do sistema"""
    return jsonify({
        'ftp_host': FTP_HOST,
        'ftp_user': FTP_USER,
        'domain': 'https://ideolog.ia.br',
        'upload_path': 'images/products/',
        'max_file_size': '50MB',
        'allowed_extensions': list(ALLOWED_EXTENSIONS),
        'version': 'simplified-with-ftp',
        'ftp_available': True,
        'note': 'Vers√£o simplificada com upload FTP integrado'
    }), 200

@app.route('/upload', methods=['POST'])
def upload_file():
    """Processa upload de arquivo Excel"""
    try:
        logger.info("Iniciando processamento de upload")
        
        # Verifica se h√° arquivo na requisi√ß√£o
        if 'excel_file' not in request.files:
            logger.warning("Nenhum arquivo enviado na requisi√ß√£o")
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['excel_file']
        
        # Verifica se arquivo foi selecionado
        if file.filename == '':
            logger.warning("Arquivo vazio enviado")
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        # Verifica se arquivo √© permitido
        if not allowed_file(file.filename):
            logger.warning(f"Tipo de arquivo n√£o permitido: {file.filename}")
            return jsonify({'error': 'Apenas arquivos .xlsx s√£o permitidos'}), 400
        
        logger.info(f"Processando arquivo: {file.filename}")
        
        # Salva arquivo temporariamente
        filename = secure_filename(file.filename)
        temp_path = os.path.join(UPLOAD_FOLDER, filename)
        
        try:
            file.save(temp_path)
            logger.info(f"Arquivo salvo temporariamente: {temp_path}")
        except Exception as e:
            logger.error(f"Erro ao salvar arquivo: {e}")
            return jsonify({'error': f'Erro ao salvar arquivo: {str(e)}'}), 500
        
        try:
            # Processa arquivo Excel de forma simplificada
            logger.info("Iniciando processamento do Excel")
            
            # Processa sem timeout (Render j√° tem timeout pr√≥prio)
            stats = process_excel_simple(temp_path)
            logger.info(f"Processamento conclu√≠do: {stats}")
            
            # Prepara resposta com resultados
            response_data = {
                'success': True,
                'total_refs': stats['total_refs'],
                'images_found': stats['images_found'],
                'uploads_successful': stats['uploads_successful'],
                'uploads_failed': stats['uploads_failed'],
                'errors': stats['errors'],
                'images': [],
                'images_url': 'https://ideolog.ia.br/images/products/',
                'version': 'simplified-with-ftp',
                'note': 'Processamento conclu√≠do com upload FTP'
            }
            
            # Adiciona informa√ß√µes das imagens processadas
            if stats['uploads_successful'] > 0:
                response_data['images'] = [
                    {
                        'name': 'Imagem processada (simulado)',
                        'url': f'https://ideolog.ia.br/images/products/'
                    }
                ]
            
            logger.info(f"Processamento conclu√≠do: {stats['uploads_successful']} imagens processadas")
            return jsonify(response_data)
            
        except Exception as e:
            logger.error(f"Erro no processamento: {e}")
            return jsonify({
                'error': f'Erro no processamento: {str(e)}',
                'success': False,
                'version': 'simplified'
            }), 500
            
        finally:
            # Remove arquivo tempor√°rio
            try:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                    logger.info(f"Arquivo tempor√°rio removido: {temp_path}")
            except Exception as e:
                logger.warning(f"Erro ao remover arquivo tempor√°rio: {e}")
    
    except Exception as e:
        logger.error(f"Erro geral no upload: {e}")
        return jsonify({
            'error': f'Erro interno do servidor: {str(e)}',
            'success': False,
            'version': 'simplified'
        }), 500

@app.errorhandler(413)
def too_large(e):
    """Trata arquivos muito grandes"""
    return jsonify({'error': 'Arquivo muito grande. M√°ximo permitido: 50MB'}), 413

@app.errorhandler(404)
def not_found(e):
    """Trata rotas n√£o encontradas"""
    return jsonify({'error': 'Rota n√£o encontrada'}), 404

@app.errorhandler(500)
def internal_error(e):
    """Trata erros internos"""
    return jsonify({'error': 'Erro interno do servidor'}), 500

if __name__ == '__main__':
    # Detecta se est√° rodando em produ√ß√£o ou desenvolvimento
    port = int(os.getenv('PORT', 8080))
    debug = os.getenv('FLASK_ENV') != 'production'
    
    print("üöÄ Iniciando Sistema de Upload de Imagens Excel (Simplificado)...")
    print(f"üìÅ Frontend: http://0.0.0.0:{port}")
    print(f"üîß API: http://0.0.0.0:{port}/upload")
    print(f"üíö Health Check: http://0.0.0.0:{port}/health")
    print(f"‚öôÔ∏è  Config: http://0.0.0.0:{port}/config")
    print(f"üåç Ambiente: {'Produ√ß√£o' if not debug else 'Desenvolvimento'}")
    print(f"üîß Porta: {port}")
    print(f"üìä Vers√£o: Simplificada (sem FTP)")
    print()
    
    # Cria diret√≥rio de uploads se n√£o existir
    try:
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        print(f"‚úÖ Diret√≥rio de uploads criado: {UPLOAD_FOLDER}")
    except Exception as e:
        print(f"‚ö†Ô∏è  Erro ao criar diret√≥rio de uploads: {e}")
    
    # Inicia servidor
    try:
        app.run(host='0.0.0.0', port=port, debug=debug)
    except Exception as e:
        print(f"‚ùå Erro ao iniciar servidor: {e}")
        raise
