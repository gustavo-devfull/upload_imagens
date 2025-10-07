#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de diagnóstico detalhado para investigar por que as imagens não estão sendo detectadas
"""

import openpyxl
import os
import logging

# Configuração de logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def diagnosticar_imagens_detalhado(file_path):
    """Diagnóstico detalhado das imagens"""
    print(f"🔍 DIAGNÓSTICO DETALHADO: {file_path}")
    print("=" * 60)
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        
        print(f"📊 Planilha: {worksheet.title}")
        print(f"📏 Dimensões: {worksheet.max_row} linhas x {worksheet.max_column} colunas")
        print()
        
        # Analisa todas as imagens
        print("🖼️ ANÁLISE DETALHADA DAS IMAGENS:")
        print(f"Total de imagens encontradas: {len(worksheet._images)}")
        
        if len(worksheet._images) == 0:
            print("❌ NENHUMA IMAGEM ENCONTRADA!")
            print()
            print("🔍 INVESTIGANDO POSSÍVEIS CAUSAS:")
            
            # Verifica se há outras planilhas
            print(f"📋 Total de planilhas: {len(workbook.worksheets)}")
            for i, sheet in enumerate(workbook.worksheets):
                print(f"  Planilha {i+1}: {sheet.title} - {len(sheet._images)} imagens")
            
            # Verifica se há objetos de desenho
            print(f"🎨 Objetos de desenho na planilha ativa: {len(worksheet._drawings)}")
            
            # Verifica se há formas
            print(f"🔷 Formas na planilha ativa: {len(worksheet._shapes)}")
            
        else:
            print("✅ IMAGENS ENCONTRADAS!")
            print()
            
            for i, image in enumerate(worksheet._images):
                print(f"🖼️ Imagem {i+1}:")
                print(f"  Tipo: {type(image)}")
                print(f"  Módulo: {image.__class__.__module__}")
                
                # Analisa o anchor
                anchor = image.anchor
                print(f"  Anchor: {anchor}")
                print(f"  Tipo do anchor: {type(anchor)}")
                
                # Tenta diferentes métodos para obter posição
                row = None
                col = None
                
                print("  🔍 Tentando determinar posição:")
                
                # Método 1: _from.row e _from.col
                if hasattr(anchor, '_from'):
                    print(f"    _from existe: {anchor._from}")
                    print(f"    Tipo _from: {type(anchor._from)}")
                    
                    if hasattr(anchor._from, 'row') and hasattr(anchor._from, 'col'):
                        row = anchor._from.row + 1
                        col = anchor._from.col + 1
                        print(f"    ✅ Posição via _from.row/col: {col}{row}")
                    elif isinstance(anchor._from, tuple) and len(anchor._from) >= 2:
                        row = anchor._from[0] + 1
                        col = anchor._from[1] + 1
                        print(f"    ✅ Posição via _from tuple: {col}{row}")
                    else:
                        print(f"    ❌ _from não tem row/col ou não é tuple")
                
                # Método 2: row e col diretos
                if hasattr(anchor, 'row') and hasattr(anchor, 'col'):
                    row = anchor.row + 1
                    col = anchor.col + 1
                    print(f"    ✅ Posição via anchor.row/col: {col}{row}")
                
                # Método 3: Propriedades alternativas
                for attr in ['row', 'col', 'left', 'top', 'right', 'bottom']:
                    if hasattr(anchor, attr):
                        value = getattr(anchor, attr)
                        print(f"    📍 anchor.{attr}: {value}")
                
                if row and col:
                    col_letter = openpyxl.utils.get_column_letter(col)
                    print(f"  📍 Posição final: {col_letter}{row}")
                    
                    # Verifica se está na coluna H
                    if col == 8:  # H = 8
                        print(f"  ✅ Está na coluna PHOTO (H)")
                    else:
                        print(f"  ❌ Está na coluna {col_letter}, não na PHOTO (H)")
                    
                    # Verifica se está a partir da linha 4
                    if row >= 4:
                        print(f"  ✅ Está a partir da linha 4")
                    else:
                        print(f"  ❌ Está antes da linha 4")
                        
                    # Verifica se há REF correspondente
                    ref_value = worksheet[f'A{row}'].value
                    if ref_value:
                        print(f"  🔍 REF correspondente: {ref_value}")
                    else:
                        print(f"  ❌ Nenhuma REF na linha {row}")
                else:
                    print(f"  ❌ Não foi possível determinar a posição")
                
                print()
        
        # Analisa coluna H especificamente
        print("📸 ANÁLISE DA COLUNA H:")
        for row in range(1, min(10, worksheet.max_row + 1)):
            cell_value = worksheet[f'H{row}'].value
            if cell_value:
                print(f"  H{row}: {str(cell_value)[:50]}")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Erro no diagnóstico: {e}")
        logger.error(f"Erro no diagnóstico detalhado: {e}")

def main():
    """Função principal"""
    # Lista arquivos Excel
    excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
    
    print("📁 Arquivos Excel disponíveis:")
    for i, file in enumerate(excel_files, 1):
        print(f"  {i}. {file}")
    
    print()
    
    # Diagnostica cada arquivo
    for excel_file in excel_files:
        diagnosticar_imagens_detalhado(excel_file)
        print("\n" + "="*60 + "\n")

if __name__ == "__main__":
    main()
