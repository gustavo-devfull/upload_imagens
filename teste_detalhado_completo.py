#!/usr/bin/env python3
"""
🔍 TESTE DETALHADO - Análise completa de arquivo Excel
Verifica imagens em todas as colunas e fornece diagnóstico completo
"""

import openpyxl
import os
from detector_imagens_melhorado import ExcelImageDetector

def teste_detalhado_completo(arquivo):
    """Faz teste detalhado completo do arquivo"""
    
    if not os.path.exists(arquivo):
        print(f"❌ Arquivo {arquivo} não existe!")
        return
    
    print(f"\n🔍 TESTE DETALHADO COMPLETO")
    print(f"📁 Arquivo: {arquivo}")
    print("=" * 70)
    
    try:
        # Carrega o arquivo
        workbook = openpyxl.load_workbook(arquivo)
        worksheet = workbook.active
        
        print(f"📊 Informações da Planilha:")
        print(f"   • Nome: {worksheet.title}")
        print(f"   • Dimensões: {worksheet.max_row} linhas x {worksheet.max_column} colunas")
        
        # Cria detector
        detector = ExcelImageDetector(debug_mode=True)
        
        # Obtém estatísticas detalhadas
        print(f"\n📊 ESTATÍSTICAS COMPLETAS:")
        stats = detector.get_detailed_image_info(worksheet)
        
        print(f"   • Total de imagens: {stats['total_images']}")
        print(f"   • Imagens com REFs: {stats['images_with_refs']}")
        print(f"   • Imagens sem REFs: {stats['images_without_refs']}")
        
        if stats['images_by_column']:
            print(f"   • Imagens por coluna:")
            for col, count in sorted(stats['images_by_column'].items()):
                print(f"     - Coluna {col}: {count} imagens")
        else:
            print(f"   • ❌ Nenhuma imagem encontrada em nenhuma coluna!")
        
        if stats['anchor_types']:
            print(f"   • Tipos de anchor:")
            for anchor_type, count in stats['anchor_types'].items():
                print(f"     - {anchor_type}: {count} imagens")
        
        # Testa detecção em múltiplas colunas
        print(f"\n🎯 TESTE EM MÚLTIPLAS COLUNAS:")
        colunas_teste = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
        
        for coluna in colunas_teste:
            images_col = detector.detect_images_in_worksheet(
                worksheet, 
                target_columns=[coluna], 
                start_row=1,  # Testa desde a linha 1
                ref_column='A'
            )
            
            if images_col:
                print(f"   ✅ Coluna {coluna}: {len(images_col)} imagens")
                for img in images_col:
                    print(f"      - REF: {img['ref']} | Posição: {img['col']}{img['row']}")
            else:
                print(f"   ❌ Coluna {coluna}: 0 imagens")
        
        # Verifica REFs disponíveis
        print(f"\n📋 REFs DISPONÍVEIS:")
        refs_encontradas = []
        for row in range(1, min(worksheet.max_row + 1, 20)):  # Verifica até linha 20
            cell_value = worksheet[f'A{row}'].value
            if cell_value and str(cell_value).strip():
                ref_value = str(cell_value).strip()
                if ref_value.upper() not in ['TOTAL', 'SUBTOTAL', '']:
                    refs_encontradas.append((row, ref_value))
        
        if refs_encontradas:
            print(f"   • {len(refs_encontradas)} REFs encontradas:")
            for row, ref in refs_encontradas[:10]:  # Mostra até 10
                print(f"     - Linha {row}: {ref}")
            if len(refs_encontradas) > 10:
                print(f"     ... e mais {len(refs_encontradas) - 10} REFs")
        else:
            print(f"   • ❌ Nenhuma REF encontrada!")
        
        # Verifica conteúdo das células
        print(f"\n🔍 CONTEÚDO DAS CÉLULAS (primeiras 10 linhas):")
        for row in range(1, min(11, worksheet.max_row + 1)):
            linha_info = []
            for col in range(1, min(8, worksheet.max_column + 1)):  # Colunas A-G
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value:
                    linha_info.append(f"{openpyxl.utils.get_column_letter(col)}{row}:{str(cell_value)[:20]}")
            
            if linha_info:
                print(f"   Linha {row}: {' | '.join(linha_info)}")
        
        workbook.close()
        
        # Resultado final
        print(f"\n{'='*70}")
        print("📋 DIAGNÓSTICO FINAL:")
        print("=" * 70)
        
        if stats['total_images'] == 0:
            print("❌ PROBLEMA IDENTIFICADO:")
            print("   • O arquivo não contém imagens inseridas")
            print("   • Para funcionar, você precisa:")
            print("     1. Abrir o arquivo no Excel")
            print("     2. Inserir imagens na coluna H")
            print("     3. Certificar-se que cada linha tem uma REF na coluna A")
            print("     4. Salvar o arquivo")
            print()
            print("💡 ARQUIVOS QUE FUNCIONAM:")
            print("   • fabrica_com_imagens.xlsx (4 imagens)")
            print("   • tartaruga.xlsx (1 imagem)")
            print("   • carrinho.xlsx (2 imagens)")
        else:
            print("✅ ARQUIVO VÁLIDO!")
            print(f"   • {stats['total_images']} imagens encontradas")
            print(f"   • Pode ser usado para upload")
            
    except Exception as e:
        print(f"❌ Erro ao processar arquivo: {e}")
        import traceback
        traceback.print_exc()

def main():
    """Função principal"""
    
    print("🔍 TESTE DETALHADO COMPLETO")
    print("=" * 70)
    
    # Lista de arquivos para testar
    arquivos_teste = [
        "produtos_Testando_v2_images_1761008370800.xlsx",
        "fabrica_com_imagens.xlsx",
        "tartaruga.xlsx",
        "carrinho.xlsx"
    ]
    
    for arquivo in arquivos_teste:
        if os.path.exists(arquivo):
            teste_detalhado_completo(arquivo)
            print(f"\n{'='*70}\n")

if __name__ == "__main__":
    main()

