#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de diagnóstico para identificar problemas no upload de imagens
"""

import openpyxl
import os
import logging
from upload_ftp_corrigido import FTPImageExtractorCorrigido

# Configuração de logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def diagnosticar_excel(file_path):
    """Diagnostica problemas no arquivo Excel"""
    
    print("🔍 DIAGNÓSTICO DO ARQUIVO EXCEL")
    print("=" * 50)
    
    if not os.path.exists(file_path):
        print(f"❌ Arquivo não encontrado: {file_path}")
        return
    
    try:
        # Carrega o arquivo Excel
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        
        print(f"✅ Arquivo carregado: {file_path}")
        print(f"📊 Planilha ativa: {worksheet.title}")
        print(f"📏 Dimensões: {worksheet.max_row} linhas x {worksheet.max_column} colunas")
        
        # Verifica coluna REF (A)
        print("\n📋 ANÁLISE DA COLUNA REF (A):")
        ref_data = []
        for row in range(4, min(10, worksheet.max_row + 1)):  # Primeiras 6 linhas
            cell_value = worksheet[f'A{row}'].value
            print(f"   Linha {row}: {cell_value}")
            if cell_value and str(cell_value).strip():
                ref_data.append((row, str(cell_value).strip()))
        
        print(f"📊 Total de REFs encontradas: {len(ref_data)}")
        
        # Verifica imagens
        print("\n🖼️ ANÁLISE DAS IMAGENS:")
        print(f"📊 Total de imagens na planilha: {len(worksheet._images)}")
        
        if len(worksheet._images) == 0:
            print("❌ Nenhuma imagem encontrada na planilha!")
            return
        
        # Analisa cada imagem
        for i, image in enumerate(worksheet._images):
            print(f"\n🖼️ Imagem {i+1}:")
            
            # Tenta determinar posição
            anchor = image.anchor
            print(f"   Tipo do anchor: {type(anchor)}")
            
            if hasattr(anchor, '_from'):
                print(f"   Anchor._from: {anchor._from}")
                if hasattr(anchor._from, 'row') and hasattr(anchor._from, 'col'):
                    row = anchor._from.row + 1
                    col = anchor._from.col + 1
                    print(f"   Posição: Linha {row}, Coluna {col}")
                    
                    # Verifica se está na coluna H (PHOTO)
                    if col == 8:  # H = 8
                        print(f"   ✅ Está na coluna H (PHOTO)")
                    else:
                        print(f"   ⚠️ Está na coluna {openpyxl.utils.get_column_letter(col)}")
                else:
                    print(f"   ⚠️ Não foi possível determinar posição")
            else:
                print(f"   ⚠️ Anchor sem atributo _from")
            
            # Verifica dados da imagem
            print(f"   Tipo da imagem: {type(image)}")
            if hasattr(image, '_data'):
                data = image._data
                if callable(data):
                    try:
                        data_bytes = data()
                        print(f"   ✅ Dados da imagem: {len(data_bytes)} bytes")
                    except Exception as e:
                        print(f"   ❌ Erro ao obter dados: {e}")
                else:
                    print(f"   ✅ Dados da imagem: {len(data)} bytes")
            else:
                print(f"   ❌ Sem dados de imagem")
        
        # Testa o extrator
        print("\n🧪 TESTE DO EXTRATOR:")
        extractor = FTPImageExtractorCorrigido("46.202.90.62", "u715606397.ideolog.ia.br", "]X9CC>t~ihWhdzNq")
        
        # Testa extração de REFs
        refs = extractor.get_ref_column_data(worksheet, 4)
        print(f"📋 REFs extraídas pelo extrator: {len(refs)}")
        for row, ref in refs:
            print(f"   Linha {row}: {ref}")
        
        # Testa extração de imagens
        images = extractor.extract_images_from_worksheet(worksheet, 4, 'H')
        print(f"🖼️ Imagens extraídas pelo extrator: {len(images)}")
        for row, img in images:
            print(f"   Linha {row}: {type(img)}")
        
        # Testa associação REF-Imagem
        print("\n🔗 TESTE DE ASSOCIAÇÃO REF-IMAGEM:")
        for ref_row, ref_value in refs:
            image = extractor.find_image_for_ref(ref_row, images)
            if image:
                print(f"   ✅ REF {ref_value} (linha {ref_row}) -> Imagem encontrada")
                
                # Testa salvamento
                try:
                    temp_path = extractor.save_image_to_temp(image, ref_value)
                    if os.path.exists(temp_path):
                        size = os.path.getsize(temp_path)
                        print(f"   ✅ Imagem salva: {temp_path} ({size} bytes)")
                        os.remove(temp_path)  # Remove arquivo de teste
                    else:
                        print(f"   ❌ Arquivo não foi criado: {temp_path}")
                except Exception as e:
                    print(f"   ❌ Erro ao salvar imagem: {e}")
            else:
                print(f"   ❌ REF {ref_value} (linha {ref_row}) -> Nenhuma imagem")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Erro durante diagnóstico: {e}")
        import traceback
        traceback.print_exc()

def main():
    """Função principal"""
    
    # Lista de arquivos para testar
    arquivos_teste = [
        "/Users/gustavo/upload/tartaruga.xlsx",
        "/Users/gustavo/upload/carrinho.xlsx"
    ]
    
    for arquivo in arquivos_teste:
        if os.path.exists(arquivo):
            print(f"\n{'='*60}")
            print(f"TESTANDO: {arquivo}")
            print(f"{'='*60}")
            diagnosticar_excel(arquivo)
            break
    else:
        print("❌ Nenhum arquivo de teste encontrado")
        print("💡 Coloque um arquivo Excel na pasta e execute novamente")

if __name__ == "__main__":
    main()
