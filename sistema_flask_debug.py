#!/usr/bin/env python3
"""
Sistema de Upload de Imagens Excel - Versão Flask Debug para Render
Detecta imagens em planilhas Excel e faz upload para FTP
Versão com logs detalhados para debug
"""

import os
import sys
import json
import tempfile
import traceback
import time
import logging

# Configuração de logging para debug
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Imports condicionais para evitar erros de deploy
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
    logger.info("✅ openpyxl disponível")
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.error("❌ openpyxl não disponível")

try:
    import ftplib
    FTP_AVAILABLE = True
    logger.info("✅ ftplib disponível")
except ImportError:
    FTP_AVAILABLE = False
    logger.error("❌ ftplib não disponível")

try:
    from flask import Flask, request, jsonify, render_template_string
    FLASK_AVAILABLE = True
    logger.info("✅ Flask disponível")
except ImportError:
    FLASK_AVAILABLE = False
    logger.error("❌ Flask não disponível")

# Configurações FTP
FTP_HOST = "46.202.90.62"
FTP_USER = "u715606397.ideolog.ia.br"
FTP_PASS = "]X9CC>t~ihWhdzNq"
FTP_DIR = "public_html/images/products"

# HTML Template
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>🚀 Sistema Upload Excel - Render Debug</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }
        
        .container {
            max-width: 800px;
            margin: 0 auto;
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 40px rgba(0,0,0,0.1);
            overflow: hidden;
        }
        
        .header {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }
        
        .header h1 {
            font-size: 2.5em;
            margin-bottom: 10px;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
        }
        
        .header p {
            font-size: 1.2em;
            opacity: 0.9;
        }
        
        .upload-area {
            padding: 40px;
            text-align: center;
        }
        
        .file-input-wrapper {
            position: relative;
            display: inline-block;
            margin: 20px 0;
        }
        
        .file-input {
            display: none;
        }
        
        .file-input-label {
            display: inline-block;
            padding: 15px 30px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border-radius: 50px;
            cursor: pointer;
            font-size: 1.1em;
            font-weight: bold;
            transition: all 0.3s ease;
            box-shadow: 0 5px 15px rgba(102, 126, 234, 0.4);
        }
        
        .file-input-label:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 25px rgba(102, 126, 234, 0.6);
        }
        
        .file-name {
            margin: 15px 0;
            font-size: 1.1em;
            color: #333;
            font-weight: 500;
        }
        
        .buttons {
            margin: 30px 0;
        }
        
        .btn {
            padding: 12px 25px;
            margin: 0 10px;
            border: none;
            border-radius: 25px;
            font-size: 1em;
            font-weight: bold;
            cursor: pointer;
            transition: all 0.3s ease;
            text-transform: uppercase;
            letter-spacing: 1px;
        }
        
        .btn-primary {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            box-shadow: 0 5px 15px rgba(102, 126, 234, 0.4);
        }
        
        .btn-primary:hover:not(:disabled) {
            transform: translateY(-2px);
            box-shadow: 0 8px 25px rgba(102, 126, 234, 0.6);
        }
        
        .btn-secondary {
            background: #f8f9fa;
            color: #6c757d;
            border: 2px solid #dee2e6;
        }
        
        .btn-secondary:hover:not(:disabled) {
            background: #e9ecef;
            border-color: #adb5bd;
        }
        
        .btn:disabled {
            opacity: 0.6;
            cursor: not-allowed;
            transform: none !important;
        }
        
        .results {
            margin: 30px 0;
            padding: 20px;
            background: #f8f9fa;
            border-radius: 15px;
            display: none;
        }
        
        .success-box {
            background: #d4edda;
            border: 1px solid #c3e6cb;
            color: #155724;
            padding: 20px;
            border-radius: 10px;
            margin-bottom: 20px;
        }
        
        .error-box {
            background: #f8d7da;
            border: 1px solid #f5c6cb;
            color: #721c24;
            padding: 20px;
            border-radius: 10px;
            margin-bottom: 20px;
        }
        
        .warning-box {
            background: #fff3cd;
            border: 1px solid #ffeaa7;
            color: #856404;
            padding: 20px;
            border-radius: 10px;
            margin-bottom: 20px;
        }
        
        .debug-box {
            background: #e2e3e5;
            border: 1px solid #d6d8db;
            color: #383d41;
            padding: 20px;
            border-radius: 10px;
            margin-bottom: 20px;
            font-family: monospace;
            font-size: 0.9em;
        }
        
        .stats {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 15px;
            margin: 20px 0;
        }
        
        .stat-item {
            background: white;
            padding: 15px;
            border-radius: 10px;
            text-align: center;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        
        .stat-number {
            font-size: 2em;
            font-weight: bold;
            color: #667eea;
            margin-bottom: 5px;
        }
        
        .stat-label {
            color: #6c757d;
            font-size: 0.9em;
        }
        
        .footer {
            background: #f8f9fa;
            padding: 20px;
            text-align: center;
            color: #6c757d;
            border-top: 1px solid #dee2e6;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🚀 Sistema Upload Excel</h1>
            <p>Versão Debug para Render - Upload de Imagens Automático</p>
        </div>
        
        <div class="upload-area">
            <h2>📁 Selecione seu arquivo Excel</h2>
            <p>Arquivos suportados: .xlsx com imagens na coluna H</p>
            
            <div class="file-input-wrapper">
                <input type="file" id="excelFile" class="file-input" accept=".xlsx" />
                <label for="excelFile" class="file-input-label">
                    📂 Escolher Arquivo Excel
                </label>
            </div>
            
            <div id="fileName" class="file-name"></div>
            
            <div class="buttons">
                <button id="uploadBtn" class="btn btn-primary" onclick="uploadFile()" disabled>
                    🚀 Fazer Upload
                </button>
                <button id="resetBtn" class="btn btn-secondary" onclick="resetForm()" disabled>
                    🔄 Resetar
                </button>
            </div>
            
            <!-- Resultados -->
            <div id="results" class="results"></div>
        </div>
        
        <div class="footer">
            <p>🔧 Sistema Debug para Render | 🚀 Upload Automático | 📊 Detecção Inteligente</p>
        </div>
    </div>

    <script>
        let selectedFile = null;
        
        document.getElementById('excelFile').addEventListener('change', handleFileSelect);
        
        function handleFileSelect(event) {
            const file = event.target.files[0];
            if (file) {
                selectedFile = file;
                document.getElementById('fileName').textContent = `📄 ${file.name} (${(file.size / 1024 / 1024).toFixed(2)} MB)`;
                document.getElementById('uploadBtn').disabled = false;
                document.getElementById('resetBtn').disabled = false;
            }
        }
        
        function uploadFile() {
            if (!selectedFile) {
                alert('Por favor, selecione um arquivo Excel.');
                return;
            }
            
            const formData = new FormData();
            formData.append('excel_file', selectedFile);
            
            // Desabilita botões
            document.getElementById('uploadBtn').disabled = true;
            document.getElementById('resetBtn').disabled = true;
            
            // Mostra resultados
            document.getElementById('results').style.display = 'block';
            document.getElementById('results').innerHTML = '<div class="debug-box">🔄 Processando arquivo...</div>';
            
            fetch('/upload', {
                method: 'POST',
                body: formData
            })
            .then(response => response.json())
            .then(data => {
                showResults(data);
            })
            .catch(error => {
                document.getElementById('results').innerHTML = `<div class="error-box"><h3>❌ Erro no Upload</h3><p>${error.message}</p></div>`;
                document.getElementById('uploadBtn').disabled = false;
                document.getElementById('resetBtn').disabled = false;
            });
        }
        
        function showResults(data) {
            const resultsDiv = document.getElementById('results');
            
            let html = '';
            
            // Debug info
            if (data.debug_info) {
                html += `<div class="debug-box">
                    <h3>🔍 Informações de Debug</h3>
                    <pre>${data.debug_info}</pre>
                </div>`;
            }
            
            if (data.success && data.uploads_successful > 0) {
                html += `<div class="success-box">
                    <h3>✅ Upload Concluído com Sucesso!</h3>
                    <p>${data.uploads_successful} imagem(ns) processada(s) e enviada(s) para o FTP.</p>
                </div>`;
                
                html += `<div class="stats">
                    <div class="stat-item">
                        <div class="stat-number">${data.total_refs}</div>
                        <div class="stat-label">REFs Processadas</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-number">${data.images_found}</div>
                        <div class="stat-label">Imagens Encontradas</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-number">${data.uploads_successful}</div>
                        <div class="stat-label">Uploads Bem-sucedidos</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-number">${data.uploads_failed}</div>
                        <div class="stat-label">Uploads Falharam</div>
                    </div>
                </div>`;
                
            } else if (data.error) {
                html += `<div class="warning-box">
                    <h3>⚠️ Problema Detectado</h3>
                    <p>${data.error}</p>`;
                if (data.suggestion) {
                    html += `<p><strong>💡 Sugestão:</strong> ${data.suggestion}</p>`;
                }
                html += `</div>`;
                
                html += `<div class="stats">
                    <div class="stat-item">
                        <div class="stat-number">${data.total_refs || 0}</div>
                        <div class="stat-label">REFs Processadas</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-number">${data.images_found || 0}</div>
                        <div class="stat-label">Imagens Encontradas</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-number">${data.uploads_successful || 0}</div>
                        <div class="stat-label">Uploads Bem-sucedidos</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-number">${data.uploads_failed || 0}</div>
                        <div class="stat-label">Uploads Falharam</div>
                    </div>
                </div>`;
            } else {
                html += `<div class="error-box">
                    <h3>❌ Erro no Processamento</h3>
                    <p>Não foi possível processar o arquivo. Verifique se é um arquivo Excel válido.</p>
                </div>`;
            }
            
            resultsDiv.innerHTML = html;
            
            // Reabilita botões
            document.getElementById('uploadBtn').disabled = false;
            document.getElementById('resetBtn').disabled = false;
        }
        
        function resetForm() {
            document.getElementById('excelFile').value = '';
            document.getElementById('fileName').textContent = '';
            document.getElementById('results').style.display = 'none';
            document.getElementById('uploadBtn').disabled = true;
            document.getElementById('resetBtn').disabled = true;
            selectedFile = null;
        }
    </script>
</body>
</html>
"""

def create_flask_app():
    """Cria aplicação Flask com debug"""
    if not FLASK_AVAILABLE:
        logger.error("❌ Flask não disponível")
        return None
    
    app = Flask(__name__)
    
    @app.route('/')
    def index():
        return render_template_string(HTML_TEMPLATE)
    
    @app.route('/health')
    def health():
        try:
            # Teste básico de dependências
            health_status = {
                "status": "ok",
                "message": "Sistema Flask Debug para Railway funcionando",
                "dependencies": {
                    "ftp": FTP_AVAILABLE,
                    "openpyxl": OPENPYXL_AVAILABLE,
                    "flask": FLASK_AVAILABLE,
                    "pil": False
                },
                "timestamp": time.time(),
                "uptime": time.time() - start_time if 'start_time' in globals() else 0
            }
            
            # Verifica se todas as dependências críticas estão disponíveis
            if not OPENPYXL_AVAILABLE or not FLASK_AVAILABLE:
                health_status["status"] = "error"
                health_status["message"] = "Dependências críticas não disponíveis"
            
            return jsonify(health_status), 200
            
        except Exception as e:
            logger.error(f"❌ Erro no health check: {e}")
            return jsonify({
                "status": "error",
                "message": f"Erro no health check: {str(e)}",
                "timestamp": time.time()
            }), 500
    
    @app.route('/config')
    def config():
        return jsonify({
            "ftp_host": FTP_HOST,
            "ftp_user": FTP_USER,
            "ftp_dir": FTP_DIR,
            "max_file_size": "10MB",
            "supported_formats": [".xlsx"],
            "features": [
                "Detecção de imagens com debug",
                "Upload FTP automático",
                "Validação de arquivos",
                "Logs detalhados",
                "Deploy no Railway"
            ]
        })
    
    @app.route('/upload', methods=['POST'])
    def upload():
        debug_info = []
        
        try:
            logger.info("=== INÍCIO DO UPLOAD ===")
            debug_info.append("=== INÍCIO DO UPLOAD ===")
            
            if not FTP_AVAILABLE or not OPENPYXL_AVAILABLE:
                error_msg = 'Dependências não disponíveis no ambiente de deploy'
                logger.error(error_msg)
                debug_info.append(f"❌ {error_msg}")
                return jsonify({
                    'error': error_msg,
                    'total_refs': 0,
                    'images_found': 0,
                    'uploads_successful': 0,
                    'uploads_failed': 1,
                    'debug_info': '\n'.join(debug_info)
                })
            
            # Verifica se o arquivo foi enviado
            if 'excel_file' not in request.files:
                error_msg = 'Arquivo não encontrado'
                logger.error(error_msg)
                debug_info.append(f"❌ {error_msg}")
                return jsonify({
                    'error': error_msg,
                    'total_refs': 0,
                    'images_found': 0,
                    'uploads_successful': 0,
                    'uploads_failed': 1,
                    'debug_info': '\n'.join(debug_info)
                })
            
            file = request.files['excel_file']
            logger.info(f"📁 Arquivo recebido: {file.filename}")
            debug_info.append(f"📁 Arquivo recebido: {file.filename}")
            
            if file.filename == '':
                error_msg = 'Nenhum arquivo selecionado'
                logger.error(error_msg)
                debug_info.append(f"❌ {error_msg}")
                return jsonify({
                    'error': error_msg,
                    'total_refs': 0,
                    'images_found': 0,
                    'uploads_successful': 0,
                    'uploads_failed': 1,
                    'debug_info': '\n'.join(debug_info)
                })
            
            # Salva arquivo temporariamente
            temp_path = os.path.join(tempfile.gettempdir(), file.filename)
            file.save(temp_path)
            logger.info(f"💾 Arquivo salvo em: {temp_path}")
            debug_info.append(f"💾 Arquivo salvo em: {temp_path}")
            
            # Processamento completo com debug
            stats = process_excel_flask_debug(temp_path, debug_info)
            
            # Verifica se o arquivo tem imagens
            if stats['images_found'] == 0:
                error_response = {
                    'error': f'Arquivo {file.filename} não contém imagens na coluna H. Use um arquivo que tenha imagens inseridas.',
                    'total_refs': stats['total_refs'],
                    'images_found': stats['images_found'],
                    'uploads_successful': 0,
                    'uploads_failed': stats['total_refs'],
                    'suggestion': 'Arquivos recomendados: tartaruga.xlsx ou carrinho.xlsx',
                    'debug_info': '\n'.join(debug_info)
                }
                os.remove(temp_path)
                return jsonify(error_response)
            
            # Remove arquivo temporário
            os.remove(temp_path)
            
            # Adiciona debug info ao resultado
            stats['debug_info'] = '\n'.join(debug_info)
            return jsonify(stats)
            
        except Exception as e:
            error_msg = str(e)
            logger.error(f"❌ Erro geral: {error_msg}")
            debug_info.append(f"❌ Erro geral: {error_msg}")
            debug_info.append(f"❌ Traceback: {traceback.format_exc()}")
            
            return jsonify({
                'error': error_msg,
                'total_refs': 0,
                'images_found': 0,
                'uploads_successful': 0,
                'uploads_failed': 1,
                'debug_info': '\n'.join(debug_info)
            })
    
    return app

def process_excel_flask_debug(file_path, debug_info):
    """Processamento completo do Excel para Flask com debug"""
    try:
        logger.info(f"📊 Carregando arquivo Excel: {file_path}")
        debug_info.append(f"📊 Carregando arquivo Excel: {file_path}")
        
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        logger.info(f"📋 Planilha ativa: {worksheet.title}")
        debug_info.append(f"📋 Planilha ativa: {worksheet.title}")
        
        # Processa REFs
        logger.info("🔍 Processando REFs...")
        debug_info.append("🔍 Processando REFs...")
        refs = get_ref_column_data_debug(worksheet, start_row=4, debug_info=debug_info)
        
        # Processa imagens
        logger.info("🖼️ Processando imagens...")
        debug_info.append("🖼️ Processando imagens...")
        images = extract_images_from_worksheet_debug(worksheet, start_row=4, photo_column='H', debug_info=debug_info)
        
        # Faz upload das imagens
        logger.info("🚀 Iniciando uploads...")
        debug_info.append("🚀 Iniciando uploads...")
        uploads_successful = 0
        uploads_failed = 0
        errors = []
        
        for image_data in images:
            try:
                logger.info(f"📤 Uploading imagem para REF: {image_data['ref']}")
                debug_info.append(f"📤 Uploading imagem para REF: {image_data['ref']}")
                
                if upload_image_to_ftp(image_data['image'], image_data['ref']):
                    uploads_successful += 1
                    logger.info(f"✅ Upload bem-sucedido: {image_data['ref']}")
                    debug_info.append(f"✅ Upload bem-sucedido: {image_data['ref']}")
                else:
                    uploads_failed += 1
                    error_msg = f"Falha no upload da imagem para REF: {image_data['ref']}"
                    errors.append(error_msg)
                    logger.error(f"❌ {error_msg}")
                    debug_info.append(f"❌ {error_msg}")
            except Exception as e:
                uploads_failed += 1
                error_msg = f"Erro no upload da imagem para REF {image_data['ref']}: {str(e)}"
                errors.append(error_msg)
                logger.error(f"❌ {error_msg}")
                debug_info.append(f"❌ {error_msg}")
        
        logger.info(f"📊 Processamento concluído: {len(refs)} REFs, {len(images)} imagens, {uploads_successful} uploads")
        debug_info.append(f"📊 Processamento concluído: {len(refs)} REFs, {len(images)} imagens, {uploads_successful} uploads")
        
        return {
            'success': True,
            'total_refs': len(refs),
            'images_found': len(images),
            'uploads_successful': uploads_successful,
            'uploads_failed': uploads_failed,
            'errors': errors,
            'message': 'Processamento Flask Debug Railway concluído com detecção de imagens'
        }
        
    except Exception as e:
        error_msg = str(e)
        logger.error(f"❌ Erro no processamento: {error_msg}")
        debug_info.append(f"❌ Erro no processamento: {error_msg}")
        debug_info.append(f"❌ Traceback: {traceback.format_exc()}")
        
        return {
            'success': False,
            'error': error_msg,
            'total_refs': 0,
            'images_found': 0,
            'uploads_successful': 0,
            'uploads_failed': 1
        }

def get_ref_column_data_debug(worksheet, start_row=4, debug_info=None):
    """Extrai dados da coluna REF com debug"""
    refs = []
    try:
        logger.info(f"🔍 Buscando REFs a partir da linha {start_row}")
        if debug_info:
            debug_info.append(f"🔍 Buscando REFs a partir da linha {start_row}")
        
        for row in range(start_row, worksheet.max_row + 1):
            cell_value = worksheet[f'A{row}'].value
            if cell_value and str(cell_value).strip() and str(cell_value).upper() not in ['TOTAL', 'SUBTOTAL', '']:
                ref_value = str(cell_value).strip()
                refs.append(ref_value)
                logger.debug(f"📋 REF encontrada na linha {row}: {ref_value}")
                if debug_info:
                    debug_info.append(f"📋 REF encontrada na linha {row}: {ref_value}")
        
        logger.info(f"📊 Total de REFs encontradas: {len(refs)}")
        if debug_info:
            debug_info.append(f"📊 Total de REFs encontradas: {len(refs)}")
            
    except Exception as e:
        logger.error(f"❌ Erro ao ler REFs: {e}")
        if debug_info:
            debug_info.append(f"❌ Erro ao ler REFs: {e}")
    return refs

def extract_images_from_worksheet_debug(worksheet, start_row=4, photo_column='H', debug_info=None):
    """Extrai imagens da planilha com debug detalhado"""
    images = []
    try:
        total_images = len(worksheet._images)
        logger.info(f"🖼️ Total de imagens no arquivo: {total_images}")
        if debug_info:
            debug_info.append(f"🖼️ Total de imagens no arquivo: {total_images}")
        
        if total_images == 0:
            logger.warning("⚠️ Nenhuma imagem encontrada no arquivo")
            if debug_info:
                debug_info.append("⚠️ Nenhuma imagem encontrada no arquivo")
            return images
        
        photo_col_num = openpyxl.utils.column_index_from_string(photo_column)
        logger.info(f"📍 Procurando imagens na coluna {photo_column} (índice {photo_col_num})")
        if debug_info:
            debug_info.append(f"📍 Procurando imagens na coluna {photo_column} (índice {photo_col_num})")
        
        for i, image in enumerate(worksheet._images):
            logger.info(f"🖼️ Analisando imagem {i+1}/{total_images}")
            if debug_info:
                debug_info.append(f"🖼️ Analisando imagem {i+1}/{total_images}")
            
            # Verifica se a imagem está na coluna correta e linha válida
            if hasattr(image, 'anchor') and image.anchor:
                anchor = image.anchor
                logger.debug(f"🔗 Anchor tipo: {type(anchor)}")
                if debug_info:
                    debug_info.append(f"🔗 Anchor tipo: {type(anchor)}")
                
                if hasattr(anchor, '_from') and anchor._from:
                    logger.debug(f"📍 Anchor._from: {anchor._from}")
                    if debug_info:
                        debug_info.append(f"📍 Anchor._from: {anchor._from}")
                    
                    col_idx = anchor._from.col
                    row_idx = anchor._from.row + 1  # openpyxl usa índice 0-based
                    
                    logger.debug(f"📍 Posição calculada: linha {row_idx}, coluna {col_idx}")
                    if debug_info:
                        debug_info.append(f"📍 Posição calculada: linha {row_idx}, coluna {col_idx}")
                    
                    # Converte índice da coluna para letra
                    col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                    logger.debug(f"📍 Coluna convertida: {col_letter}")
                    if debug_info:
                        debug_info.append(f"📍 Coluna convertida: {col_letter}")
                    
                    if col_letter == photo_column and row_idx >= start_row:
                        logger.info(f"✅ Imagem encontrada na coluna {photo_column}, linha {row_idx}")
                        if debug_info:
                            debug_info.append(f"✅ Imagem encontrada na coluna {photo_column}, linha {row_idx}")
                        
                        # Busca REF correspondente
                        ref_cell = worksheet[f'A{row_idx}']
                        if ref_cell.value:
                            ref_value = str(ref_cell.value).strip()
                            if ref_value and ref_value.upper() not in ['TOTAL', 'SUBTOTAL', '']:
                                logger.info(f"📋 REF correspondente: {ref_value}")
                                if debug_info:
                                    debug_info.append(f"📋 REF correspondente: {ref_value}")
                                
                                images.append({
                                    'image': image,
                                    'ref': ref_value,
                                    'row': row_idx,
                                    'col': col_letter
                                })
                            else:
                                logger.warning(f"⚠️ REF inválida na linha {row_idx}: {ref_value}")
                                if debug_info:
                                    debug_info.append(f"⚠️ REF inválida na linha {row_idx}: {ref_value}")
                        else:
                            logger.warning(f"⚠️ Sem REF na linha {row_idx}")
                            if debug_info:
                                debug_info.append(f"⚠️ Sem REF na linha {row_idx}")
                    else:
                        logger.debug(f"❌ Imagem não está na coluna {photo_column} ou linha < {start_row}")
                        if debug_info:
                            debug_info.append(f"❌ Imagem não está na coluna {photo_column} ou linha < {start_row}")
                else:
                    logger.warning(f"⚠️ Anchor sem _from")
                    if debug_info:
                        debug_info.append(f"⚠️ Anchor sem _from")
            else:
                logger.warning(f"⚠️ Imagem sem anchor")
                if debug_info:
                    debug_info.append(f"⚠️ Imagem sem anchor")
        
        logger.info(f"📊 Total de imagens válidas encontradas: {len(images)}")
        if debug_info:
            debug_info.append(f"📊 Total de imagens válidas encontradas: {len(images)}")
            
    except Exception as e:
        logger.error(f"❌ Erro ao extrair imagens: {e}")
        if debug_info:
            debug_info.append(f"❌ Erro ao extrair imagens: {e}")
    return images

def upload_image_to_ftp(image, ref_value):
    """Faz upload da imagem para FTP"""
    try:
        logger.info(f"🚀 Iniciando upload FTP para REF: {ref_value}")
        
        # Salva imagem temporariamente
        temp_image_path = save_image_to_temp(image, ref_value)
        logger.info(f"💾 Imagem salva temporariamente: {temp_image_path}")
        
        # Conecta ao FTP
        logger.info(f"🔌 Conectando ao FTP: {FTP_HOST}")
        ftp = ftplib.FTP()
        ftp.connect(FTP_HOST, 21, timeout=300)
        ftp.login(FTP_USER, FTP_PASS)
        logger.info("✅ Conectado ao FTP")
        
        # Cria diretórios se necessário
        try:
            ftp.cwd('public_html')
            logger.info("📁 Diretório public_html encontrado")
        except:
            ftp.mkd('public_html')
            ftp.cwd('public_html')
            logger.info("📁 Diretório public_html criado")
        
        try:
            ftp.cwd('images')
            logger.info("📁 Diretório images encontrado")
        except:
            ftp.mkd('images')
            ftp.cwd('images')
            logger.info("📁 Diretório images criado")
        
        try:
            ftp.cwd('products')
            logger.info("📁 Diretório products encontrado")
        except:
            ftp.mkd('products')
            ftp.cwd('products')
            logger.info("📁 Diretório products criado")
        
        # Upload do arquivo
        logger.info(f"📤 Fazendo upload: {ref_value}.jpg")
        with open(temp_image_path, 'rb') as file:
            ftp.storbinary(f'STOR {ref_value}.jpg', file)
        
        ftp.quit()
        logger.info("✅ Upload concluído com sucesso")
        
        # Remove arquivo temporário
        os.remove(temp_image_path)
        logger.info("🗑️ Arquivo temporário removido")
        
        return True
        
    except Exception as e:
        logger.error(f"❌ Erro no upload FTP: {e}")
        return False

def save_image_to_temp(image, ref_value):
    """Salva imagem temporariamente"""
    try:
        logger.info(f"💾 Salvando imagem temporariamente para REF: {ref_value}")
        
        # Sanitiza nome do arquivo
        safe_ref = "".join(c for c in ref_value if c.isalnum() or c in ('-', '_')).rstrip()
        if not safe_ref:
            safe_ref = f"image_{int(time.time())}"
        
        temp_path = os.path.join(tempfile.gettempdir(), f"{safe_ref}.jpg")
        logger.info(f"📁 Caminho temporário: {temp_path}")
        
        # Tenta diferentes métodos para extrair dados da imagem
        image_data = None
        
        # Método 1: _data() (mais comum)
        if hasattr(image, '_data') and callable(image._data):
            try:
                image_data = image._data()
                logger.info("✅ Dados extraídos via _data()")
            except Exception as e:
                logger.warning(f"⚠️ Erro ao usar _data(): {e}")
        
        # Método 2: ref (BytesIO)
        if not image_data and hasattr(image, 'ref'):
            try:
                image_data = image.ref.read()
                logger.info("✅ Dados extraídos via ref")
            except Exception as e:
                logger.warning(f"⚠️ Erro ao usar ref: {e}")
        
        # Método 3: data
        if not image_data and hasattr(image, 'data'):
            try:
                image_data = image.data
                logger.info("✅ Dados extraídos via data")
            except Exception as e:
                logger.warning(f"⚠️ Erro ao usar data: {e}")
        
        # Método 4: image.save() como último recurso
        if not image_data:
            try:
                image.save(temp_path)
                logger.info("✅ Imagem salva via save()")
                return temp_path
            except Exception as e:
                logger.warning(f"⚠️ Erro ao usar save(): {e}")
        
        # Salva dados extraídos
        if image_data:
            with open(temp_path, 'wb') as f:
                f.write(image_data)
            logger.info(f"✅ Dados salvos em: {temp_path}")
            return temp_path
        
        raise Exception("Não foi possível extrair dados da imagem")
        
    except Exception as e:
        logger.error(f"❌ Erro ao salvar imagem: {e}")
        raise

def start_flask_server():
    """Inicia o servidor Flask com debug"""
    global start_time
    start_time = time.time()
    
    port = int(os.getenv('PORT', 8080))  # Railway usa PORT automático
    
    logger.info("🚀 Iniciando Sistema Flask Debug para Railway...")
    print("🚀 Iniciando Sistema Flask Debug para Railway...")
    print("=" * 50)
    print(f"📁 Frontend: http://localhost:{port}")
    print(f"🔧 API: http://localhost:{port}/upload")
    print(f"⚙️  Config: http://localhost:{port}/config")
    print(f"💚 Health: http://localhost:{port}/health")
    print("=" * 50)
    print("🔧 Dependências:")
    print(f"   • FTP: {'✅' if FTP_AVAILABLE else '❌'}")
    print(f"   • openpyxl: {'✅' if OPENPYXL_AVAILABLE else '❌'}")
    print(f"   • Flask: {'✅' if FLASK_AVAILABLE else '❌'}")
    print(f"   • PIL: ❌ (não usado nesta versão)")
    print("=" * 50)
    
    if not FLASK_AVAILABLE:
        logger.error("❌ Flask não disponível. Instalando...")
        print("❌ Flask não disponível. Instalando...")
        os.system("pip install flask")
        print("🔄 Reinicie o sistema após a instalação.")
        return
    
    app = create_flask_app()
    if app:
        try:
            logger.info(f"✅ Servidor Flask iniciado na porta {port}")
            print(f"✅ Servidor Flask iniciado na porta {port}")
            print("🔄 Aguardando conexões...")
            app.run(host='0.0.0.0', port=port, debug=False, threaded=True)
        except Exception as e:
            logger.error(f"❌ Erro ao iniciar servidor Flask: {e}")
            print(f"❌ Erro ao iniciar servidor Flask: {e}")
            sys.exit(1)

if __name__ == "__main__":
    start_flask_server()
