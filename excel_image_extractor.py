#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistema para ler planilha Excel, extrair imagens associadas à coluna REF
e fazer upload para servidor FTP
"""

import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import ftplib
import os
import io
import pandas as pd
from typing import List, Tuple, Optional
import logging

# Configuração de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ExcelImageExtractor:
    """Classe para extrair imagens de planilhas Excel e fazer upload para FTP"""
    
    def __init__(self, ftp_host: str, ftp_user: str, ftp_password: str):
        """
        Inicializa o extrator com credenciais FTP
        
        Args:
            ftp_host: Host do servidor FTP
            ftp_user: Usuário FTP
            ftp_password: Senha FTP
        """
        self.ftp_host = ftp_host
        self.ftp_user = ftp_user
        self.ftp_password = ftp_password
        
    def read_excel_file(self, file_path: str) -> openpyxl.Workbook:
        """
        Lê arquivo Excel
        
        Args:
            file_path: Caminho para o arquivo Excel
            
        Returns:
            Workbook do openpyxl
        """
        try:
            workbook = openpyxl.load_workbook(file_path)
            logger.info(f"Arquivo Excel carregado: {file_path}")
            return workbook
        except Exception as e:
            logger.error(f"Erro ao carregar arquivo Excel: {e}")
            raise
    
    def get_ref_column_data(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, start_row: int = 4) -> List[Tuple[int, str]]:
        """
        Extrai dados da coluna REF a partir de uma linha específica
        
        Args:
            worksheet: Planilha do Excel
            start_row: Linha inicial para leitura (padrão: 4)
            
        Returns:
            Lista de tuplas (linha, valor_ref)
        """
        ref_data = []
        
        # Procura pela coluna REF (assumindo que está na coluna A) a partir da linha especificada
        for row in worksheet.iter_rows(min_row=start_row, max_row=worksheet.max_row, min_col=1, max_col=1):
            cell = row[0]
            if cell.value and str(cell.value).strip():
                ref_data.append((cell.row, str(cell.value).strip()))
                
        logger.info(f"Encontradas {len(ref_data)} entradas na coluna REF (a partir da linha {start_row})")
        return ref_data
    
    def extract_images_from_worksheet(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, start_row: int = 4, photo_column: str = 'H') -> List[Tuple[int, Image]]:
        """
        Extrai imagens da planilha, focando em imagens sobrepostas à coluna PHOTO
        
        Args:
            worksheet: Planilha do Excel
            start_row: Linha inicial para considerar imagens (padrão: 4)
            photo_column: Coluna onde estão as imagens (padrão: 'H' para PHOTO)
            
        Returns:
            Lista de tuplas (linha, imagem)
        """
        images = []
        
        # Converte coluna para número (A=1, B=2, C=3, etc.)
        photo_col_num = openpyxl.utils.column_index_from_string(photo_column)
        
        # Percorre todas as imagens na planilha
        for image in worksheet._images:
            # Tenta determinar a linha e coluna da imagem baseado na posição
            anchor = image.anchor
            
            # Método 1: Usando _from.row e _from.col
            if hasattr(anchor, '_from') and hasattr(anchor._from, 'row') and hasattr(anchor._from, 'col'):
                row = anchor._from.row + 1  # openpyxl usa índice 0-based
                col = anchor._from.col + 1
            # Método 2: Usando _from como tupla
            elif hasattr(anchor, '_from') and isinstance(anchor._from, tuple) and len(anchor._from) >= 2:
                row = anchor._from[0] + 1
                col = anchor._from[1] + 1
            # Método 3: Usando propriedades alternativas
            elif hasattr(anchor, 'row') and hasattr(anchor, 'col'):
                row = anchor.row + 1
                col = anchor.col + 1
            else:
                # Se não conseguir determinar a posição, pula esta imagem
                logger.warning("Não foi possível determinar a posição de uma imagem")
                continue
            
            # Só considera imagens a partir da linha especificada e na coluna PHOTO
            if row >= start_row and col == photo_col_num:
                images.append((row, image))
                logger.debug(f"Imagem encontrada na linha {row}, coluna {photo_column}")
            
        logger.info(f"Encontradas {len(images)} imagens na coluna {photo_column} (a partir da linha {start_row})")
        return images
    
    def find_image_for_ref(self, ref_row: int, images: List[Tuple[int, Image]], tolerance: int = 2) -> Optional[Image]:
        """
        Encontra a imagem mais próxima de uma linha REF
        
        Args:
            ref_row: Linha da coluna REF
            images: Lista de imagens extraídas
            tolerance: Tolerância para considerar imagem próxima
            
        Returns:
            Imagem mais próxima ou None
        """
        closest_image = None
        min_distance = float('inf')
        
        for img_row, image in images:
            distance = abs(img_row - ref_row)
            if distance <= tolerance and distance < min_distance:
                min_distance = distance
                closest_image = image
                
        return closest_image
    
    def save_image_to_temp(self, image: Image, ref_value: str) -> str:
        """
        Salva imagem em arquivo temporário
        
        Args:
            image: Imagem do openpyxl
            ref_value: Valor da REF para nome do arquivo
            
        Returns:
            Caminho do arquivo temporário
        """
        try:
            # Cria nome do arquivo baseado na REF (com tratamento para REFs inválidas)
            safe_ref = "".join(c for c in ref_value if c.isalnum() or c in ('-', '_')).strip()
            if not safe_ref or safe_ref.upper() in ['TOTAL', 'SUBTOTAL', 'SUM', 'COUNT']:
                safe_ref = f"imagem_{hash(ref_value) % 10000}"
            
            filename = f"{safe_ref}.jpg"
            temp_path = os.path.join("/tmp", filename)
            
            # Salva a imagem - método robusto para diferentes tipos de objeto
            logger.debug(f"Tipo do objeto image: {type(image)}")
            
            # Tenta diferentes métodos de salvamento
            success = False
            
            # Método 1: Usando dados brutos (_data) - MÉTODO PRINCIPAL
            if hasattr(image, '_data') and not success:
                try:
                    data = image._data
                    if callable(data):
                        data = data()  # Se for um método, chama ele
                    
                    if isinstance(data, bytes) and len(data) > 0:
                        with open(temp_path, 'wb') as f:
                            f.write(data)
                            f.flush()
                        success = True
                        logger.debug("Imagem salva usando _data")
                except Exception as e:
                    logger.debug(f"Erro ao salvar usando _data: {e}")
            
            # Método 2: Usando ref (BytesIO) - APENAS SE _data FALHAR
            if hasattr(image, 'ref') and not success:
                try:
                    ref_obj = image.ref
                    if hasattr(ref_obj, 'read'):
                        # É um BytesIO, mas pode estar fechado
                        try:
                            ref_obj.seek(0)  # Volta para o início
                            data = ref_obj.read()
                            if isinstance(data, bytes) and len(data) > 0:
                                with open(temp_path, 'wb') as f:
                                    f.write(data)
                                    f.flush()
                                success = True
                                logger.debug("Imagem salva usando ref (BytesIO)")
                        except ValueError:
                            # BytesIO fechado, pula este método
                            logger.debug("BytesIO fechado, pulando método ref")
                    elif callable(ref_obj):
                        data = ref_obj()
                        if isinstance(data, bytes) and len(data) > 0:
                            with open(temp_path, 'wb') as f:
                                f.write(data)
                                f.flush()
                            success = True
                            logger.debug("Imagem salva usando ref (callable)")
                except Exception as e:
                    logger.debug(f"Erro ao salvar usando ref: {e}")
            
            # Método 3: Usando save direto
            if hasattr(image, 'save') and callable(getattr(image, 'save')) and not success:
                try:
                    image.save(temp_path)
                    success = True
                    logger.debug("Imagem salva usando save direto")
                except Exception as e:
                    logger.debug(f"Erro ao salvar usando save direto: {e}")
            
            # Método 4: Último recurso - tentar obter dados de diferentes formas
            if not success:
                data = None
                for attr in ['_data', 'ref', 'data', 'image']:
                    if hasattr(image, attr):
                        try:
                            potential_data = getattr(image, attr)
                            if callable(potential_data):
                                potential_data = potential_data()
                            if isinstance(potential_data, bytes) and len(potential_data) > 0:
                                data = potential_data
                                break
                        except Exception as e:
                            logger.debug(f"Erro ao obter {attr}: {e}")
                            continue
                
                if data:
                    with open(temp_path, 'wb') as f:
                        f.write(data)
                        f.flush()
                    success = True
                    logger.debug("Imagem salva usando método de fallback")
            
            if not success:
                raise AttributeError("Não foi possível obter dados da imagem")
            
            # Verifica se arquivo foi criado corretamente
            if not os.path.exists(temp_path):
                raise FileNotFoundError(f"Arquivo temporário não foi criado: {temp_path}")
            
            # Verifica se arquivo tem conteúdo
            file_size = os.path.getsize(temp_path)
            if file_size == 0:
                raise ValueError(f"Arquivo temporário está vazio: {temp_path}")
            
            logger.info(f"Imagem salva temporariamente: {temp_path} ({file_size} bytes)")
            return temp_path
            
        except AttributeError as e:
            logger.error(f"Erro de atributo ao salvar imagem: {e}")
            logger.error(f"Tipo do objeto image: {type(image)}")
            logger.error(f"Atributos disponíveis: {[attr for attr in dir(image) if not attr.startswith('_')]}")
            raise
        except Exception as e:
            logger.error(f"Erro ao salvar imagem temporária: {e}")
            raise
    
    def upload_to_ftp(self, local_file_path: str, remote_filename: str) -> bool:
        """
        Faz upload do arquivo para o servidor FTP na pasta images/products/
        
        Args:
            local_file_path: Caminho do arquivo local
            remote_filename: Nome do arquivo no servidor
            
        Returns:
            True se upload foi bem-sucedido
        """
        try:
            # Verifica se arquivo existe e tem conteúdo
            if not os.path.exists(local_file_path):
                logger.error(f"Arquivo local não encontrado: {local_file_path}")
                return False
            
            file_size = os.path.getsize(local_file_path)
            if file_size == 0:
                logger.error(f"Arquivo local está vazio: {local_file_path}")
                return False
            
            logger.debug(f"Fazendo upload de {local_file_path} ({file_size} bytes) como {remote_filename}")
            
            with ftplib.FTP(timeout=300) as ftp:  # Timeout de 5 minutos
                ftp.connect(self.ftp_host, 21)  # Conecta explicitamente na porta 21
                ftp.login(self.ftp_user, self.ftp_password)
                
                # Cria diretórios se não existirem
                try:
                    ftp.mkd('images')
                    logger.debug("Diretório 'images' criado ou já existe")
                except:
                    logger.debug("Diretório 'images' já existe")
                
                try:
                    ftp.cwd('images')
                    ftp.mkd('products')
                    logger.debug("Diretório 'images/products' criado ou já existe")
                except:
                    logger.debug("Diretório 'images/products' já existe")
                
                # Volta para o diretório raiz
                ftp.cwd('/')
                
                # Caminho completo para upload
                remote_path = f"images/products/{remote_filename}"
                
                # Abre arquivo local e faz upload
                with open(local_file_path, 'rb') as file:
                    ftp.storbinary(f'STOR {remote_path}', file)
                    
            logger.info(f"Upload concluído: {remote_path}")
            logger.info(f"Imagem disponível em: https://{self.ftp_host}/{remote_path}")
            return True
            
        except Exception as e:
            logger.error(f"Erro no upload FTP: {e}")
            return False
    
    def process_excel_file(self, excel_file_path: str, start_row: int = 4, photo_column: str = 'H') -> dict:
        """
        Processa arquivo Excel completo: extrai REFs, imagens e faz upload
        
        Args:
            excel_file_path: Caminho para o arquivo Excel
            start_row: Linha inicial para leitura (padrão: 4)
            photo_column: Coluna onde estão as imagens (padrão: 'H' para PHOTO)
            
        Returns:
            Dicionário com estatísticas do processamento
        """
        stats = {
            'total_refs': 0,
            'images_found': 0,
            'uploads_successful': 0,
            'uploads_failed': 0,
            'errors': [],
            'start_row': start_row,
            'photo_column': photo_column
        }
        
        try:
            # Carrega o arquivo Excel
            workbook = self.read_excel_file(excel_file_path)
            worksheet = workbook.active
            
            # Extrai dados da coluna REF a partir da linha especificada
            ref_data = self.get_ref_column_data(worksheet, start_row)
            stats['total_refs'] = len(ref_data)
            
            # Extrai imagens da coluna PHOTO a partir da linha especificada
            images = self.extract_images_from_worksheet(worksheet, start_row, photo_column)
            stats['images_found'] = len(images)
            
            # Processa cada REF
            for ref_row, ref_value in ref_data:
                try:
                    # Encontra imagem associada na mesma linha
                    image = self.find_image_for_ref(ref_row, images)
                    
                    if image:
                        temp_path = None
                        try:
                            # Salva imagem temporariamente
                            temp_path = self.save_image_to_temp(image, ref_value)
                            
                            # Faz upload para FTP
                            remote_filename = f"{ref_value}.jpg"
                            if self.upload_to_ftp(temp_path, remote_filename):
                                stats['uploads_successful'] += 1
                                logger.info(f"✅ REF {ref_value} (linha {ref_row}) processada com sucesso")
                            else:
                                stats['uploads_failed'] += 1
                                
                        finally:
                            # Remove arquivo temporário se foi criado
                            if temp_path and os.path.exists(temp_path):
                                try:
                                    os.remove(temp_path)
                                    logger.debug(f"Arquivo temporário removido: {temp_path}")
                                except Exception as e:
                                    logger.warning(f"Erro ao remover arquivo temporário {temp_path}: {e}")
                    else:
                        logger.warning(f"Nenhuma imagem encontrada para REF {ref_value} (linha {ref_row}) na coluna {photo_column}")
                        
                except Exception as e:
                    error_msg = f"Erro ao processar REF {ref_value}: {e}"
                    logger.error(error_msg)
                    stats['errors'].append(error_msg)
                    stats['uploads_failed'] += 1
            
            workbook.close()
            
        except Exception as e:
            error_msg = f"Erro geral no processamento: {e}"
            logger.error(error_msg)
            stats['errors'].append(error_msg)
            
        return stats


def main():
    """Função principal"""
    
    # Configurações FTP
    FTP_HOST = "gpreto.space"
    FTP_USER = "u715606397.nova"
    FTP_PASSWORD = "]X9CC>t~ihWhdzNq"
    
    # Caminho do arquivo Excel (modifique conforme necessário)
    excel_file = input("Digite o caminho para o arquivo Excel: ").strip()
    
    if not os.path.exists(excel_file):
        print(f"Arquivo não encontrado: {excel_file}")
        return
    
    # Cria instância do extrator
    extractor = ExcelImageExtractor(FTP_HOST, FTP_USER, FTP_PASSWORD)
    
    # Processa o arquivo (lendo a partir da linha 4, coluna PHOTO)
    print("Iniciando processamento...")
    print("📋 Lendo planilha a partir da linha 4...")
    print("🖼️  Procurando imagens na coluna PHOTO (H)...")
    stats = extractor.process_excel_file(excel_file, start_row=4, photo_column='H')
    
    # Exibe resultados
    print("\n" + "="*50)
    print("RESULTADOS DO PROCESSAMENTO")
    print("="*50)
    print(f"Linha inicial: {stats['start_row']}")
    print(f"Coluna das imagens: {stats['photo_column']}")
    print(f"Total de REFs encontradas: {stats['total_refs']}")
    print(f"Imagens encontradas: {stats['images_found']}")
    print(f"Uploads bem-sucedidos: {stats['uploads_successful']}")
    print(f"Uploads falharam: {stats['uploads_failed']}")
    
    if stats['errors']:
        print(f"\nErros encontrados ({len(stats['errors'])}):")
        for error in stats['errors']:
            print(f"  - {error}")
    
    print("\nProcessamento concluído!")


if __name__ == "__main__":
    main()
