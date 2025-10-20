#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Análise detalhada do arquivo nova.xlsx
"""

import openpyxl
import os
import zipfile

def analise_detalhada_nova_xlsx():
    """Análise detalhada do arquivo nova.xlsx"""
    
    print("🔍 Análise Detalhada do arquivo nova.xlsx")
    print("=" * 50)
    
    arquivo = "nova.xlsx"
    
    if not os.path.exists(arquivo):
        print(f"❌ Arquivo {arquivo} não encontrado")
        return
    
    try:
        # Análise 1: Estrutura do arquivo
        print(f"📁 Informações do arquivo:")
        file_size = os.path.getsize(arquivo)
        print(f"   Tamanho: {file_size} bytes ({file_size/1024/1024:.2f} MB)")
        
        # Análise 2: Estrutura interna do Excel
        print(f"\n📊 Estrutura interna do Excel:")
        with zipfile.ZipFile(arquivo, 'r') as zip_file:
            file_list = zip_file.namelist()
            print(f"   Total de arquivos internos: {len(file_list)}")
            
            # Procura por arquivos de imagem
            image_files = [f for f in file_list if any(f.endswith(ext) for ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp'])]
            print(f"   Arquivos de imagem encontrados: {len(image_files)}")
            
            if image_files:
                print(f"   📷 Arquivos de imagem:")
                for img_file in image_files:
                    print(f"      {img_file}")
            
            # Procura por arquivos de desenho
            drawing_files = [f for f in file_list if 'drawing' in f.lower()]
            print(f"   Arquivos de desenho encontrados: {len(drawing_files)}")
            
            if drawing_files:
                print(f"   🎨 Arquivos de desenho:")
                for draw_file in drawing_files:
                    print(f"      {draw_file}")
        
        # Análise 3: Conteúdo das planilhas
        print(f"\n📋 Conteúdo das planilhas:")
        workbook = openpyxl.load_workbook(arquivo, data_only=True, read_only=False)
        
        print(f"   Planilhas disponíveis: {workbook.sheetnames}")
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            print(f"\n   📊 Planilha: {sheet_name}")
            print(f"      Dimensões: {worksheet.max_row} linhas x {worksheet.max_column} colunas")
            
            # Verifica imagens nesta planilha
            total_imagens = len(worksheet._images)
            print(f"      Imagens: {total_imagens}")
            
            if total_imagens > 0:
                print(f"      🖼️ Imagens encontradas:")
                for i, image in enumerate(worksheet._images):
                    print(f"         Imagem {i+1}: {image.width}x{image.height}")
            
            # Verifica conteúdo das primeiras linhas
            print(f"      📝 Primeiras linhas:")
            for row in range(1, min(6, worksheet.max_row + 1)):
                row_data = []
                for col in range(1, min(6, worksheet.max_column + 1)):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        row_data.append(str(cell_value)[:20])  # Limita tamanho
                    else:
                        row_data.append("")
                print(f"         Linha {row}: {row_data}")
        
        workbook.close()
        
        # Análise 4: Verifica se há objetos incorporados
        print(f"\n🔍 Verificando objetos incorporados:")
        try:
            workbook = openpyxl.load_workbook(arquivo, data_only=True, read_only=False)
            worksheet = workbook.active
            
            # Verifica se há objetos incorporados
            if hasattr(worksheet, '_rels'):
                print(f"   Relacionamentos encontrados: {len(worksheet._rels)}")
            
            # Verifica se há objetos OLE
            if hasattr(worksheet, '_ole_objects'):
                print(f"   Objetos OLE encontrados: {len(worksheet._ole_objects)}")
            
            workbook.close()
            
        except Exception as e:
            print(f"   ❌ Erro ao verificar objetos: {e}")
        
        # Análise 5: Verifica se há hiperlinks ou referências
        print(f"\n🔗 Verificando hiperlinks e referências:")
        try:
            workbook = openpyxl.load_workbook(arquivo, data_only=True, read_only=False)
            worksheet = workbook.active
            
            # Verifica hiperlinks
            if hasattr(worksheet, '_hyperlinks'):
                print(f"   Hiperlinks encontrados: {len(worksheet._hyperlinks)}")
                for i, hyperlink in enumerate(worksheet._hyperlinks[:5]):
                    print(f"      {i+1}: {hyperlink.target}")
            
            workbook.close()
            
        except Exception as e:
            print(f"   ❌ Erro ao verificar hiperlinks: {e}")
        
    except Exception as e:
        print(f"❌ Erro na análise: {e}")
        import traceback
        traceback.print_exc()

def verificar_formato_imagens():
    """Verifica se há imagens em formato diferente"""
    
    print(f"\n🔍 Verificando formatos alternativos de imagem:")
    
    try:
        workbook = openpyxl.load_workbook("nova.xlsx", data_only=True, read_only=False)
        worksheet = workbook.active
        
        # Verifica se há células com conteúdo que pode ser imagem
        print(f"📊 Verificando células com conteúdo suspeito:")
        
        for row in range(1, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                
                if cell.value:
                    cell_value = str(cell.value)
                    
                    # Verifica se há referências a imagens
                    if any(keyword in cell_value.lower() for keyword in ['image', 'img', 'photo', 'pic', '.jpg', '.png', '.gif']):
                        col_letter = openpyxl.utils.get_column_letter(col)
                        print(f"   Célula {col_letter}{row}: {cell_value[:50]}...")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Erro: {e}")

if __name__ == "__main__":
    analise_detalhada_nova_xlsx()
    verificar_formato_imagens()
    
    print(f"\n💡 CONCLUSÕES:")
    print(f"1. Se não há imagens, o arquivo não pode ser usado para exportação")
    print(f"2. Se há referências a imagens, pode ser necessário inserir as imagens")
    print(f"3. Use arquivos como tartaruga.xlsx ou carrinho.xlsx que têm imagens")

